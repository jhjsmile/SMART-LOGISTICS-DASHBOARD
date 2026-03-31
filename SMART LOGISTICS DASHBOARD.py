# ═══════════════════════════════════════════════════════════════
#  보안 개선 사항 (v24.2)
# ═══════════════════════════════════════════════════════════════
# 
#  적용 완료:
# 1. Supabase users 테이블에서 사용자 로드 (평문 비밀번호 제거)
# 2. 마스터 비밀번호를 환경변수/Supabase로 이동
# 3. delete_all_rows에 Soft delete + 백업 추가
#
#  추가 작업 필요:
# 4. session_state 메모리 최적화 (페이징)
# 5. 엑셀 파싱 Validation 강화  
# 6. Supabase RLS 정책 설정
# 7. 캐시 무효화 개선
# 8. Google Drive → Supabase Storage 이전
#
#  상세 내용: 보안_취약점_수정_가이드.md 참고
# ═══════════════════════════════════════════════════════════════


import re
import os
import html as html_mod
import threading
import requests
import streamlit as st
import pandas as pd
import plotly.express as px
import hashlib
import calendar
import io
from datetime import datetime, timezone, timedelta, date
from supabase import create_client, Client
from streamlit_autorefresh import st_autorefresh
from modules.realtime import start_realtime, pop_changed_tables, is_running
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload

# ── 모듈 임포트 ──────────────────────────────────────────────────
from modules.utils import (
    get_now_kst_str, _inject_autofocus, notify_new_arrivals,
    _send_telegram, _get_tg_creds, upload_img_to_drive,
    _TELEGRAM_BOT_TOKEN, _TELEGRAM_CHAT_ID, _TG_SENT_CACHE,
)
from modules.auth import (
    hash_pw, verify_pw, get_master_pw_hash,
    _parse_custom_perms, check_perm, _BCRYPT_AVAILABLE,
)
from modules.database import (
    get_supabase, keep_supabase_alive,
    _clear_production_cache, _clear_schedule_cache, _clear_plan_cache,
    _clear_master_cache, _clear_audit_cache, _clear_all_cache,
    _clear_help_request_cache, _clear_access_request_cache,
    clear_cache_for_tables,
    load_realtime_ledger, load_production_history, archive_old_completed,
    insert_row, update_row,
    delete_all_rows, delete_production_row_by_sn,
    load_app_setting, save_app_setting,
    submit_help_request, load_help_requests,
    submit_access_request, load_access_requests, review_access_request,
    insert_audit_log, load_audit_log, load_audit_log_by_date, load_oqc_fail_audit_log,
    delete_all_audit_log, delete_audit_log_row,
    insert_material_serials, load_material_serials,
    load_material_serials_bulk, search_material_by_sn,
    update_material_serial_sn,
    delete_all_material_serial, delete_material_serial_row,
    load_schedule, insert_schedule, update_schedule, delete_schedule,
    delete_all_production_schedule,
    insert_schedule_change_log,
    delete_all_schedule_change_log, delete_schedule_change_log_row,
    load_model_master, upsert_model_master,
    delete_model_from_master, delete_item_from_master,
    delete_all_master_by_group, sync_master_to_session,
    load_production_plan, save_production_plan,
    delete_production_plan_row, delete_all_production_plan,
    insert_plan_change_log, load_plan_change_log,
    delete_all_plan_change_log, delete_plan_change_log_row,
    check_login_lockout, record_login_failure, clear_login_attempts,
    insert_stoppage_log, load_stoppage_log,
    update_stoppage_log, delete_stoppage_log_row,
)
from modules.calendar_view import (
    show_inline_day_panel, render_calendar_weekly, render_calendar_monthly,
    _xp, _rerun, clear_cal,
)
from modules.kpi_dashboard import render_kpi_dashboard


# =================================================================
# 상수 정의
# =================================================================
# 성능 설정
AUTO_REFRESH_INTERVAL_MS = 15000  # Realtime 폴백용 15초 폴링 (변경 감지는 Realtime이 담당)
MAX_FUNCTION_LINES = 200  # 함수 최대 라인 수 가이드

# UI 설정
PDF_VIEWER_HEIGHT_PX = 900
IFRAME_BORDER_RADIUS_PX = 10

# 데이터베이스
DEFAULT_PAGE_SIZE = 100
MAX_QUERY_RESULTS = 1000
MAX_AUDIT_LOG_ROWS = 200
MAX_PLAN_LOG_ROWS  = 500
MAX_LOGIN_ATTEMPTS = 5
LOGIN_LOCKOUT_SECONDS = 300  # 5분

# 파일 업로드
MAX_UPLOAD_SIZE_MB = 200
ALLOWED_FILE_EXTENSIONS = ['.xlsx', '.xls', '.csv']

# 색상 설정
COLOR_SUCCESS = "#28a745"
COLOR_ERROR = "#dc3545"
COLOR_WARNING = "#ffc107"
COLOR_INFO = "#17a2b8"


# =================================================================
# DataFrame 성능 최적화 예시
# =================================================================
# 
#  느린 방법 (iterrows 사용):
#  성능 개선 예시: 위 DataFrame 성능 최적화 섹션 참고
#  벡터화 예: df['result'] = df['col1'] + df['col2']  (현재 코드는 10-100배 느림)
#  벡터화 예: df['result'] = df['col1'] + df['col2']  (현재 코드는 10-100배 느림)
#  벡터화 예: df['result'] = df['col1'] + df['col2']  (현재 코드는 10-100배 느림)
# for idx, row in df.iterrows():
#     df.at[idx, 'result'] = row['a'] + row['b']
#
#  빠른 방법 (벡터화):
# df['result'] = df['a'] + df['b']
#
#  느린 방법 (반복문으로 필터링):
# result = []
# for idx, row in df.iterrows():
#     if row['status'] == 'active':
#         result.append(row)
#
#  빠른 방법 (boolean indexing):
# result = df[df['status'] == 'active']
#
#  느린 방법 (apply with iterrows):
# for idx, row in df.iterrows():
#     process_row(row)
#
#  빠른 방법 (apply 또는 map):
# df.apply(lambda row: process_row(row), axis=1)
# # 또는 단일 컬럼인 경우:
# df['column'].map(process_value)
#
# =================================================================

# =================================================================
# 1. 시스템 전역 설정 (v1.0.0 - 반응형)
# =================================================================
st.set_page_config(
    page_title="생산 통합 관리 시스템 v1.0.0",
    layout="wide",
    initial_sidebar_state="expanded"
)

KST = timezone(timedelta(hours=9))

# ── Supabase Realtime 리스너 시작 (최초 1회) ──────────────────────
try:
    _sb_url = st.secrets["supabase"]["url"]
    _sb_key = st.secrets["supabase"]["key"]
    start_realtime(_sb_url, _sb_key)
except Exception as _rt_err:
    pass  # secrets 없는 환경(로컬 테스트 등)에서는 무시

# ── 폴링 트리거 (Realtime 폴백 / 탭 wake-up 대비) ────────────────
# 마스터 관리 페이지는 자동 새로고침 제외 (입력 중 끊김 방지)
_on_master_page = st.session_state.get("current_line") == "마스터 관리"
_refresh_count = st_autorefresh(
    interval=86400000 if _on_master_page else AUTO_REFRESH_INTERVAL_MS,
    key="pms_auto_refresh"
)
if _refresh_count:
    st.session_state["_last_refresh_count"] = _refresh_count

# ── Realtime 변경 감지 → 해당 테이블 캐시만 초기화 ───────────────
_rt_changed = pop_changed_tables()
if _rt_changed:
    clear_cache_for_tables(_rt_changed)
    # production 변경 시 session_state도 즉시 갱신
    if "production" in _rt_changed and st.session_state.get("login_status"):
        st.session_state.production_db = load_realtime_ledger()
    if "production_schedule" in _rt_changed and st.session_state.get("login_status"):
        st.session_state.schedule_db = load_schedule()

# ── 일별 아카이브: 완료 후 30일 이상 된 레코드를 production_history로 이동 ──
# 로그인한 세션에서 하루 1번만 실행 (UI 차단 없음 — 실패해도 무시)
if (st.session_state.get("login_status")
        and st.session_state.get("_archive_date") != str(date.today())):
    archive_old_completed(days=30)
    st.session_state["_archive_date"] = str(date.today())

PRODUCTION_GROUPS   = ["제조1반", "제조2반", "제조3반"]
CALENDAR_EDIT_ROLES = ["master", "admin", "control_tower", "schedule_manager"]

# ── 사용 설명서 PDF (외부 파일 로드) ────────────────────────────────
# PDF 파일을 소스 코드와 같은 폴더에 위치시키세요: PMS_v1.0.0_사용설명서.pdf
import os as _os, base64 as _b64_loader
_PDF_PATH = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "PMS_v1.0.0_사용설명서.pdf")
def _load_manual_pdf_b64() -> str:
    if _os.path.exists(_PDF_PATH):
        with open(_PDF_PATH, "rb") as _f:
            return _b64_loader.b64encode(_f.read()).decode("utf-8")
    return ""
_MANUAL_PDF_B64 = _load_manual_pdf_b64()

ROLES = {
    "master":           ["생산 지표 관리", "조립 라인", "검사 라인", "포장 라인", "OQC 라인", "생산 현황 리포트", "불량 공정", "수리 현황 리포트", "생산 중단 일지", "마스터 관리", "작업자 매뉴얼", "관리자 매뉴얼", "플로우차트"],
    "admin":            ["생산 지표 관리", "조립 라인", "검사 라인", "포장 라인", "OQC 라인", "생산 현황 리포트", "불량 공정", "수리 현황 리포트", "생산 중단 일지", "마스터 관리", "작업자 매뉴얼", "관리자 매뉴얼", "플로우차트"],
    "control_tower":    ["생산 지표 관리", "생산 현황 리포트", "수리 현황 리포트", "생산 중단 일지", "마스터 관리", "작업자 매뉴얼", "관리자 매뉴얼", "플로우차트"],
    "assembly_team":    ["조립 라인", "생산 중단 일지", "작업자 매뉴얼", "플로우차트"],
    "qc_team":          ["검사 라인", "불량 공정", "생산 중단 일지", "작업자 매뉴얼", "플로우차트"],
    "packing_team":     ["포장 라인", "생산 중단 일지", "작업자 매뉴얼", "플로우차트"],
    "schedule_manager": ["생산 지표 관리", "생산 중단 일지", "작업자 매뉴얼", "플로우차트"],
    "oqc_team":         ["OQC 라인", "생산 중단 일지", "작업자 매뉴얼", "플로우차트"],
}

ROLE_LABELS = {
    "master":        " 마스터 관리자",
    "admin":         " 관리자",
    "control_tower": " 컨트롤 타워",
    "assembly_team": " 조립 담당자",
    "qc_team":       " 검사 담당자",
    "packing_team":  " 포장 담당자",
    "schedule_manager": " 일정 관리자",
    "oqc_team":       " OQC 품질팀",
}

# ── 권한 레벨 (읽기 / 쓰기 / 수정) ─────────────────────────────
PERM_ACTIONS       = ["read", "write", "edit"]
PERM_ACTION_LABELS = {"read": "읽기", "write": "쓰기", "edit": "수정"}


SCHEDULE_COLORS = {
    "조립계획": "#7eb8e8",
    "포장계획": "#7ec8a0",
    "출하계획": "#f0c878",
    "특이사항": "#e8908a",
    "기타":     "#b49fd4",
}
# 일정 등록 폼에서 선택 가능한 계획 카테고리 (특이사항/기타 제외)
PLAN_CATEGORIES = ["조립계획", "포장계획", "출하계획"]
calendar.setfirstweekday(6)  # 일요일 시작

# ── 상태값 상수 ─────────────────────────────────────────────────
WIP_STATES    = ['조립중', '수리 완료(재투입)']
DONE_STATES   = ['검사대기','검사중','OQC대기','OQC중','출하승인','포장대기','포장중','완료']
ACTIVE_STATES = ['조립중','검사대기','검사중','OQC대기','OQC중','출하승인','포장대기','포장중','수리 완료(재투입)','불량 처리 중']

# ── 상태 스타일 (모듈 레벨 상수) ───────────────────────────────
STATUS_STYLE = {
    '검사대기': ('#fff8e1','#f4922a','#f5a623',''),
    '검사중':   ('#e3f2fd','#1565c0','#1976d2',''),
    '포장대기': ('#f3e5f5','#6a1b9a','#8e24aa',''),
    '포장중':   ('#fff3e0','#e07a18','#f4922a',''),
    '완료':     ('#e8f5e9','#1b5e20','#388e3c',''),
    'OQC대기':  ('#fff8e1','#f57f17','#ffa000',''),
    'OQC중':    ('#e8f4fd','#0d47a1','#1565c0',''),
    '출하승인': ('#e0f2f1','#004d40','#00796b',''),
    '조립중':   ('#eceff1','#37474f','#546e7a',''),
    '수리 완료(재투입)': ('#fff3e0','#e07a18','#f4a040',''),
    '불량 처리 중': ('#ffebee','#b71c1c','#c62828',''),
    '교체됨':      ('#e8eaf6','#283593','#3949ab',''),
}

# ── 상태별 배경색 (STATE_CLR / STATE_CLR2 공통 상수, 두 곳에서 재사용) ──
STATUS_BG = {
    '조립중':           '#fff3d4',
    '검사대기':         '#fff3d4',
    '검사중':           '#ddeeff',
    '포장대기':         '#ede0f5',
    '포장중':           '#fde8d4',
    '완료':             '#d4f0e2',
    '불량 처리 중':     '#fde8e7',
    '수리 완료(재투입)':'#e8f4fd',
    'OQC대기':          '#fff3d4',
    'OQC중':            '#ddeeff',
    '출하승인':         '#d4f0e2',
    '부적합(OQC)':      '#fde8e7',
    '교체됨':           '#e8e8f4',
}

from modules.styles import inject_styles, inject_js
inject_styles()
inject_js()

# =================================================================
# 2. 보안 유틸리티
# =================================================================

# =================================================================
# 3. Supabase 연결 및 DB 함수
# =================================================================

# 앱 최초 기동 시 1회만 실행 (autorefresh마다 재실행 방지)
if "supabase_alive_checked" not in st.session_state:
    keep_supabase_alive()
    st.session_state["supabase_alive_checked"] = True

# ── 사이드바 Realtime 연결 상태 표시 ─────────────────────────────
with st.sidebar:
    if is_running():
        st.caption(" 실시간 연결")
    else:
        st.caption(" 폴링 모드")



# =================================================================
# 4. 옵티미스틱 업데이트 헬퍼
# =================================================================
# DB 재조회(load_realtime_ledger) 없이 메모리 즉시 반영
# → update_row + insert_audit_log 2회 DB 쓰기만으로 완결
# → Realtime이 백그라운드에서 전체 캐시 갱신

def _prod_update(sn: str, data: dict) -> None:
    """단일 행 옵티미스틱 업데이트.
    인메모리 업데이트 성공 시 캐시 무효화 불필요 — Realtime 구독이 DB 변경을
    감지해 자동으로 캐시를 초기화함. 시리얼이 메모리에 없을 때만 DB 재조회."""
    _db = st.session_state.get("production_db", pd.DataFrame())
    if _db.empty or "시리얼" not in _db.columns:
        _clear_production_cache()
        st.session_state.production_db = load_realtime_ledger()
        return
    _mask = _db["시리얼"] == sn
    if not _mask.any():
        _clear_production_cache()
        st.session_state.production_db = load_realtime_ledger()
        return
    for col, val in data.items():
        if col in _db.columns:
            st.session_state.production_db.loc[_mask, col] = val


def _prod_bulk_update(updates: list) -> None:
    """다중 행 옵티미스틱 업데이트.
    updates: [{"sn": ..., "data": {...}}, ...]
    인메모리 업데이트 성공 시 캐시 무효화 불필요 — Realtime 구독이 처리."""
    _db = st.session_state.get("production_db", pd.DataFrame())
    if _db.empty or "시리얼" not in _db.columns:
        _clear_production_cache()
        st.session_state.production_db = load_realtime_ledger()
        return
    for item in updates:
        _mask = _db["시리얼"] == item["sn"]
        if not _mask.any():
            continue
        for col, val in item["data"].items():
            if col in _db.columns:
                st.session_state.production_db.loc[_mask, col] = val


def _run_bulk_db_ops(ops: list) -> list:
    """update_row + insert_audit_log를 ThreadPoolExecutor로 병렬 실행.
    ops: [{"sn": str, "data": dict, "audit": dict}, ...]
      audit 키는 insert_audit_log의 kwargs (시리얼, 모델, 반, 이전상태, 이후상태, 작업자, [비고]).
    반환: _prod_bulk_update에 넘길 [{"sn":..., "data":...}, ...] 리스트.
    max_workers=5: Supabase httpx 클라이언트는 동시 요청 안전, 과부하 방지를 위해 제한."""
    if not ops:
        return []
    from concurrent.futures import ThreadPoolExecutor
    def _one(op):
        update_row(op["sn"], op["data"])
        insert_audit_log(**op["audit"])
    with ThreadPoolExecutor(max_workers=min(len(ops), 5)) as _ex:
        list(_ex.map(_one, ops))
    return [{"sn": o["sn"], "data": o["data"]} for o in ops]


# =================================================================
# 5. 캘린더 다이얼로그
# =================================================================


# ── 일정 변경 로그 ───────────────────────────────────────────────
SCH_CHANGE_REASONS = [
    "(선택 필수)",
    "영업 요구량 변경 (주문 취소)",
    "영업 요구량 변경 (물량 증가)",
    "긴급 주문 (Rush Order)",
    "자재 수급 문제 (입고 지연)",
    "자재 수급 문제 (불량 자재)",
    "설비 고장 / 유지보수",
    "인력 변동 (부족/결원)",
    "품질 문제 (불량 발생)",
    "계획 오입력 수정",
    "기타 (직접 입력)",
]


# =================================================================
# st.rerun() 사용 가이드
# =================================================================
#  st.rerun()은 전체 스크립트를 재실행하므로 성능에 영향을 줍니다.
# 권장 패턴:
#   1. st.session_state 업데이트 후 자연스러운 리렌더링 활용
#   2. 조건문으로 불필요한 rerun 방지
#   3. 연속된 rerun() 호출 금지
# =================================================================


# =================================================================
# 5. 세션 상태 초기화
# =================================================================

if 'schedule_db'     not in st.session_state: st.session_state.schedule_db     = load_schedule()
if 'production_plan' not in st.session_state: st.session_state.production_plan = load_production_plan()
if 'production_db'   not in st.session_state: st.session_state.production_db   = pd.DataFrame()
if 'cal_year'         not in st.session_state: st.session_state.cal_year         = datetime.now(KST).year
if 'cal_month'        not in st.session_state: st.session_state.cal_month        = datetime.now(KST).month
if 'cal_month_year'   not in st.session_state: st.session_state.cal_month_year   = datetime.now(KST).year
if 'cal_month_month'  not in st.session_state: st.session_state.cal_month_month  = datetime.now(KST).month
if 'cal_view'        not in st.session_state: st.session_state.cal_view        = "주별"
if 'cal_week_idx'    not in st.session_state: st.session_state.cal_week_idx    = 0
# ── 드롭박스 옵션 기본값 ──────────────────────────────────────────
_DD_DEFAULTS = {
    "dropdown_oqc_defect": [
        "(선택)",
        "외관 불량 (스크래치/변형)", "기능 불량 (동작 이상)",
        "라벨 / 刻印 오류", "포장 불량", "치수 불량",
        "이물질 혼입", "수량 부족", "서류 오류",
        "기타 (직접 입력)",
    ],
    "dropdown_defect_cause": [
        "(선택)",
        "납땜 불량", "부품 미삽", "부품 오삽", "부품 불량",
        "기구 파손", "기구 간섭", "나사 체결 불량",
        "소프트웨어 오류", "펌웨어 오류", "설정 오류",
        "외관 불량 (스크래치)", "외관 불량 (변형)",
        "통신 불량", "전원 불량", "센서 불량",
        "기타 (직접 입력)",
    ],
    "dropdown_repair_action": [
        "(선택)",
        "재납땜", "부품 교체", "부품 재삽입",
        "기구 교체", "나사 재체결",
        "펌웨어 재설치", "소프트웨어 초기화", "설정 재조정",
        "외관 교체", "세척 후 재검사",
        "재검사 후 양품 확인", "폐기 처리",
        "기타 (직접 입력)",
    ],
    "dropdown_mat_name": [],
}
for _dd_key, _dd_default in _DD_DEFAULTS.items():
    if _dd_key not in st.session_state:
        _loaded = load_app_setting(_dd_key)
        st.session_state[_dd_key] = _loaded if _loaded is not None else _dd_default
if 'cal_action'      not in st.session_state: st.session_state.cal_action      = None
if 'cal_action_data' not in st.session_state: st.session_state.cal_action_data = None
if 'cal_action_sub'      not in st.session_state: st.session_state.cal_action_sub      = None
if 'cal_action_sub_data' not in st.session_state: st.session_state.cal_action_sub_data = None

if 'user_db' not in st.session_state:
    #  보안 개선: secrets.toml에서 비밀번호 해시 로드
    # .streamlit/secrets.toml에 다음 내용 추가 필요:
    # [default_users]
    # admin_hash = "your_bcrypt_hash_here"
    # master_hash = "your_bcrypt_hash_here"
    # control_tower_hash = "your_bcrypt_hash_here"
    
    try:
        # Supabase에서 사용자 정보 로드 시도
        sb = get_supabase()
        # users 테이블에서 로드
        result = sb.table("users").select("*").execute()
        if result.data:
            # DB에서 로드 성공
            import json as _json
            _user_db = {}
            for user in result.data:
                _entry = {
                    'pw_hash': user['password_hash'],
                    'role': user['role'],
                }
                if user.get('custom_permissions'):
                    try:
                        _entry['custom_permissions'] = _json.loads(user['custom_permissions'])
                    except (_json.JSONDecodeError, ValueError):
                        pass  # 잘못된 JSON은 무시하고 커스텀 권한 없이 로그인 허용
                _user_db[user['username']] = _entry
            st.session_state.user_db = _user_db
        else:
            # DB에 데이터 없으면 secrets.toml의 해시 사용 (기본 패스워드 사용 금지)
            _admin_hash = st.secrets.get("default_users", {}).get("admin_hash")
            if _admin_hash:
                st.session_state.user_db = {
                    "admin": {"pw_hash": _admin_hash, "role": "admin"},
                }
            else:
                st.session_state.user_db = {}
                st.sidebar.error(" DB에 사용자 없음 & secrets에 admin_hash 미설정: [default_users] admin_hash를 추가하세요.")
    except Exception as e:
        # Supabase 연결 실패 시 임시 계정 (경고 표시)
        # 보안: 평문 비밀번호 하드코딩 제거 → secrets.toml 또는 환경변수에서 해시 로드
        st.sidebar.warning(" Supabase 연결 실패: 로컬 임시 계정으로 실행 중입니다.")
        _fb_users = {}
        try:
            _fb_cfg = st.secrets.get("fallback_users", {})
            if _fb_cfg.get("master_hash"):
                _fb_users["master"] = {"pw_hash": _fb_cfg["master_hash"], "role": "master"}
            if _fb_cfg.get("admin_hash"):
                _fb_users["admin"] = {"pw_hash": _fb_cfg["admin_hash"], "role": "admin"}
        except Exception:
            pass
        if not _fb_users:
            st.sidebar.error(" Supabase 미연결 & 임시 계정 미설정: secrets.toml에 [fallback_users] 섹션을 추가하세요.")
        st.session_state.user_db = _fb_users

if 'group_master_models' not in st.session_state:
    st.session_state.group_master_models = {"제조1반": [], "제조2반": [], "제조3반": []}

if 'group_master_items' not in st.session_state:
    st.session_state.group_master_items = {"제조1반": {}, "제조2반": {}, "제조3반": {}}

# DB model_master → session 동기화 (앱 시작 시 1회)
if 'master_synced' not in st.session_state:
    sync_master_to_session()
    st.session_state.master_synced = True

if 'login_status'            not in st.session_state: st.session_state.login_status            = False
if 'user_role'               not in st.session_state: st.session_state.user_role               = None
if 'user_id'                 not in st.session_state: st.session_state.user_id                 = None
if 'admin_authenticated'     not in st.session_state: st.session_state.admin_authenticated     = False
if 'user_custom_permissions' not in st.session_state: st.session_state.user_custom_permissions = None
if 'user_permission_levels'  not in st.session_state: st.session_state.user_permission_levels  = {}
if 'selected_group'      not in st.session_state: st.session_state.selected_group      = "제조2반"
if 'current_line'        not in st.session_state: st.session_state.current_line        = "현황판"
if 'confirm_target'      not in st.session_state: st.session_state.confirm_target      = None
if 'wait_checked'        not in st.session_state: st.session_state.wait_checked        = {}
if 'wait_scan_cnt'       not in st.session_state: st.session_state.wait_scan_cnt       = {}

# =================================================================
# 6. 로그인
# =================================================================

if 'show_signup' not in st.session_state: st.session_state.show_signup = False

if not st.session_state.login_status:
    _, c_col, _ = st.columns([1, 1.2, 1])
    with c_col:
        st.markdown("<h2 class='centered-title'> 생산 통합 관리 시스템</h2>", unsafe_allow_html=True)

        if not st.session_state.show_signup:
            # ── 로그인 폼 ──────────────────────────────────────────
            with st.form("gate_login"):
                in_id = st.text_input("아이디(ID)")
                in_pw = st.text_input("비밀번호(PW)", type="password")
                if st.form_submit_button("인증 시작", use_container_width=True):
                    # ── 로그인 시도 제한 (Brute-force 방어 — 서버사이드 프로세스 공유) ──
                    _is_locked, _remain_sec = check_login_lockout(in_id)
                    if _is_locked:
                        st.error(f" 로그인 잠금 중입니다. {_remain_sec}초 후 다시 시도하세요.")
                        st.stop()
                    user_info = st.session_state.user_db.get(in_id)
                    if user_info and verify_pw(in_pw, user_info["pw_hash"]):
                        clear_login_attempts(in_id)
                        if _BCRYPT_AVAILABLE and not user_info["pw_hash"].startswith("$2"):
                            new_hash = hash_pw(in_pw)
                            st.session_state.user_db[in_id]["pw_hash"] = new_hash
                            try:
                                get_supabase().table("users").update(
                                    {"password_hash": new_hash}
                                ).eq("username", in_id).execute()
                            except Exception:
                                pass
                        _role = user_info.get("role", "")
                        if _role not in ROLES:
                            st.error(f" 허용되지 않은 계정 권한입니다. (role={_role})")
                            st.stop()
                        st.session_state.user_id       = in_id
                        st.session_state.user_role     = _role
                        _raw_perms = user_info.get("custom_permissions", None)
                        _pages, _levels = _parse_custom_perms(_raw_perms)
                        st.session_state.user_custom_permissions = _pages
                        st.session_state.user_permission_levels  = _levels
                        st.session_state.production_db = load_realtime_ledger()
                        st.session_state.schedule_db   = load_schedule()
                        st.session_state.login_status  = True
                        st.rerun()
                    else:
                        _attempts = record_login_failure(in_id, MAX_LOGIN_ATTEMPTS, LOGIN_LOCKOUT_SECONDS)
                        _remain_attempts = MAX_LOGIN_ATTEMPTS - _attempts
                        if _attempts >= MAX_LOGIN_ATTEMPTS:
                            st.error(f" 로그인 {MAX_LOGIN_ATTEMPTS}회 실패로 {LOGIN_LOCKOUT_SECONDS//60}분 동안 잠금됩니다.")
                        else:
                            st.error(f"로그인 정보가 올바르지 않습니다. (남은 시도: {_remain_attempts}회)")
            st.markdown("<div style='text-align:center;margin-top:8px;'>", unsafe_allow_html=True)
            if st.button(" 계정이 없으신가요? 가입 / 권한 요청", use_container_width=True):
                st.session_state.show_signup = True
                st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)

        else:
            # ── 계정 신청 폼 ───────────────────────────────────────
            st.markdown("<div class='section-title'> 가입 / 권한 요청</div>", unsafe_allow_html=True)
            st.caption("아래 양식을 작성해 주세요. 관리자 승인 후 로그인 가능합니다.")
            _REQ_ROLES = {
                "assembly_team":    " 조립 담당자",
                "qc_team":          " 검사 담당자",
                "oqc_team":         " OQC 품질팀",
                "packing_team":     " 포장 담당자",
                "schedule_manager": " 일정 관리자",
                "control_tower":    " 컨트롤 타워",
            }
            with st.form("access_request_form"):
                rq_name  = st.text_input("이름 *", placeholder="홍길동")
                rq_id    = st.text_input("사용할 아이디 *", placeholder="영문/숫자 조합")
                rq_pw    = st.text_input("비밀번호 *", type="password")
                rq_pw2   = st.text_input("비밀번호 확인 *", type="password")
                rq_dept  = st.selectbox("소속 반", ["전체/공통"] + PRODUCTION_GROUPS)
                rq_role  = st.selectbox("요청 권한",
                                        list(_REQ_ROLES.keys()),
                                        format_func=lambda k: _REQ_ROLES[k])
                rq_reason = st.text_area("신청 사유 *",
                                         placeholder="업무 내용 및 접근이 필요한 이유를 입력해 주세요.",
                                         height=80)
                sb1, sb2 = st.columns(2)
                submitted = sb1.form_submit_button(" 신청 제출", use_container_width=True, type="primary")
                go_back   = sb2.form_submit_button("← 로그인으로", use_container_width=True)

                if go_back:
                    st.session_state.show_signup = False
                    st.rerun()
                if submitted:
                    if not all([rq_name.strip(), rq_id.strip(), rq_pw, rq_reason.strip()]):
                        st.error("이름·아이디·비밀번호·신청 사유는 필수 입력 항목입니다.")
                    elif len(rq_pw) < 8:
                        st.error("비밀번호는 최소 8자 이상이어야 합니다.")
                    elif not re.search(r'[A-Za-z]', rq_pw) or not re.search(r'[0-9]', rq_pw):
                        st.error("비밀번호는 영문자와 숫자를 각각 1자 이상 포함해야 합니다.")
                    elif rq_pw != rq_pw2:
                        st.error("비밀번호가 일치하지 않습니다.")
                    elif rq_id.strip() in st.session_state.user_db:
                        st.error("이미 사용 중인 아이디입니다.")
                    else:
                        _ok = submit_access_request(
                            username=rq_id.strip(), pw_hash=hash_pw(rq_pw),
                            name=rq_name.strip(), department=rq_dept,
                            requested_role=rq_role, reason=rq_reason.strip()
                        )
                        if _ok is True:
                            st.success(" 신청 완료! 관리자 승인 후 로그인 가능합니다.")
                        else:
                            st.error(f"신청 중 오류가 발생했습니다: {_ok}")
    st.stop()

# =================================================================
# 7. 사이드바
# =================================================================


st.sidebar.markdown("###  생산 관리 시스템 v1.0.0")
st.sidebar.markdown(f"**{ROLE_LABELS.get(st.session_state.user_role, '')}**")
st.sidebar.caption(f"ID: {st.session_state.user_id}")
st.sidebar.divider()

#  커스텀 권한이 있으면 우선 적용, 없으면 기본 역할 권한
allowed_nav = st.session_state.get("user_custom_permissions", None)
if allowed_nav is None:
    allowed_nav = ROLES.get(st.session_state.user_role, [])

if st.sidebar.button(" 메인 현황판", use_container_width=True,
    type="primary" if st.session_state.current_line == "현황판" else "secondary"):
    clear_cal()
    st.session_state.production_db = load_realtime_ledger()
    st.session_state.schedule_db   = load_schedule()
    st.session_state.current_line  = "현황판"
    st.rerun()

if "생산 지표 관리" in allowed_nav:
    if st.sidebar.button(" 생산 지표 관리", use_container_width=True,
        type="primary" if st.session_state.current_line == "생산 지표 관리" else "secondary"):
        clear_cal()
        st.session_state.production_db = load_realtime_ledger()
        st.session_state.schedule_db   = load_schedule()
        st.session_state.current_line  = "생산 지표 관리"
        st.rerun()

st.sidebar.divider()

for group in PRODUCTION_GROUPS:
    exp = (st.session_state.selected_group == group
           and st.session_state.current_line in ["조립 라인", "검사 라인", "포장 라인"])
    with st.sidebar.expander(f" {group}", expanded=exp):
        for p in ["조립 라인", "검사 라인", "포장 라인"]:
            # 반별 권한: "라인::반" 신형 또는 "라인" 구형(= 전 반 허용) 모두 인정
            if f"{p}::{group}" in allowed_nav or p in allowed_nav:
                active = (st.session_state.selected_group == group and st.session_state.current_line == p)
                if st.button(f"{p} 현황", key=f"nav_{group}_{p}", use_container_width=True,
                             type="primary" if active else "secondary"):
                    clear_cal()
                    st.session_state.selected_group = group
                    st.session_state.current_line   = p
                    st.session_state.production_db  = load_realtime_ledger()
                    st.rerun()
        if group == PRODUCTION_GROUPS[-1] and "불량 공정" in allowed_nav:
            if st.sidebar.button(" 불량 공정", key="nav_defect", use_container_width=True,
                type="primary" if st.session_state.current_line == "불량 공정" else "secondary"):
                clear_cal()
                st.session_state.current_line  = "불량 공정"
                st.session_state.production_db = load_realtime_ledger()
                st.rerun()
        if group == PRODUCTION_GROUPS[-1] and "OQC 라인" in allowed_nav:
            if st.sidebar.button(" OQC 라인", key="nav_oqc", use_container_width=True,
                type="primary" if st.session_state.current_line == "OQC 라인" else "secondary"):
                clear_cal()
                st.session_state.current_line  = "OQC 라인"
                st.session_state.production_db = load_realtime_ledger()
                st.rerun()

st.sidebar.divider()

for p in ["생산 현황 리포트", "수리 현황 리포트"]:
    if p in allowed_nav:
        if st.sidebar.button(p, key=f"fnav_{p}", use_container_width=True,
            type="primary" if st.session_state.current_line == p else "secondary"):
            clear_cal()
            st.session_state.current_line  = p
            st.session_state.production_db = load_realtime_ledger()
            st.rerun()

if "생산 중단 일지" in allowed_nav:
    if st.sidebar.button(" 생산 중단 일지", key="fnav_stoppage", use_container_width=True,
        type="primary" if st.session_state.current_line == "생산 중단 일지" else "secondary"):
        clear_cal()
        st.session_state.current_line = "생산 중단 일지"
        st.rerun()

if "마스터 관리" in allowed_nav:
    st.sidebar.divider()
    if st.sidebar.button(" 마스터 데이터 관리", use_container_width=True,
        type="primary" if st.session_state.current_line == "마스터 관리" else "secondary"):
        clear_cal()
        st.session_state.current_line = "마스터 관리"
        st.rerun()

if "작업자 매뉴얼" in allowed_nav or "관리자 매뉴얼" in allowed_nav:
    st.sidebar.divider()
if "작업자 매뉴얼" in allowed_nav:
    if st.sidebar.button(" 작업자 매뉴얼", use_container_width=True,
        type="primary" if st.session_state.current_line == "작업자 매뉴얼" else "secondary"):
        clear_cal()
        st.session_state.current_line = "작업자 매뉴얼"
        st.rerun()
if "관리자 매뉴얼" in allowed_nav:
    if st.sidebar.button(" 관리자 매뉴얼", use_container_width=True,
        type="primary" if st.session_state.current_line == "관리자 매뉴얼" else "secondary"):
        clear_cal()
        st.session_state.current_line = "관리자 매뉴얼"
        st.rerun()
if "플로우차트" in allowed_nav:
    if st.sidebar.button(" 시스템 플로우차트", use_container_width=True,
        type="primary" if st.session_state.current_line == "플로우차트" else "secondary"):
        clear_cal()
        st.session_state.current_line = "플로우차트"
        st.rerun()

# ── 관리자 도움 요청 (항상 표시) ─────────────────────────────────
st.sidebar.divider()
_is_admin = st.session_state.user_role in ['admin', 'master']
if _is_admin:
    _help_open = load_help_requests(status="open")
    _help_cnt  = len(_help_open)
    _help_badge = f"  {_help_cnt}건" if _help_cnt > 0 else ""
    _help_exp_label = f"🆘 도움 요청 관리{_help_badge}"
    with st.sidebar.expander(_help_exp_label, expanded=(_help_cnt > 0)):
        if not _help_open.empty:
            for _hr in _help_open.to_dict('records'):
                _hr_req  = html_mod.escape(str(_hr.get('requester', '')))
                _hr_role = html_mod.escape(str(_hr.get('role', '')))
                _hr_page = html_mod.escape(str(_hr.get('page', '')))
                _hr_msg  = html_mod.escape(str(_hr.get('message', '')))
                _hr_time = html_mod.escape(str(_hr.get('created_at', ''))[:16])
                st.markdown(
                    f"**{_hr_req}** ({_hr_role})"
                    f"  \n {_hr_page}  \n {_hr_msg}"
                    f"  \n<small style='color:#aaa;'>{_hr_time}</small>",
                    unsafe_allow_html=True
                )
                _hr_id = _hr.get('id')
                if st.button(" 처리 완료", key=f"hr_close_{_hr_id}", use_container_width=True):
                    try:
                        get_supabase().table("help_requests").update({
                            "status": "closed",
                            "reviewed_by": st.session_state.user_id,
                            "reviewed_at": get_now_kst_str()
                        }).eq("id", _hr_id).execute()
                        _clear_help_request_cache()
                        st.rerun()
                    except Exception:
                        st.error("처리 실패")
                st.markdown("---")
        else:
            st.caption("대기 중인 도움 요청이 없습니다.")
else:
    with st.sidebar.expander("🆘 관리자 도움 요청", expanded=False):
        if 'help_sent' not in st.session_state: st.session_state.help_sent = False
        if st.session_state.help_sent:
            st.success(" 요청이 전송됐습니다.")
            if st.button("다시 요청하기", key="help_reset", use_container_width=True):
                st.session_state.help_sent = False
                st.rerun()
        else:
            with st.form("help_request_form"):
                _help_msg = st.text_area("요청 내용",
                    placeholder="도움이 필요한 내용을 입력해 주세요.", height=80, label_visibility="collapsed")
                if st.form_submit_button(" 요청 전송", use_container_width=True, type="primary"):
                    if _help_msg.strip():
                        _ok, _tg_err = submit_help_request(
                            requester=st.session_state.user_id,
                            role=st.session_state.user_role,
                            page=st.session_state.current_line,
                            message=_help_msg.strip()
                        )
                        if _ok:
                            st.session_state.help_sent = True
                            st.rerun()
                        else:
                            st.error(f"전송 실패: {_tg_err}")
                    else:
                        st.warning("내용을 입력해 주세요.")

if st.sidebar.button(" 로그아웃", use_container_width=True):
    for k in ['login_status','user_role','user_id','admin_authenticated']:
        st.session_state[k] = False if k == 'login_status' else None
    st.rerun()
# =================================================================
# 10. 페이지 렌더링
# =================================================================

curr_g = st.session_state.selected_group
curr_l = st.session_state.current_line

# ── 페이지 이동 감지 → expander 상태 초기화 ──────────────────────────
# 새 메뉴/반으로 이동하면 모든 _xp_* 키를 삭제하여 접힌 상태로 시작.
# 같은 페이지에서 버튼 클릭(rerun) 시에는 상태가 유지된다.
_xp_page_sig_now = f"{curr_l}||{curr_g}"
if st.session_state.get("_xp_page_sig") != _xp_page_sig_now:
    for _xp_k in [k for k in st.session_state if k.startswith("_xp_") and k != "_xp_page_sig"]:
        del st.session_state[_xp_k]
    # 페이지 이동 시 스캔 체크 상태 초기화
    for _ck_k in [k for k in st.session_state if k.startswith((
        "wait_ck_", "hist_ck_", "asm_checked_", "oqc_in_ck_", "oqc_ck_"
    ))]:
        st.session_state[_ck_k] = {}
    st.session_state["_xp_page_sig"] = _xp_page_sig_now
# ─────────────────────────────────────────────────────────────────────

# =================================================================
# 관리자 호출 플로팅 버튼 (항상 표시)
# =================================================================

# 플로팅 버튼 + 모달 주입 — Telegram fetch 방식 (페이지 이동 없음)
_adm_tg_token = _TELEGRAM_BOT_TOKEN
_adm_tg_chat  = _TELEGRAM_CHAT_ID
_adm_caller   = st.session_state.get("user_id", "미상")

import json as _json
_adm_vars_js = (
    f"<script>"
    f"var _ADM_TG_TOKEN={_json.dumps(_adm_tg_token)};"
    f"var _ADM_TG_CHAT={_json.dumps(_adm_tg_chat)};"
    f"var _ADM_CALLER={_json.dumps(_adm_caller)};"
    f"</script>"
)
_adm_main_js = """
<script>
(function(){
    var pdoc = window.parent.document;
    var _old = pdoc.getElementById('adm_float_btn');
    if (_old) _old.remove();
    var _oldO = pdoc.getElementById('adm_modal_overlay');
    if (_oldO) _oldO.remove();
    var _oldS = pdoc.getElementById('adm_float_style');
    if (_oldS) _oldS.remove();

    var s = pdoc.createElement('style');
    s.id = 'adm_float_style';
    s.textContent = [
        '#adm_float_btn{position:fixed;bottom:70px;right:28px;z-index:999990;',
        'background:#e74c3c;color:#fff;border:none;border-radius:50px;',
        'padding:13px 22px;font-size:0.92rem;font-weight:700;cursor:pointer;',
        'box-shadow:0 4px 18px rgba(231,76,60,0.45);',
        'transition:background 0.18s,transform 0.15s;}',
        '#adm_float_btn:hover{background:#c0392b;transform:scale(1.06);}',
        '#adm_modal_overlay{display:none;position:fixed;top:0;left:0;',
        'width:100%;height:100%;background:rgba(0,0,0,0.52);',
        'z-index:999995;align-items:center;justify-content:center;}',
        '#adm_modal_box{background:#fff;border-radius:18px;padding:38px 44px 32px;',
        'min-width:340px;max-width:460px;',
        'box-shadow:0 12px 48px rgba(0,0,0,0.35);text-align:center;}',
        '@keyframes admPopIn{from{transform:scale(0.75);opacity:0}to{transform:scale(1);opacity:1}}',
        '#adm_msg_input{width:100%;padding:10px 14px;border:1.5px solid #ddd;',
        'border-radius:10px;font-size:0.95rem;margin:14px 0 20px;',
        'box-sizing:border-box;outline:none;}',
        '#adm_msg_input:focus{border-color:#e74c3c;}',
        '.adm_btn_row{display:flex;gap:10px;justify-content:center;}',
        '.adm_submit_btn{background:#e74c3c;color:#fff;border:none;border-radius:10px;',
        'padding:11px 30px;font-size:0.97rem;font-weight:700;cursor:pointer;}',
        '.adm_submit_btn:hover{background:#c0392b;}',
        '.adm_cancel_btn{background:#eee;color:#555;border:none;border-radius:10px;',
        'padding:11px 24px;font-size:0.97rem;cursor:pointer;}',
        '.adm_cancel_btn:hover{background:#ddd;}'
    ].join('');
    pdoc.head.appendChild(s);

    var btn = pdoc.createElement('button');
    btn.id = 'adm_float_btn';
    btn.innerHTML = ' 관리자 호출';
    btn.onclick = function() {
        pdoc.getElementById('adm_modal_overlay').style.display = 'flex';
        pdoc.getElementById('adm_msg_input').focus();
    };
    pdoc.body.appendChild(btn);

    window.parent._admClose = function() {
        pdoc.getElementById('adm_modal_overlay').style.display = 'none';
        pdoc.getElementById('adm_msg_input').value = '';
    };

    window.parent._admSubmit = function() {
        var msg = pdoc.getElementById('adm_msg_input').value.trim() || '(사유 없음)';
        window.parent._admClose();
        if (!_ADM_TG_TOKEN || !_ADM_TG_CHAT) {
            alert('텔레그램 설정이 없어 전송할 수 없습니다.');
            return;
        }
        var now = new Date().toLocaleString('ko-KR', {timeZone:'Asia/Seoul'});
        var text = ' 관리자 호출\n작업자: ' + _ADM_CALLER + '\n메시지: ' + msg + '\n시각: ' + now;
        fetch('https://api.telegram.org/bot' + _ADM_TG_TOKEN + '/sendMessage', {
            method: 'POST',
            headers: {'Content-Type': 'application/json'},
            body: JSON.stringify({chat_id: _ADM_TG_CHAT, text: text, parse_mode: 'HTML'})
        }).then(function(r) {
            if (r.ok) {
                var toast = pdoc.createElement('div');
                toast.style.cssText = 'position:fixed;bottom:100px;left:50%;transform:translateX(-50%);'
                    + 'background:#1e8449;color:#fff;padding:12px 28px;border-radius:50px;'
                    + 'font-weight:700;font-size:0.95rem;z-index:9999999;'
                    + 'box-shadow:0 4px 16px rgba(0,0,0,0.3);animation:admPopIn 0.25s ease;';
                toast.textContent = ' 관리자에게 호출을 전송했습니다';
                pdoc.body.appendChild(toast);
                setTimeout(function() { toast.remove(); }, 3000);
            } else {
                alert('전송 실패. 텔레그램 설정을 확인해주세요.');
            }
        }).catch(function() {
            alert('네트워크 오류로 전송에 실패했습니다.');
        });
    };

    var overlay = pdoc.createElement('div');
    overlay.id = 'adm_modal_overlay';
    overlay.innerHTML = '<div id="adm_modal_box">'
        + '<div style="font-size:2.4rem;margin-bottom:6px;"></div>'
        + '<div style="font-size:1.25rem;font-weight:800;color:#1a1a2e;margin-bottom:4px;">관리자 호출</div>'
        + '<div style="font-size:0.85rem;color:#888;margin-bottom:2px;">호출 사유를 입력하세요 (선택)</div>'
        + '<input id="adm_msg_input" type="text" placeholder="예: 라인 이상, 자재 부족, 품질 문제..." maxlength="100"/>'
        + '<div class="adm_btn_row">'
        + '<button id="adm_submit_btn" class="adm_submit_btn"> 호출하기</button>'
        + '<button id="adm_cancel_btn" class="adm_cancel_btn">취소</button>'
        + '</div></div>';
    pdoc.body.appendChild(overlay);

    pdoc.getElementById('adm_submit_btn').addEventListener('click', function() { window.parent._admSubmit(); });
    pdoc.getElementById('adm_cancel_btn').addEventListener('click', function() { window.parent._admClose(); });
    pdoc.getElementById('adm_msg_input').addEventListener('keydown', function(e) {
        if (e.key === 'Enter') window.parent._admSubmit();
        if (e.key === 'Escape') window.parent._admClose();
    });
})();
</script>
"""
st.components.v1.html(_adm_vars_js + _adm_main_js, height=0)

# ── 현황판 ──────────────────────────────────────────────────────
if curr_l == "현황판":
    st.markdown("<h2 class='centered-title'> 생산 통합 현황판</h2>", unsafe_allow_html=True)
    st.caption(f" 마지막 업데이트: {get_now_kst_str()}")

    db_all = st.session_state.production_db

    # 차트 (데이터 있을 때만)
    if not db_all.empty:
        st.markdown("<div class='section-title'> 실시간 차트</div>", unsafe_allow_html=True)
        ch1, ch2, ch3 = st.columns([2.5, 1.5, 1.2])
        with ch1:
            fig = px.bar(
                db_all.groupby(['반','라인']).size().reset_index(name='수량'),
                x='라인', y='수량', color='반', barmode='group',
                title="반별 공정 진행 현황", template="plotly_white",
                text='수량',
                category_orders={"라인": ["조립 라인", "검사 라인", "OQC 라인", "포장 라인"]}
            )
            fig.update_traces(textposition='outside', textfont_size=11)
            fig.update_yaxes(dtick=5)
            fig.update_layout(
                margin=dict(t=50, b=50, l=20, r=30),
                legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5),
                title=dict(font=dict(size=13), x=0, xanchor='left', pad=dict(t=4)),
                uniformtext_minsize=9, uniformtext_mode='hide'
            )
            st.plotly_chart(fig, use_container_width=True, key="dashboard_bar")
        with ch2:
            fig2 = px.pie(
                db_all.groupby('상태').size().reset_index(name='수량'),
                values='수량', names='상태', hole=0.5, title="<b>전체 상태 비중</b>"
            )
            fig2.update_layout(margin=dict(t=40,b=20))
            st.plotly_chart(fig2, use_container_width=True, key="dashboard_pie")
        with ch3:
            fig3 = px.bar(
                db_all.groupby('반').size().reset_index(name='수량'),
                x='반', y='수량', color='반',
                title="<b>반별 총 투입</b>", template="plotly_white",
                text='수량'
            )
            fig3.update_traces(textposition='outside', textfont_size=12)
            fig3.update_yaxes(dtick=5)
            fig3.update_layout(margin=dict(t=50, b=20), showlegend=False)
            st.plotly_chart(fig3, use_container_width=True, key="dashboard_bar2")

    st.divider()

    # 요약 카드 (6열로 넓게)
    st.markdown("<div class='section-title'> 전체 반 생산 요약</div>", unsafe_allow_html=True)
    col1, col2, col3, col4 = st.columns(4)
    total      = len(db_all)
    completed  = len(db_all[(db_all['라인']=='포장 라인')&(db_all['상태']=='완료')])
    in_prog    = len(db_all[db_all['상태'].isin(ACTIVE_STATES)])
    defects    = len(db_all[db_all['상태'].str.contains('불량|부적합',na=False)])
    col1.markdown(f"<div class='stat-box'><div class='stat-label'> 총 투입</div><div class='stat-value'>{total}</div></div>", unsafe_allow_html=True)
    col2.markdown(f"<div class='stat-box'><div class='stat-label'> 최종 완료</div><div class='stat-value'>{completed}</div></div>", unsafe_allow_html=True)
    col3.markdown(f"<div class='stat-box'><div class='stat-label'> 작업 중</div><div class='stat-value'>{in_prog}</div></div>", unsafe_allow_html=True)
    col4.markdown(f"<div class='stat-box'><div class='stat-label'> 불량 이슈</div><div class='stat-value'>{defects}</div></div>", unsafe_allow_html=True)

    # 반별 상세 보기 토글
    _show_ban_detail = st.toggle("반별 상세 보기", value=True, key="dash_ban_detail")
    if _show_ban_detail:
        from datetime import date as _date
        _today    = _date.today()
        _mth_from = _today.strftime('%Y-%m-01')
        _mth_to   = _today.strftime('%Y-%m-%d')
        _hist = load_production_history(_mth_from, _mth_to)

        # 이번달 조립 계획 수량 (달성률 게이지용)
        _sch_db  = st.session_state.get('schedule_db', pd.DataFrame())
        _mth_str = _today.strftime('%Y-%m')
        if not _sch_db.empty:
            _mth_sch = _sch_db[
                _sch_db['날짜'].astype(str).str.startswith(_mth_str) &
                (_sch_db['카테고리'] == '조립계획')
            ]
        else:
            _mth_sch = pd.DataFrame()

        _BAN_CLR_CARD = {"제조1반": "#2471a3", "제조2반": "#1e8449", "제조3반": "#6c3483"}
        _ban_cols = st.columns(len(PRODUCTION_GROUPS))
        for _bi, _g in enumerate(PRODUCTION_GROUPS):
            _h = _hist[_hist['반'] == _g] if not _hist.empty else pd.DataFrame()
            _d = db_all[db_all['반'] == _g]

            _총투입   = len(_h)
            _누적완료 = int(len(_h[(_h['라인'] == '포장 라인') & (_h['상태'] == '완료')])) if not _h.empty else 0
            _진행중   = len(_d[_d['상태'].isin(ACTIVE_STATES)])
            _불량     = len(_d[_d['상태'].str.contains('불량|부적합', na=False)])

            # 이번달 달성률
            _bp_rows  = _mth_sch[_mth_sch['반'] == _g] if not _mth_sch.empty else pd.DataFrame()
            _ban_plan = int(pd.to_numeric(_bp_rows['조립수'], errors='coerce').fillna(0).sum()) if not _bp_rows.empty else 0
            _달성률   = round(_누적완료 / _ban_plan * 100, 1) if _ban_plan > 0 else 0
            _gauge_w  = min(int(_달성률), 100)
            _gauge_c  = "#1e8449" if _달성률 >= 100 else "#d68910" if _달성률 >= 70 else "#c0392b"
            _pct_txt  = f"{_달성률}%" if _ban_plan > 0 else "계획 미등록"
            _clr      = _BAN_CLR_CARD.get(_g, "#888")

            with _ban_cols[_bi]:
                st.markdown(
                    f"<div style='background:#fffdf8;border:1.5px solid {_clr}44;border-radius:14px;padding:14px 16px;box-sizing:border-box;'>"
                    f"<div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:6px;'>"
                    f"<div style='font-size:clamp(0.85rem,1.2vw,1rem);font-weight:bold;color:{_clr};'>{_g}</div>"
                    f"<div style='font-size:clamp(1rem,1.8vw,1.35rem);font-weight:bold;color:{_gauge_c};'>{_pct_txt}</div>"
                    f"</div>"
                    f"<div style='background:#e8e2d8;border-radius:99px;height:6px;margin-bottom:10px;overflow:hidden;'>"
                    f"<div style='background:{_gauge_c};width:{_gauge_w}%;height:100%;border-radius:99px;'></div>"
                    f"</div>"
                    f"<div style='display:grid;grid-template-columns:1fr 1fr;gap:6px;'>"
                    f"<div style='background:#f0f4f8;border-radius:8px;padding:8px 6px;text-align:center;'>"
                    f"<div style='font-size:0.62rem;color:#8a7f72;font-weight:bold;margin-bottom:3px;'>이번달 투입</div>"
                    f"<div style='font-size:clamp(1.1rem,2vw,1.6rem);color:#5a96c8;font-weight:bold;'>{_총투입}</div></div>"
                    f"<div style='background:#f0f4f8;border-radius:8px;padding:8px 6px;text-align:center;'>"
                    f"<div style='font-size:0.62rem;color:#8a7f72;font-weight:bold;margin-bottom:3px;'>누적 완료</div>"
                    f"<div style='font-size:clamp(1.1rem,2vw,1.6rem);color:#4da875;font-weight:bold;'>{_누적완료}</div></div>"
                    f"<div style='background:#f0f4f8;border-radius:8px;padding:8px 6px;text-align:center;'>"
                    f"<div style='font-size:0.62rem;color:#8a7f72;font-weight:bold;margin-bottom:3px;'>진행 중</div>"
                    f"<div style='font-size:clamp(1.1rem,2vw,1.6rem);color:#e8a838;font-weight:bold;'>{_진행중}</div></div>"
                    f"<div style='background:{'#fde8e7' if _불량 > 0 else '#f0f4f8'};border-radius:8px;padding:8px 6px;text-align:center;'>"
                    f"<div style='font-size:0.62rem;color:#8a7f72;font-weight:bold;margin-bottom:3px;'>불량·부적합</div>"
                    f"<div style='font-size:clamp(1.1rem,2vw,1.6rem);color:{'#c8605a' if _불량 > 0 else '#aaa'};font-weight:bold;'>{_불량}</div></div>"
                    f"</div></div>",
                    unsafe_allow_html=True
                )

    st.divider()

    # ── 모델별 생산 현황 (혼류 대응) ─────────────────────────────────
    if not db_all.empty:
        st.markdown("<div class='section-title'> 모델별 생산 현황</div>", unsafe_allow_html=True)
        _m_total  = db_all.groupby(['반', '모델']).size().rename('투입')
        _m_active = db_all[db_all['상태'].isin(ACTIVE_STATES)].groupby(['반', '모델']).size().rename('진행중')
        _m_done   = db_all[(db_all['라인'] == '포장 라인') & (db_all['상태'] == '완료')].groupby(['반', '모델']).size().rename('완료')
        _m_ng     = db_all[db_all['상태'].str.contains('불량|부적합', na=False)].groupby(['반', '모델']).size().rename('불량')
        _mdl_df   = pd.concat([_m_total, _m_active, _m_done, _m_ng], axis=1).fillna(0).astype(int).reset_index()
        _mdl_df   = _mdl_df[_mdl_df['투입'] > 0].sort_values(['반', '투입'], ascending=[True, False]).reset_index(drop=True)

        if not _mdl_df.empty:
            _BAN_HDR = {"제조1반": "#2471a3", "제조2반": "#1e8449", "제조3반": "#6c3483"}
            _tbl = (
                "<div style='overflow-x:auto;'>"
                "<table style='width:100%;border-collapse:collapse;font-size:0.82rem;'>"
                "<tr style='background:#1B3A5C;color:#fff;font-weight:700;'>"
                "<th style='padding:7px 10px;text-align:left;'>반</th>"
                "<th style='padding:7px 10px;text-align:left;'>모델</th>"
                "<th style='padding:7px 10px;text-align:center;'>투입</th>"
                "<th style='padding:7px 10px;text-align:center;'>진행중</th>"
                "<th style='padding:7px 10px;text-align:center;'>완료</th>"
                "<th style='padding:7px 10px;text-align:center;'>불량</th>"
                "<th style='padding:7px 10px;text-align:left;min-width:120px;'>진행률</th>"
                "</tr>"
            )
            for _ri, _mr in _mdl_df.iterrows():
                _bg = "#f8f9fa" if _ri % 2 == 0 else "#ffffff"
                _hc = _BAN_HDR.get(_mr['반'], "#888")
                _ng_bg = "background:#fde8e7;" if _mr['불량'] > 0 else ""
                _ng_cl = "color:#c0392b;font-weight:bold;" if _mr['불량'] > 0 else "color:#aaa;"
                _pct  = round(_mr['완료'] / max(_mr['투입'], 1) * 100)
                _pc   = "#1e8449" if _pct >= 80 else "#d68910" if _pct >= 50 else "#c0392b"
                _tbl += (
                    f"<tr style='background:{_bg};'>"
                    f"<td style='padding:6px 10px;'><span style='background:{_hc}22;color:{_hc};font-weight:700;"
                    f"padding:2px 8px;border-radius:5px;font-size:0.78rem;'>{_mr['반'][:3]}</span></td>"
                    f"<td style='padding:6px 10px;font-weight:600;'>{_mr['모델']}</td>"
                    f"<td style='padding:6px 10px;text-align:center;'>{_mr['투입']}</td>"
                    f"<td style='padding:6px 10px;text-align:center;color:#d68910;font-weight:600;'>{_mr['진행중']}</td>"
                    f"<td style='padding:6px 10px;text-align:center;color:#1e8449;font-weight:600;'>{_mr['완료']}</td>"
                    f"<td style='padding:6px 10px;text-align:center;{_ng_bg}'><span style='{_ng_cl}'>{_mr['불량']}</span></td>"
                    f"<td style='padding:6px 12px;'>"
                    f"<div style='display:flex;align-items:center;gap:6px;'>"
                    f"<div style='flex:1;background:#e8e2d8;border-radius:99px;height:5px;overflow:hidden;'>"
                    f"<div style='background:{_pc};width:{_pct}%;height:100%;border-radius:99px;'></div></div>"
                    f"<span style='font-size:0.75rem;color:{_pc};font-weight:700;min-width:32px;'>{_pct}%</span>"
                    f"</div></td>"
                    f"</tr>"
                )
            _tbl += "</table></div>"
            st.markdown(_tbl, unsafe_allow_html=True)

    st.divider()

    # 캘린더 (월별)
    st.markdown("<div class='section-title'> 생산 일정 캘린더</div>", unsafe_allow_html=True)
    if st.session_state.user_role in CALENDAR_EDIT_ROLES and check_perm("생산 지표 관리", "edit"):
        st.caption(" 날짜 버튼 클릭 → 일정 상세/추가/수정/삭제")
    else:
        st.caption(" 조회만 가능합니다.")
    render_calendar_monthly()

    # ── 일정 상세 인라인 패널 (날짜 클릭 시 캘린더 바로 아래 표시) ──
    if st.session_state.get("cal_action") in ("view_day", "add", "edit"):
        show_inline_day_panel()

# ── 조립 라인 ────────────────────────────────────────────────────
elif curr_l == "조립 라인":
    st.markdown(f"<h2 class='centered-title'> {curr_g} 신규 조립 현황</h2>", unsafe_allow_html=True)

    # ── 오늘 일정 알림 & 팝업 ─────────────────────────────────
    today_str   = datetime.now(KST).strftime('%Y-%m-%d')
    sch_all     = st.session_state.schedule_db
    # 조립 라인은 조립계획만, 포장 라인은 포장계획만 표시
    LINE_SCH_FILTER = {"조립 라인": "조립계획", "포장 라인": "포장계획"}
    sch_cat_filter  = LINE_SCH_FILTER.get(curr_l)
    if not sch_all.empty:
        _mask = (sch_all['날짜'] == today_str) & (sch_all['반'] == curr_g)
        if sch_cat_filter:
            _mask = _mask & (sch_all['카테고리'] == sch_cat_filter)
        today_sch = sch_all[_mask]
    else:
        today_sch = pd.DataFrame()

    # 변경 감지: 마지막 확인 이후 등록된 일정
    last_seen_key = f"sch_last_seen_{curr_g}"
    _first_load = last_seen_key not in st.session_state
    if _first_load:
        st.session_state[last_seen_key] = ""
    sch_ids_now   = ",".join(sorted(str(i) for i in today_sch['id'].tolist())) if not today_sch.empty else ""
    has_new_sch   = (sch_ids_now != st.session_state[last_seen_key]) and not today_sch.empty

    # 변경 알림 팝업
    if has_new_sch and not st.session_state.get(f"sch_popup_dismissed_{curr_g}", False):
        with st.container():
            st.warning(f" 오늘 생산 일정이 등록/변경되었습니다!\n\n{today_str} 기준 **{curr_g}** 일정 **{len(today_sch)}건**이 있습니다. 아래에서 확인하세요.")
            ack_c1, ack_c2 = st.columns([3, 1])
            if ack_c2.button(" 확인했습니다", key=f"sch_ack_{curr_g}", use_container_width=True, type="primary"):
                st.session_state[last_seen_key] = sch_ids_now
                st.session_state[f"sch_popup_dismissed_{curr_g}"] = True
                st.rerun()

    # 오늘 일정 카드
    _today_label = f" 오늘({today_str}) {curr_g} 작업 일정" + (f"  ·  {len(today_sch)}건" if not today_sch.empty else "  ·  없음")
    with st.expander(_today_label, expanded=_xp("asm_sch_today"), key="_xp_asm_sch_today"):
        if today_sch.empty:
            st.info("오늘 등록된 작업 일정이 없습니다.")
        else:
            th = st.columns([1.2, 2.8, 1.5, 0.8, 1.8, 2.5])
            for col, txt in zip(th, ["유형", "모델명", "P/N", "처리수", "출하계획", "특이사항"]):
                col.markdown(f"<p style='font-size:0.72rem;font-weight:700;color:#8a7f72;margin:0;padding-bottom:3px;border-bottom:2px solid #e0d8c8;'>{txt}</p>", unsafe_allow_html=True)
            # 성능: iterrows → to_dict('records')
            for sr in today_sch.to_dict('records'):
                cat   = str(sr.get('카테고리', '기타'))
                color = SCHEDULE_COLORS.get(cat, "#888")
                cat_esc = html_mod.escape(cat)
                model = str(sr.get('모델명', ''))
                pn    = str(sr.get('pn', ''))
                qty   = sr.get('조립수', 0)
                try:
                    qty = int(float(qty)) if qty not in (None, '', 'nan') else 0
                except (ValueError, TypeError):
                    qty = 0
                ship  = str(sr.get('출하계획', ''))
                note  = str(sr.get('특이사항', ''))
                rc = st.columns([1.2, 2.8, 1.5, 0.8, 1.8, 2.5])
                rc[0].markdown(f"<span style='background:{color}22;color:{color};border-left:3px solid {color};padding:1px 6px;border-radius:4px;font-size:0.75rem;font-weight:bold;'>{cat_esc}</span>", unsafe_allow_html=True)
                rc[1].write(model)
                rc[2].caption(pn if pn and pn != 'nan' else "-")
                rc[3].write(f"**{qty:,}**")
                rc[4].caption(ship if ship and ship != 'nan' else "-")
                rc[5].caption(f" {note}" if note and note != 'nan' else "-")

    # 이번 달 + 다음 달 일정 (오늘 이후)
    _next_mth = (date.today().replace(day=1) + timedelta(days=32)).strftime('%Y-%m')
    _month_sch_pre = sch_all[
        (sch_all['날짜'] >= today_str) &
        (sch_all['날짜'].str[:7].isin([today_str[:7], _next_mth])) &
        (sch_all['반'] == curr_g)
    ] if not sch_all.empty else pd.DataFrame()
    _month_sch_cnt = len(_month_sch_pre)
    with st.expander(f" {curr_g} 앞으로의 일정  ·  {_month_sch_cnt}건" if _month_sch_cnt else f" {curr_g} 앞으로의 일정  ·  없음", expanded=_xp("asm_sch_month"), key="_xp_asm_sch_month"):
        month_sch = _month_sch_pre
        if not month_sch.empty:
            show_cols = ['날짜','카테고리','모델명','pn','조립수','출하계획','특이사항']
            show_cols = [c for c in show_cols if c in month_sch.columns]
            st.dataframe(month_sch[show_cols].sort_values('날짜').rename(columns={'조립수': '처리수'}), use_container_width=True, hide_index=True)
        else:
            st.info("앞으로 등록된 일정이 없습니다.")

    st.divider()
    db_v = st.session_state.production_db
    db_g = db_v[db_v['반'] == curr_g]                                       # 현재 반 전체
    f_df = db_v[(db_v['반'] == curr_g) & (db_v['라인'] == "조립 라인")]
    _ASM_DEFECT_CAUSES = st.session_state.get('dropdown_defect_cause', ['(선택)', '기타 (직접 입력)'])

    # ──  오늘의 목표 달성 현황 ─────────────────────────────────
    _plan_qty = int(pd.to_numeric(today_sch['조립수'], errors='coerce').fillna(0).sum()) if not today_sch.empty else 0
    _done_today = len(f_df[
        f_df['시간'].astype(str).str.startswith(today_str) &
        f_df['상태'].isin(['검사대기','검사중','OQC대기','OQC중','출하승인','포장대기','포장중','완료'])
    ]) if not f_df.empty else 0
    _wip_today  = len(f_df[f_df['시간'].astype(str).str.startswith(today_str) & f_df['상태'].isin(['조립중','수리 완료(재투입)'])]) if not f_df.empty else 0

    if _plan_qty > 0:
        _real_pct = int(_done_today / _plan_qty * 100)
        _bar_pct  = min(_real_pct, 100)
        _over     = max(_done_today - _plan_qty, 0)
        _remain   = max(_plan_qty - _done_today, 0)
        if _real_pct >= 100:
            _bar_color = "#28a745"; _emoji = ""
            _msg = " 목표 달성! 수고하셨습니다!"
        elif _real_pct >= 80:
            _bar_color = "#28a745"; _emoji = ""
            _msg = f" 거의 다 왔어요! {_remain}개만 더!"
        elif _real_pct >= 50:
            _bar_color = "#ffc107"; _emoji = ""
            _msg = f" 절반 넘었어요! {_remain}개 남았어요!"
        else:
            _bar_color = "#e67e22"; _emoji = ""
            _msg = f" 파이팅! 목표까지 {_remain}개 남았어요!"

        # 초과 블록 HTML (한 줄로 유지 - 들여쓰기 4칸 이상이면 마크다운이 코드블록으로 처리)
        if _over > 0:
            _over_block = f"<div style='text-align:center;padding:12px 20px;background:#e8f8ee;border-radius:12px;border:2px solid #28a745;min-width:100px;'><div style='color:#28a745;font-size:2.2rem;font-weight:900;line-height:1;'>+{_over}</div><div style='color:#28a745;font-size:0.78rem;font-weight:600;margin-top:2px;'>초과 달성</div></div>"
        else:
            _over_block = "<div style='text-align:center;padding:12px 20px;background:#f5f5f5;border-radius:12px;border:2px solid #ddd;min-width:100px;'><div style='color:#aaa;font-size:2.2rem;font-weight:900;line-height:1;'>-</div><div style='color:#aaa;font-size:0.78rem;font-weight:600;margin-top:2px;'>초과 없음</div></div>"

        st.markdown(f"""
        <div style='background:#ffffff;border-radius:16px;padding:24px 28px;margin-bottom:16px;
                    border:2px solid {_bar_color};box-shadow:0 4px 16px rgba(0,0,0,0.1);'>
            <div style='display:flex;align-items:center;justify-content:space-between;margin-bottom:16px;'>
                <span style='color:#1a1a2e;font-size:1.1rem;font-weight:700;'> 오늘의 목표 달성 현황</span>
                <span style='color:#888;font-size:0.85rem;'>{today_str}</span>
            </div>
            <div style='display:flex;align-items:center;gap:16px;margin-bottom:14px;'>
                <div style='text-align:center;padding:12px 20px;background:#f0f4ff;
                            border-radius:12px;border:2px solid {_bar_color};min-width:100px;'>
                    <div style='color:{_bar_color};font-size:2.2rem;font-weight:900;line-height:1;'>{_done_today}</div>
                    <div style='color:#555;font-size:0.78rem;font-weight:600;margin-top:2px;'>오늘 누적</div>
                </div>
                <div style='color:#bbb;font-size:1.6rem;font-weight:300;'>/</div>
                <div style='text-align:center;padding:12px 20px;background:#fafafa;
                            border-radius:12px;border:2px solid #ddd;min-width:100px;'>
                    <div style='color:#444;font-size:2.2rem;font-weight:900;line-height:1;'>{_plan_qty}</div>
                    <div style='color:#888;font-size:0.78rem;font-weight:600;margin-top:2px;'>오늘 목표</div>
                </div>
                <div style='flex:1;'></div>
                {_over_block}
                <span style='font-size:2rem;'>{_emoji}</span>
            </div>
            <div style='background:#e9ecef;border-radius:8px;height:14px;overflow:hidden;margin-bottom:10px;'>
                <div style='background:{_bar_color};width:{_bar_pct}%;height:100%;border-radius:8px;
                            transition:width 0.4s ease;'></div>
            </div>
            <div style='display:flex;justify-content:space-between;align-items:center;'>
                <span style='color:#333;font-size:0.92rem;font-weight:600;'>{_msg}</span>
                <span style='color:#666;font-size:0.82rem;'>{_real_pct}% · 작업 중: {_wip_today}개</span>
            </div>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.info(f" 오늘({today_str}) 등록된 조립 계획이 없습니다. 일정을 등록하면 달성 현황이 표시됩니다.", icon="ℹ")

    # ── 모델/품목별 수량 카운트 + 생산 이력 ─────────────────────────
    if not f_df.empty:
        if curr_g != "제조2반":
            with st.expander(f" {curr_g} 조립 라인 수량 현황  ·  {len(f_df)}건", expanded=_xp("asm_cnt"), key="_xp_asm_cnt"):
                grp = f_df.groupby(['모델','품목코드'])
                count_rows = []
                for (model, pn), gdf in grp:
                    total  = len(gdf)
                    done   = len(gdf[gdf['상태'].isin(['검사대기','검사중','OQC대기','OQC중','출하승인','포장대기','포장중','완료'])])
                    wip    = len(gdf[gdf['상태'].isin(['조립중','수리 완료(재투입)'])])
                    defect = len(gdf[gdf['상태'].str.contains('불량|부적합', na=False)])
                    count_rows.append((model, pn, total, done, wip, defect))

                for (model, pn, total, done, wip, defect) in count_rows:
                    pct = int(done / total * 100) if total > 0 else 0
                    with st.container(border=True):
                        mc1, mc2 = st.columns([3, 1])
                        mc1.markdown(f"**{model}**" + (f" `{pn}`" if pn else ""))
                        mc2.markdown(f"완료율 **{pct}%**")
                        st.progress(min(pct, 100) / 100)
                        sc1, sc2, sc3, sc4 = st.columns(4)
                        sc1.metric("전체", total)
                        sc2.metric(" 완료", done)
                        sc3.metric(" 작업중", wip)
                        sc4.metric(" 불량", defect, delta=None if defect == 0 else f"{defect}건", delta_color="inverse")

        _hist_wip        = len(db_g[db_g['상태'].isin(WIP_STATES)])
        _hist_done_today = len(db_g[
            db_g['시간'].astype(str).str.startswith(today_str) &
            db_g['상태'].isin(DONE_STATES)
        ])
        with st.expander(f" {curr_g} 생산 이력  ·  조립 대기 {_hist_wip}건 / 당일 완료 {_hist_done_today}건", expanded=_xp("asm_hist"), key="_xp_asm_hist"):
            _asm_chk_key = f"asm_checked_{curr_g}"
            if _asm_chk_key not in st.session_state:
                st.session_state[_asm_chk_key] = {}

            _asm_search_cnt = f"asm_search_cnt_{curr_g}"
            if _asm_search_cnt not in st.session_state:
                st.session_state[_asm_search_cnt] = 0
            _asm_search_key = f"sn_search_{curr_g}"
            sc1, sc2 = st.columns([2, 2])
            sn_search = sc1.text_input(" 시리얼 검색", placeholder="S/N 스캔 또는 입력...", key=_asm_search_key)
            if sn_search.strip():
                # 반 전체(db_g)에서 검색 — 이미 다른 라인으로 이동한 시리얼도 조회 가능
                f_df_view = db_g[db_g['시리얼'].str.contains(sn_search.strip(), case=False, na=False)]
                if f_df_view.empty:
                    st.warning(f" **'{sn_search.strip()}'** 에 해당하는 시리얼이 없습니다.")
                # 자동 체크는 조립 라인의 처리 가능 상태(actionable)만 적용
                _asm_cb_ver_now = st.session_state[_asm_search_cnt]
                _wip_mask = f_df_view['상태'].isin(["조립중", "수리 완료(재투입)"])
                for _si in f_df_view.index[_wip_mask]:
                    st.session_state[_asm_chk_key][str(_si)] = True
                    st.session_state[f"asm_cb_{curr_g}_{_si}_{_asm_cb_ver_now}"] = True
            else:
                # 검색어 없을 때는 조립 라인 항목만 표시
                f_df_view = f_df

            checked_idxs = [k for k,v in st.session_state[_asm_chk_key].items() if v]
            if checked_idxs:
                ba1, ba2, ba3, ba4 = st.columns([2, 1.5, 1, 1])
                ba1.markdown(f"<span style='color:#2E75B6;font-weight:700;'> {len(checked_idxs)}개 선택됨</span>", unsafe_allow_html=True)
                _bulk_ng_cause = ba2.selectbox("불량 원인", _ASM_DEFECT_CAUSES, key=f"bulk_ng_cause_{curr_g}", label_visibility="collapsed")
                if _bulk_ng_cause == "기타 (직접 입력)":
                    _bulk_ng_cause_final = st.text_input("직접 입력", key=f"bulk_ng_cause_txt_{curr_g}", placeholder="불량 원인 직접 입력")
                else:
                    _bulk_ng_cause_final = _bulk_ng_cause
                if ba3.button(" 일괄 완료", key=f"bulk_ok_{curr_g}", type="primary", use_container_width=True,
                             disabled=not check_perm(f"조립 라인::{curr_g}", "write")):
                    _ops = []
                    for ci in checked_idxs:
                        ci_int = int(ci)
                        if ci_int in f_df.index:
                            _r = f_df.loc[ci_int]
                            _upd = {'상태':'검사대기','시간':get_now_kst_str()}
                            _ops.append({"sn": _r['시리얼'], "data": _upd,
                                "audit": {"시리얼": _r['시리얼'], "모델": _r['모델'], "반": curr_g,
                                         "이전상태": _r['상태'], "이후상태": '검사대기',
                                         "작업자": st.session_state.user_id}})
                    st.session_state[_asm_chk_key] = {}
                    st.session_state[_asm_search_cnt] += 1  # 체크박스 키 리셋
                    _prod_bulk_update(_run_bulk_db_ops(_ops))
                    _rerun("asm_hist")
                if ba4.button(" 일괄 불량", key=f"bulk_ng_{curr_g}", use_container_width=True,
                             disabled=not check_perm(f"조립 라인::{curr_g}", "write")):
                    if _bulk_ng_cause_final in ["(선택)", ""]:
                        st.warning(" 불량 원인을 먼저 선택해주세요.")
                    else:
                        _ops = []
                        for ci in checked_idxs:
                            ci_int = int(ci)
                            if ci_int in f_df.index:
                                _r = f_df.loc[ci_int]
                                _upd = {'상태':'불량 처리 중','시간':get_now_kst_str(),
                                    '증상': f'불량입고출처: 조립 라인 | 불량원인: {_bulk_ng_cause_final}'}
                                _ops.append({"sn": _r['시리얼'], "data": _upd,
                                    "audit": {"시리얼": _r['시리얼'], "모델": _r['모델'], "반": curr_g,
                                             "이전상태": _r['상태'], "이후상태": '불량 처리 중',
                                             "작업자": st.session_state.user_id}})
                        st.session_state[_asm_chk_key] = {}
                        st.session_state[_asm_search_cnt] += 1  # 체크박스 키 리셋
                        _prod_bulk_update(_run_bulk_db_ops(_ops))
                        _rerun("asm_hist")

            # STATUS_STYLE: 모듈 상수 사용 (상단 정의 참조)
            h = st.columns([0.4, 2.0, 1.8, 1.4, 1.6, 1.1, 1.1])
            for col, txt in zip(h, ["","기록 시간","모델","품목","시리얼","",""]):
                col.markdown(f"<p style='font-size:0.72rem;font-weight:700;color:#8a7f72;margin:0;text-align:center;'>{txt}</p>", unsafe_allow_html=True)

            _asm_cb_ver = st.session_state[_asm_search_cnt]  # 스캔 시 변경 → 체크박스 강제 재렌더
            # 자재 시리얼 일괄 조회 (N+1 방지)
            _asm_bulk_sns = tuple(f_df_view['시리얼'].unique().tolist())
            _asm_bulk_mats = load_material_serials_bulk(_asm_bulk_sns) if _asm_bulk_sns else pd.DataFrame()
            for row in f_df_view.sort_values('시간', ascending=False).reset_index().to_dict('records'):
                idx = row['index']
                is_actionable = row['상태'] in ["조립중", "수리 완료(재투입)"]
                r = st.columns([0.4, 2.0, 1.8, 1.4, 1.6, 1.1, 1.1])
                if is_actionable:
                    _ck = r[0].checkbox("", key=f"asm_cb_{curr_g}_{idx}_{_asm_cb_ver}",
                        value=st.session_state[_asm_chk_key].get(str(idx), False),
                        label_visibility="collapsed")
                    st.session_state[_asm_chk_key][str(idx)] = _ck
                else:
                    _ck = False
                r[1].caption(str(row['시간'])[:16]); r[2].caption(row['모델'])
                r[3].caption(row['품목코드'])
                _asm_mc = len(_asm_bulk_mats[_asm_bulk_mats['메인시리얼'] == row['시리얼']]) if not _asm_bulk_mats.empty else 0
                _asm_tog_key = f"mat_tog_{row['시리얼']}_{curr_g}_asm"
                _asm_btn_lbl = f"{row['시리얼']}\n자재 {_asm_mc}개" if _asm_mc > 0 else row['시리얼']
                if r[4].button(_asm_btn_lbl, key=f"sntog_asm_{idx}_{_asm_cb_ver}",
                               use_container_width=True, help="클릭하여 자재 시리얼 조회"):
                    st.session_state[_asm_tog_key] = not st.session_state.get(_asm_tog_key, False)
                if is_actionable:
                    if r[5].button("완료", key=f"ok_{idx}", use_container_width=True, type="primary"):
                        _upd = {'상태':'검사대기','시간':get_now_kst_str()}
                        update_row(row['시리얼'], _upd)
                        insert_audit_log(시리얼=row['시리얼'], 모델=row['모델'], 반=curr_g,
                            이전상태=row['상태'], 이후상태='검사대기', 작업자=st.session_state.user_id)
                        st.session_state[_asm_chk_key].pop(str(idx), None)
                        st.session_state[_asm_search_cnt] += 1  # 체크박스 키 리셋
                        _prod_update(row['시리얼'], _upd)
                        _rerun("asm_hist")
                    _ng_open_key = f"_ng_open_asm_{idx}"
                    if not st.session_state.get(_ng_open_key):
                        if r[6].button("불량", key=f"ng_{idx}", use_container_width=True):
                            st.session_state[_ng_open_key] = True
                            _rerun("asm_hist")
                    else:
                        r[6].button("불량", key=f"ng_{idx}", use_container_width=True, disabled=True)
                else:
                    s = row['상태']
                    s_esc = html_mod.escape(str(s))
                    if "불량" in str(s):
                        r[5].markdown(f"<div style='background:#fde8e7;color:#7a2e2a;padding:2px 6px;border-radius:5px;text-align:center;font-weight:bold;font-size:0.75rem;'> {s_esc}</div>", unsafe_allow_html=True)
                    else:
                        bg,tc,bc,ic = STATUS_STYLE.get(s, ('#f5f2ec','#5a5048','#c8b89a','•'))
                        ic_esc = html_mod.escape(str(ic))
                        r[5].markdown(f"<div style='background:{bg};color:{tc};padding:2px 6px;border-radius:5px;text-align:center;font-weight:bold;border:1px solid {bc};font-size:0.75rem;'>{ic_esc} {s_esc}</div>", unsafe_allow_html=True)
                # ── 불량 원인 선택 패널 (개별  클릭 후) ──
                _ng_open_key = f"_ng_open_asm_{idx}"
                if st.session_state.get(_ng_open_key):
                    with st.container(border=True):
                        st.caption(f" 불량 원인 입력 — `{row['시리얼']}`")
                        _nc1, _nc2, _nc3 = st.columns([2.5, 1, 1])
                        _cause_ng = _nc1.selectbox("불량 원인", _ASM_DEFECT_CAUSES, key=f"_ng_cause_asm_{idx}", label_visibility="collapsed")
                        if _cause_ng == "기타 (직접 입력)":
                            _cause_ng_final = st.text_input("직접 입력", key=f"_ng_cause_asm_txt_{idx}", placeholder="불량 원인 직접 입력")
                        else:
                            _cause_ng_final = _cause_ng
                        if _nc2.button("확정", key=f"_ng_confirm_asm_{idx}", type="primary", use_container_width=True):
                            if _cause_ng_final in ["(선택)", ""]:
                                st.warning(" 불량 원인을 선택해주세요.")
                            else:
                                _upd = {'상태':'불량 처리 중','시간':get_now_kst_str(),
                                    '증상': f'불량입고출처: 조립 라인 | 불량원인: {_cause_ng_final}'}
                                update_row(row['시리얼'], _upd)
                                insert_audit_log(시리얼=row['시리얼'], 모델=row['모델'], 반=curr_g,
                                    이전상태=row['상태'], 이후상태='불량 처리 중', 작업자=st.session_state.user_id)
                                st.session_state[_asm_chk_key].pop(str(idx), None)
                                st.session_state[_asm_search_cnt] += 1
                                st.session_state.pop(_ng_open_key, None)
                                _prod_update(row['시리얼'], _upd)
                                _rerun("asm_hist")
                        if _nc3.button("취소", key=f"_ng_cancel_asm_{idx}", use_container_width=True):
                            st.session_state.pop(_ng_open_key, None)
                            _rerun("asm_hist")
                # ── 자재 시리얼 토글 표시 ──
                if st.session_state.get(_asm_tog_key, False):
                    _asm_row_mats = _asm_bulk_mats[_asm_bulk_mats['메인시리얼'] == row['시리얼']] if not _asm_bulk_mats.empty else pd.DataFrame()
                    with st.container(border=True):
                        st.caption(f" 자재 시리얼 — `{row['시리얼']}`")
                        if not _asm_row_mats.empty:
                            for _am in _asm_row_mats.to_dict('records'):
                                amc1, amc2 = st.columns([2, 4])
                                amc1.caption(_am.get('자재명', ''))
                                amc2.caption(f"`{_am.get('자재시리얼', '')}`")
                        else:
                            st.caption("등록된 자재 시리얼 없음")
    else:
        st.info("등록된 생산 내역이 없습니다.")

    # 자재 목록 마스터
    MAT_NAME_OPTIONS = st.session_state.get("dropdown_mat_name") or []

    _mat_list_key  = f"mat_list_{curr_g}"
    _scan_sn_key   = f"scan_sn_{curr_g}"
    _mat_name_key  = f"mat_name_sel_{curr_g}"

    if _mat_list_key not in st.session_state:
        st.session_state[_mat_list_key] = []

    with st.container(border=True):
        st.markdown(f"####  {curr_g} 신규 생산 등록")

        g_models     = st.session_state.group_master_models.get(curr_g, [])
        target_model = st.selectbox("투입 모델 선택", ["선택하세요."] + g_models, key=f"model_sel_{curr_g}")
        g_items      = st.session_state.group_master_items.get(curr_g, {}).get(target_model, [])

        ef1, ef2 = st.columns(2)
        target_item = ef1.selectbox("품목 코드",
            (g_items if g_items else ["(품목코드 없음)"]) if target_model != "선택하세요." else ["모델 선택 대기"],
            key=f"item_sel_{curr_g}")
        _msn_cnt_key = f"msn_cnt_{curr_g}"
        if _msn_cnt_key not in st.session_state:
            st.session_state[_msn_cnt_key] = 0
        _msn_field_key = f"sn_input_{curr_g}_{st.session_state[_msn_cnt_key]}"
        target_sn = ef2.text_input(
            " 메인 S/N",
            placeholder="S/N 입력 후 아래 버튼으로 등록",
            key=_msn_field_key)
        ef2.caption(" 자재 시리얼 입력 완료 후 [생산 시작 등록] 버튼을 누르세요")

        st.divider()

        st.markdown("<p style='font-size:0.88rem;font-weight:700;color:#5a4f45;margin:0 0 6px 0;'> 자재 시리얼</p>", unsafe_allow_html=True)

        sc1, sc2, sc3 = st.columns([2, 3, 1])
        sel_mat_name = sc1.selectbox("자재명 선택", MAT_NAME_OPTIONS, key=_mat_name_key)

        _scan_counter_key = f"scan_cnt_{curr_g}"
        if _scan_counter_key not in st.session_state:
            st.session_state[_scan_counter_key] = 0
        _scan_processed_key = f"scan_proc_{curr_g}"
        if _scan_processed_key not in st.session_state:
            st.session_state[_scan_processed_key] = ""
        _scan_field_key = f"{_scan_sn_key}_{st.session_state[_scan_counter_key]}"

        scan_input = sc2.text_input(
            "자재 S/N 스캔",
            placeholder="바코드 스캔 → 자동 추가 (Enter)",
            key=_scan_field_key,
        )
        if st.session_state.pop("_autofocus_after_rerun", None) == _scan_field_key:
            _inject_autofocus("자재 S/N 스캔")
        sc2.caption(" 스캐너로 스캔하면 Enter가 자동 입력됩니다")

        if scan_input.strip():
            import time as _time
            _scanned = scan_input.strip()
            _scan_ts_key  = f"scan_ts_{curr_g}"
            _last_ts  = st.session_state.get(_scan_ts_key, 0)
            _last_val = st.session_state.get(_scan_processed_key, "")
            # 2초 이내 동일 값 → 중복 스킵 (렉/스캐너 이중 전송 방지)
            if _scanned == _last_val and (_time.monotonic() - _last_ts) < 5.0:
                st.session_state[_scan_counter_key] += 1
                st.rerun()
            st.session_state[_scan_processed_key] = _scanned
            st.session_state[_scan_ts_key] = _time.monotonic()
            already = any(m["자재시리얼"] == _scanned
                         for m in st.session_state[_mat_list_key])
            if not already:
                st.session_state[_mat_list_key].append({
                    "자재명": sel_mat_name,
                    "자재시리얼": _scanned
                })
            else:
                st.toast(f" 이미 추가된 자재 S/N: {_scanned}")
            st.session_state["_autofocus_after_rerun"] = f"{_scan_sn_key}_{st.session_state[_scan_counter_key] + 1}"
            st.session_state[_scan_counter_key] += 1
            st.rerun()

        if sc3.button(" 추가", key=f"mat_manual_add_{curr_g}", use_container_width=True):
            st.session_state[_mat_list_key].append({
                "자재명": sel_mat_name, "자재시리얼": ""
            })
            st.rerun()

        mat_list_now = st.session_state[_mat_list_key]
        if mat_list_now:
            st.markdown(f"<p style='font-size:0.78rem;color:#8a7f72;margin:6px 0 2px 0;'>등록 예정 자재: <b>{len(mat_list_now)}개</b></p>", unsafe_allow_html=True)
            lh1, lh2, lh3 = st.columns([2, 4, 1])
            lh1.markdown("<p style='font-size:0.7rem;font-weight:700;color:#aaa;margin:0;'>자재명</p>", unsafe_allow_html=True)
            lh2.markdown("<p style='font-size:0.7rem;font-weight:700;color:#aaa;margin:0;'>자재 S/N</p>", unsafe_allow_html=True)

            updated_list = []
            _should_rerun = False
            for mi, mat in enumerate(mat_list_now):
                lc1, lc2, lc3 = st.columns([2, 4, 1])
                new_name = lc1.selectbox("", MAT_NAME_OPTIONS,
                    index=MAT_NAME_OPTIONS.index(mat["자재명"]) if mat["자재명"] in MAT_NAME_OPTIONS else 0,
                    key=f"mat_nm_{curr_g}_{mi}", label_visibility="collapsed")
                new_sn = lc2.text_input("", value=mat["자재시리얼"],
                    key=f"mat_sv_{curr_g}_{mi}", label_visibility="collapsed",
                    placeholder="S/N 직접 입력 또는 스캔")
                if not lc3.button("삭제", key=f"mat_del_{curr_g}_{mi}", help="삭제"):
                    updated_list.append({"자재명": new_name, "자재시리얼": new_sn})
                else:
                    _should_rerun = True  # 삭제 버튼 클릭됨 — 해당 항목은 updated_list에 추가 안 됨

            st.session_state[_mat_list_key] = updated_list  # 먼저 저장
            if _should_rerun:
                st.rerun()  # 저장 후 rerun

            if st.button(" 전체 초기화", key=f"mat_clear_{curr_g}", type="secondary"):
                st.session_state[_mat_list_key] = []
                st.rerun()
        else:
            st.caption("자재 없음 — 스캔하거나  추가 버튼을 누르세요")

        st.divider()

        def _do_register_sn(sn_val):
            if insert_row({
                '시간': get_now_kst_str(), '반': curr_g, '라인': "조립 라인",
                '모델': target_model, '품목코드': target_item,
                '시리얼': sn_val, '상태': '조립중',
                '증상': '', '수리': '', '작업자': st.session_state.user_id
            }):
                insert_audit_log(시리얼=sn_val, 모델=target_model, 반=curr_g,
                    이전상태="-", 이후상태="조립중", 작업자=st.session_state.user_id)
                valid_mats = [m for m in st.session_state[_mat_list_key] if m["자재시리얼"].strip()]
                if valid_mats:
                    insert_material_serials(
                        메인시리얼=sn_val, 모델=target_model,
                        반=curr_g, 자재목록=valid_mats,
                        작업자=st.session_state.user_id
                    )
                st.session_state[_mat_list_key] = []
                st.session_state[f"scan_cnt_{curr_g}"] = 0
                st.session_state[_msn_cnt_key] += 1
                _clear_production_cache()
                st.session_state.production_db = load_realtime_ledger()
                st.toast(f" 등록 완료: {sn_val}")
                st.rerun()

        _start_col, _ = st.columns([1, 2])
        if _start_col.button("▶ 생산 시작 등록", use_container_width=True, type="primary", key=f"start_btn_{curr_g}"):
            if target_model != "선택하세요." and target_item not in [None, "", "모델 선택 대기", "(품목코드 없음)"] and target_sn.strip():
                _do_register_sn(target_sn.strip())
            else:
                st.warning("모델, 품목코드, 메인 S/N을 모두 입력해주세요.")

    # ── 기존 제품 자재 시리얼 추가 ────────────────────────────────────
    with st.expander(" 기존 제품 자재 시리얼 추가 등록", expanded=_xp("asm_mat"), key="_xp_asm_mat"):
        st.caption("메인 S/N을 입력하면 기존 등록된 자재를 조회하고 누락된 자재를 추가할 수 있습니다.")

        _add_mat_sn_cnt_key = f"add_mat_sn_cnt_{curr_g}"
        if _add_mat_sn_cnt_key not in st.session_state:
            st.session_state[_add_mat_sn_cnt_key] = 0
        _add_main_sn_key = f"add_main_sn_{curr_g}_{st.session_state[_add_mat_sn_cnt_key]}"
        add_main_sn = st.text_input(" 메인 S/N 조회", placeholder="기존 등록된 메인 S/N 입력", key=_add_main_sn_key)

        if add_main_sn.strip():
            _exist_row = db_v[db_v['시리얼'] == add_main_sn.strip()]
            if _exist_row.empty:
                st.warning(f" 등록되지 않은 S/N입니다: **{add_main_sn.strip()}**")
            else:
                _er = _exist_row.iloc[0]
                st.success(f" 등록 확인: **{_er['모델']}** `{_er['품목코드']}` — 상태: {_er['상태']}")

                existing_mats = load_material_serials(add_main_sn.strip())
                if not existing_mats.empty:
                    st.markdown(f"<p style='font-size:0.78rem;color:#8a7f72;margin:6px 0 4px 0;'>기존 등록 자재: <b>{len(existing_mats)}개</b></p>", unsafe_allow_html=True)
                    emh1, emh2 = st.columns([2, 4])
                    emh1.markdown("<p style='font-size:0.7rem;font-weight:700;color:#aaa;margin:0;'>자재명</p>", unsafe_allow_html=True)
                    emh2.markdown("<p style='font-size:0.7rem;font-weight:700;color:#aaa;margin:0;'>자재 S/N</p>", unsafe_allow_html=True)
                    for mat_r in existing_mats.to_dict('records'):
                        emc1, emc2 = st.columns([2, 4])
                        emc1.caption(mat_r.get('자재명', ''))
                        emc2.caption(f"`{mat_r.get('자재시리얼', '')}`")
                else:
                    st.info("기존 등록된 자재 시리얼이 없습니다.")

                st.divider()
                st.markdown("<p style='font-size:0.85rem;font-weight:700;color:#5a4f45;margin:0 0 6px 0;'> 자재 추가 등록</p>", unsafe_allow_html=True)

                _add_mat_list_key = f"add_mat_list_{curr_g}"
                if _add_mat_list_key not in st.session_state:
                    st.session_state[_add_mat_list_key] = []
                _add_scan_cnt_key = f"add_scan_cnt_{curr_g}"
                if _add_scan_cnt_key not in st.session_state:
                    st.session_state[_add_scan_cnt_key] = 0

                asc1, asc2, asc3 = st.columns([2, 3, 1])
                add_sel_mat_name = asc1.selectbox("자재명 선택", MAT_NAME_OPTIONS, key=f"add_mat_nm_sel_{curr_g}")
                _add_scan_field_key = f"add_scan_sn_{curr_g}_{st.session_state[_add_scan_cnt_key]}"
                add_scan_input = asc2.text_input("자재 S/N 스캔", placeholder="바코드 스캔 → 자동 추가 (Enter)", key=_add_scan_field_key)
                if st.session_state.pop("_autofocus_after_rerun", None) == _add_scan_field_key:
                    _inject_autofocus("자재 S/N 스캔")
                asc2.caption(" 스캐너로 스캔하면 Enter가 자동 입력됩니다")

                if add_scan_input.strip():
                    _already = any(m["자재시리얼"] == add_scan_input.strip() for m in st.session_state[_add_mat_list_key])
                    if not _already:
                        st.session_state[_add_mat_list_key].append({"자재명": add_sel_mat_name, "자재시리얼": add_scan_input.strip()})
                    else:
                        st.toast(f" 이미 추가된 자재 S/N: {add_scan_input.strip()}")
                    st.session_state["_autofocus_after_rerun"] = f"add_scan_sn_{curr_g}_{st.session_state[_add_scan_cnt_key] + 1}"
                    st.session_state[_add_scan_cnt_key] += 1
                    _rerun("asm_mat")

                if asc3.button(" 추가", key=f"add_mat_manual_{curr_g}", use_container_width=True):
                    st.session_state[_add_mat_list_key].append({"자재명": add_sel_mat_name, "자재시리얼": ""})
                    _rerun("asm_mat")

                add_mat_list_now = st.session_state[_add_mat_list_key]
                if add_mat_list_now:
                    st.markdown(f"<p style='font-size:0.78rem;color:#8a7f72;margin:6px 0 2px 0;'>추가 예정 자재: <b>{len(add_mat_list_now)}개</b></p>", unsafe_allow_html=True)
                    add_updated_list = []
                    _add_should_rerun = False
                    for ami, amat in enumerate(add_mat_list_now):
                        alc1, alc2, alc3 = st.columns([2, 4, 1])
                        anew_name = alc1.selectbox("", MAT_NAME_OPTIONS,
                            index=MAT_NAME_OPTIONS.index(amat["자재명"]) if amat["자재명"] in MAT_NAME_OPTIONS else 0,
                            key=f"add_mat_nm_{curr_g}_{ami}", label_visibility="collapsed")
                        anew_sn = alc2.text_input("", value=amat["자재시리얼"],
                            key=f"add_mat_sv_{curr_g}_{ami}", label_visibility="collapsed",
                            placeholder="S/N 직접 입력 또는 스캔")
                        if not alc3.button("삭제", key=f"add_mat_del_{curr_g}_{ami}", help="삭제"):
                            add_updated_list.append({"자재명": anew_name, "자재시리얼": anew_sn})
                        else:
                            _add_should_rerun = True
                    st.session_state[_add_mat_list_key] = add_updated_list
                    if _add_should_rerun:
                        _rerun("asm_mat")

                    _mat_btn_col, _ = st.columns([1, 2])
                    if _mat_btn_col.button(" 자재 시리얼 추가 저장", key=f"add_mat_save_{curr_g}", type="primary", use_container_width=True):
                        valid_add_mats = [m for m in st.session_state[_add_mat_list_key] if m["자재시리얼"].strip()]
                        if valid_add_mats:
                            if insert_material_serials(
                                메인시리얼=add_main_sn.strip(), 모델=_er['모델'],
                                반=curr_g, 자재목록=valid_add_mats,
                                작업자=st.session_state.user_id
                            ):
                                st.session_state[_add_mat_list_key] = []
                                st.session_state[_add_mat_sn_cnt_key] += 1
                                st.session_state[_add_scan_cnt_key] = 0
                                st.cache_data.clear()
                                st.toast(f" {len(valid_add_mats)}개 자재 시리얼 추가 완료")
                                _rerun("asm_mat")
                        else:
                            st.warning("추가할 자재 시리얼을 입력해주세요.")
                else:
                    st.caption("자재 없음 — 스캔하거나  추가 버튼을 누르세요")


# ── 검사 / 포장 라인 ─────────────────────────────────────────────
elif curr_l in ["검사 라인", "포장 라인"]:
    st.markdown(f"<h2 class='centered-title'> {curr_g} {curr_l} 현황</h2>", unsafe_allow_html=True)
    prev = "조립 라인" if curr_l == "검사 라인" else "OQC 라인"

    db_s = st.session_state.production_db
    if curr_l == "검사 라인":
        wait_list = db_s[(db_s['반']==curr_g)&(db_s['상태'].isin(['검사대기','수리 완료(재투입)']))]
    else:
        wait_list = db_s[(db_s['반']==curr_g)&(db_s['상태']=='출하승인')]
    _wait_cnt = len(wait_list)
    DEFECT_CAUSES = st.session_state.get('dropdown_defect_cause', ['(선택)', '기타 (직접 입력)'])


    _wck_key     = f"wait_ck_{curr_g}_{curr_l}"
    _wscan_cnt   = f"wscan_cnt_{curr_g}_{curr_l}"
    if _wck_key   not in st.session_state: st.session_state[_wck_key]   = {}
    if _wscan_cnt not in st.session_state: st.session_state[_wscan_cnt] = 0

    with st.expander(f" 이전 공정({prev}) 완료 — 입고 대기" + (f"  ·  {_wait_cnt}건" if _wait_cnt else "  ·  없음"), expanded=_xp("chk_wait"), key="_xp_chk_wait"):
        if not wait_list.empty:
            _wscan_key = f"wscan_{curr_g}_{curr_l}_{st.session_state[_wscan_cnt]}"
            ws1, ws2 = st.columns([3, 3])
            w_scan = ws1.text_input(" 시리얼 스캔/검색", placeholder="스캔 또는 입력 → 자동 체크",
                                    key=_wscan_key)
            if st.session_state.pop("_autofocus_after_rerun", None) == _wscan_key:
                _inject_autofocus(placeholder="스캔 또는 입력 → 자동 체크")
            if w_scan.strip():
                matched_sn = wait_list[wait_list['시리얼'].str.contains(
                    w_scan.strip(), case=False, na=False)]
                if not matched_sn.empty:
                    for wi in matched_sn.index:
                        st.session_state[_wck_key][str(wi)] = True
                    st.session_state["_autofocus_after_rerun"] = f"wscan_{curr_g}_{curr_l}_{st.session_state[_wscan_cnt] + 1}"
                    st.session_state[_wscan_cnt] += 1  # 키 변경 → 체크박스 강제 재렌더
                    _rerun("chk_wait")
                else:
                    ws1.warning(f"**'{w_scan.strip()}'** — 대기 목록에 없습니다.")

            w_checked = [k for k,v in st.session_state[_wck_key].items() if v]
            if w_checked:
                wba1, wba2, wba3 = st.columns([3, 1, 1])
                wba1.markdown(f"<span style='color:#2E75B6;font-weight:700;'> {len(w_checked)}개 선택됨</span>",
                              unsafe_allow_html=True)
                if wba2.button(" 일괄 입고", key=f"wait_bulk_{curr_g}_{curr_l}",
                               type="primary", use_container_width=True):
                    _next_s = '검사중' if curr_l == '검사 라인' else '포장중'
                    _ops = []
                    for wi in w_checked:
                        wi_int = int(wi)
                        if wi_int in wait_list.index:
                            _wr = wait_list.loc[wi_int]
                            _upd = {'시간': get_now_kst_str(),
                                '라인': curr_l, '상태': _next_s,
                                '작업자': st.session_state.user_id}
                            _ops.append({"sn": _wr['시리얼'], "data": _upd,
                                "audit": {"시리얼": _wr['시리얼'], "모델": _wr['모델'],
                                         "반": curr_g, "이전상태": _wr['상태'], "이후상태": _next_s,
                                         "작업자": st.session_state.user_id}})
                    st.session_state[_wck_key] = {}
                    st.session_state[_wscan_cnt] += 1  # 체크박스 키 리셋
                    _prod_bulk_update(_run_bulk_db_ops(_ops))
                    _rerun("chk_wait")
                if wba3.button(" 선택 해제", key=f"wait_unck_{curr_g}_{curr_l}",
                               use_container_width=True):
                    st.session_state[_wck_key] = {}
                    st.session_state[_wscan_cnt] += 1  # 체크박스 키 리셋
                    _rerun("chk_wait")

            st.markdown("<hr style='margin:8px 0;border-color:#e0d8c8;'>", unsafe_allow_html=True)

            grp_w = wait_list.groupby(['모델','품목코드'])
            for (w_model, w_pn), w_gdf in grp_w:
                with st.container(border=True):
                    wc1, wc2 = st.columns([4, 1])
                    wc1.markdown(f"**{w_model}**" + (f"  `{w_pn}`" if w_pn else ""))
                    wc2.caption(f"{len(w_gdf)}대")
                    _wcb_ver = st.session_state[_wscan_cnt]
                    for wrow in w_gdf.reset_index().to_dict('records'):
                        widx = wrow['index']
                        wr1, wr2, wr3 = st.columns([0.5, 3, 1.2])
                        _wck = wr1.checkbox("", key=f"wck_{curr_g}_{curr_l}_{widx}_{_wcb_ver}",
                            value=st.session_state[_wck_key].get(str(widx), False),
                            label_visibility="collapsed")
                        st.session_state[_wck_key][str(widx)] = _wck
                        wr2.markdown(f"`{wrow['시리얼']}`  <span style='color:#999;font-size:0.75rem;'>{str(wrow.get('시간',''))[:16]}</span>",
                                    unsafe_allow_html=True)
                        if wr3.button(" 입고", key=f"in_{widx}", use_container_width=True):
                            _next_s = '검사중' if curr_l == '검사 라인' else '포장중'
                            _upd = {'시간': get_now_kst_str(),
                                '라인': curr_l, '상태': _next_s,
                                '작업자': st.session_state.user_id}
                            update_row(wrow['시리얼'], _upd)
                            insert_audit_log(시리얼=wrow['시리얼'], 모델=wrow['모델'],
                                반=curr_g, 이전상태=wrow['상태'], 이후상태=_next_s,
                                작업자=st.session_state.user_id)
                            st.session_state[_wck_key].pop(str(widx), None)
                            _prod_update(wrow['시리얼'], _upd)
                            _rerun("chk_wait")
        else:
            st.info("입고 대기 물량 없음")

    st.divider()
    f_df = db_s[(db_s['반']==curr_g)&(db_s['라인']==curr_l)]
    _hist_cnt = len(f_df)

    _hck_key   = f"hist_ck_{curr_g}_{curr_l}"
    _hsrch_cnt = f"hsrch_cnt_{curr_g}_{curr_l}"
    if _hck_key   not in st.session_state: st.session_state[_hck_key]   = {}
    if _hsrch_cnt not in st.session_state: st.session_state[_hsrch_cnt] = 0

    with st.expander(f" {curr_g} {curr_l} 이력" + (f"  ·  {_hist_cnt}건" if _hist_cnt else "  ·  없음"), expanded=_xp("chk_hist"), key="_xp_chk_hist"):
        if not f_df.empty:
            _hsrch_key = f"hsrch_{curr_g}_{curr_l}_{st.session_state[_hsrch_cnt]}"
            hs1, hs2 = st.columns([3, 3])
            _sn_search_qp = hs1.text_input(" 시리얼 스캔/검색",
                placeholder="스캔 또는 입력 → 자동 체크", key=_hsrch_key)
            if st.session_state.pop("_autofocus_after_rerun", None) == _hsrch_key:
                _inject_autofocus(placeholder="스캔 또는 입력 → 자동 체크")

            f_df_view = f_df
            if _sn_search_qp.strip():
                _search_result = f_df[f_df['시리얼'].str.contains(
                    _sn_search_qp.strip(), case=False, na=False)]
                if _search_result.empty:
                    hs1.warning(f"**'{_sn_search_qp.strip()}'** — 해당 시리얼이 없습니다.")
                else:
                    # 자동 체크는 처리 가능 상태(actionable)인 항목만 적용
                    for _hi in _search_result[_search_result['상태'].isin(["검사중","포장중","수리 완료(재투입)"])].index:
                        st.session_state[_hck_key][str(_hi)] = True
                    st.session_state["_autofocus_after_rerun"] = f"hsrch_{curr_g}_{curr_l}_{st.session_state[_hsrch_cnt] + 1}"
                    st.session_state[_hsrch_cnt] += 1  # 키 변경 → 입력 초기화 후 전체 목록 표시
                    _rerun("chk_hist")

            _h_checked = [k for k,v in st.session_state[_hck_key].items() if v]
            if _h_checked:
                btn_lbl = "검사 합격" if curr_l == "검사 라인" else "포장 완료"
                hba1, hba2, hba3, hba4, hba5 = st.columns([2, 1.5, 1.2, 1.0, 0.8])
                hba1.markdown(f"<span style='color:#2E75B6;font-weight:700;'> {len(_h_checked)}개 선택됨</span>",
                              unsafe_allow_html=True)
                _hist_bulk_ng_cause = hba2.selectbox("불량 원인", DEFECT_CAUSES, key=f"hist_bulk_ng_cause_{curr_g}_{curr_l}", label_visibility="collapsed")
                if _hist_bulk_ng_cause == "기타 (직접 입력)":
                    _hist_bulk_ng_cause_final = st.text_input("직접 입력", key=f"hist_bulk_ng_cause_txt_{curr_g}_{curr_l}", placeholder="불량 원인 직접 입력")
                else:
                    _hist_bulk_ng_cause_final = _hist_bulk_ng_cause
                if hba3.button(f" 일괄 {btn_lbl}", key=f"hist_bulk_ok_{curr_g}_{curr_l}",
                               type="primary", use_container_width=True,
                               disabled=not check_perm(f"{curr_l}::{curr_g}", "write")):
                    _ok_s  = 'OQC대기' if curr_l == '검사 라인' else '완료'
                    _prv_s = '검사중'  if curr_l == '검사 라인' else '포장중'
                    _ops = []
                    for ci in _h_checked:
                        ci_int = int(ci)
                        if ci_int in f_df.index:
                            _r = f_df.loc[ci_int]
                            if _r['상태'] in ["검사중","포장중","수리 완료(재투입)"]:
                                _upd = {'상태':_ok_s,'시간':get_now_kst_str()}
                                _ops.append({"sn": _r['시리얼'], "data": _upd,
                                    "audit": {"시리얼": _r['시리얼'], "모델": _r['모델'], "반": curr_g,
                                             "이전상태": _r['상태'], "이후상태": _ok_s,
                                             "작업자": st.session_state.user_id}})
                    st.session_state[_hck_key] = {}
                    st.session_state[_hsrch_cnt] += 1  # 체크박스 키 리셋
                    _prod_bulk_update(_run_bulk_db_ops(_ops))
                    _rerun("chk_hist")
                if hba4.button(" 일괄 불량", key=f"hist_bulk_ng_{curr_g}_{curr_l}",
                               use_container_width=True):
                    if _hist_bulk_ng_cause_final in ["(선택)", ""]:
                        st.warning(" 불량 원인을 먼저 선택해주세요.")
                    else:
                        _ops = []
                        for ci in _h_checked:
                            ci_int = int(ci)
                            if ci_int in f_df.index:
                                _r = f_df.loc[ci_int]
                                if _r['상태'] in ["검사중","포장중","수리 완료(재투입)"]:
                                    _upd = {'상태':'불량 처리 중','시간':get_now_kst_str(),
                                        '증상': f'불량입고출처: {curr_l} | 불량원인: {_hist_bulk_ng_cause_final}'}
                                    _ops.append({"sn": _r['시리얼'], "data": _upd,
                                        "audit": {"시리얼": _r['시리얼'], "모델": _r['모델'], "반": curr_g,
                                                 "이전상태": _r['상태'], "이후상태": '불량 처리 중',
                                                 "작업자": st.session_state.user_id}})
                        st.session_state[_hck_key] = {}
                        st.session_state[_hsrch_cnt] += 1  # 체크박스 키 리셋
                        _prod_bulk_update(_run_bulk_db_ops(_ops))
                        _rerun("chk_hist")
                if hba5.button("해제", key=f"hist_unck_{curr_g}_{curr_l}",
                               use_container_width=True, help="선택 해제"):
                    st.session_state[_hck_key] = {}
                    st.session_state[_hsrch_cnt] += 1  # 체크박스 키 리셋
                    _rerun("chk_hist")

            # Bug fix: STATUS_STYLE2는 전역 STATUS_STYLE과 중복 — 전역 상수 재사용
            STATUS_STYLE2 = STATUS_STYLE

            h = st.columns([0.4, 1.8, 1.8, 1.3, 1.6, 2.2])
            for col, txt in zip(h, ["","기록 시간","모델","품목","시리얼","상태"]):
                col.markdown(f"<p style='font-size:0.72rem;font-weight:700;color:#546e7a;margin:0;text-align:center;text-transform:uppercase;letter-spacing:0.05em;'>{txt}</p>",
                             unsafe_allow_html=True)

            _hcb_ver = st.session_state[_hsrch_cnt]
            # 자재 시리얼 일괄 조회 (N+1 방지)
            _hist_bulk_sns = tuple(f_df_view['시리얼'].unique().tolist())
            _hist_bulk_mats = load_material_serials_bulk(_hist_bulk_sns) if _hist_bulk_sns else pd.DataFrame()
            for row in f_df_view.sort_values('시간', ascending=False).reset_index().to_dict('records'):
                idx = row['index']
                is_act = row['상태'] in ["검사중","포장중","수리 완료(재투입)"]
                r = st.columns([0.4, 1.8, 1.8, 1.3, 1.6, 2.2])
                if is_act:
                    _hck = r[0].checkbox("", key=f"hck_{curr_g}_{curr_l}_{idx}_{_hcb_ver}",
                        value=st.session_state[_hck_key].get(str(idx), False),
                        label_visibility="collapsed")
                    st.session_state[_hck_key][str(idx)] = _hck
                else:
                    _hck = False
                r[1].caption(str(row['시간'])[:16])
                r[2].caption(row['모델'])
                r[3].caption(row['품목코드'])
                _hist_mc = len(_hist_bulk_mats[_hist_bulk_mats['메인시리얼'] == row['시리얼']]) if not _hist_bulk_mats.empty else 0
                _hist_tog_key = f"mat_tog_{row['시리얼']}_{curr_g}_{curr_l}"
                _hist_btn_lbl = f"{row['시리얼']}\n자재 {_hist_mc}개" if _hist_mc > 0 else row['시리얼']
                if r[4].button(_hist_btn_lbl, key=f"sntog_hist_{idx}_{_hcb_ver}",
                               use_container_width=True, help="클릭하여 자재 시리얼 조회"):
                    st.session_state[_hist_tog_key] = not st.session_state.get(_hist_tog_key, False)
                if is_act:
                    btn_lbl = "합격" if curr_l == "검사 라인" else "완료"
                    _act_b1, _act_b2 = r[5].columns(2)
                    if _act_b1.button(btn_lbl, key=f"ok_{idx}", use_container_width=True, type="primary"):
                        _ok_s  = 'OQC대기' if curr_l == '검사 라인' else '완료'
                        _prv_s = '검사중'  if curr_l == '검사 라인' else '포장중'
                        _upd = {'상태':_ok_s,'시간':get_now_kst_str()}
                        update_row(row['시리얼'], _upd)
                        insert_audit_log(시리얼=row['시리얼'], 모델=row['모델'], 반=curr_g,
                            이전상태=row['상태'], 이후상태=_ok_s, 작업자=st.session_state.user_id)
                        st.session_state[_hck_key].pop(str(idx), None)
                        st.session_state[_hsrch_cnt] += 1  # 체크박스 키 리셋
                        _prod_update(row['시리얼'], _upd)
                        _rerun("chk_hist")
                    _hist_ng_open_key = f"_ng_open_hist_{idx}"
                    if not st.session_state.get(_hist_ng_open_key):
                        if _act_b2.button("불량", key=f"ng_{idx}", use_container_width=True):
                            st.session_state[_hist_ng_open_key] = True
                            _rerun("chk_hist")
                    else:
                        _act_b2.button("불량", key=f"ng_{idx}", use_container_width=True, disabled=True)
                else:
                    s2 = row['상태']
                    s2_esc = html_mod.escape(str(s2))
                    if "불량" in str(s2):
                        r[5].markdown(f"<div style='background:#fde8e7;color:#7a2e2a;padding:3px 8px;border-radius:3px;text-align:center;font-weight:700;font-size:0.75rem;border:1px solid #e87e7a;'>{s2_esc}</div>", unsafe_allow_html=True)
                    else:
                        bg2,tc2,bc2,ic2 = STATUS_STYLE2.get(s2, ('#eceff1','#546e7a','#90a4ae',''))
                        s2_esc = html_mod.escape(str(s2))
                        r[5].markdown(f"<div style='background:{bg2};color:{tc2};padding:3px 8px;border-radius:3px;text-align:center;font-weight:700;border:1px solid {bc2};font-size:0.75rem;'>{s2_esc}</div>", unsafe_allow_html=True)
                # ── 불량 원인 선택 패널 (개별  클릭 후) ──
                _hist_ng_open_key = f"_ng_open_hist_{idx}"
                if st.session_state.get(_hist_ng_open_key):
                    with st.container(border=True):
                        st.caption(f" 불량 원인 입력 — `{row['시리얼']}`")
                        _hnc1, _hnc2, _hnc3 = st.columns([2.5, 1, 1])
                        _hist_cause_ng = _hnc1.selectbox("불량 원인", DEFECT_CAUSES, key=f"_ng_cause_hist_{idx}", label_visibility="collapsed")
                        if _hist_cause_ng == "기타 (직접 입력)":
                            _hist_cause_ng_final = st.text_input("직접 입력", key=f"_ng_cause_hist_txt_{idx}", placeholder="불량 원인 직접 입력")
                        else:
                            _hist_cause_ng_final = _hist_cause_ng
                        if _hnc2.button("확정", key=f"_ng_confirm_hist_{idx}", type="primary", use_container_width=True):
                            if _hist_cause_ng_final in ["(선택)", ""]:
                                st.warning(" 불량 원인을 선택해주세요.")
                            else:
                                _upd = {'상태':'불량 처리 중','시간':get_now_kst_str(),
                                    '증상': f'불량입고출처: {curr_l} | 불량원인: {_hist_cause_ng_final}'}
                                update_row(row['시리얼'], _upd)
                                insert_audit_log(시리얼=row['시리얼'], 모델=row['모델'], 반=curr_g,
                                    이전상태=row['상태'], 이후상태='불량 처리 중', 작업자=st.session_state.user_id)
                                st.session_state[_hck_key].pop(str(idx), None)
                                st.session_state[_hsrch_cnt] += 1
                                st.session_state.pop(_hist_ng_open_key, None)
                                _prod_update(row['시리얼'], _upd)
                                _rerun("chk_hist")
                        if _hnc3.button("취소", key=f"_ng_cancel_hist_{idx}", use_container_width=True):
                            st.session_state.pop(_hist_ng_open_key, None)
                            _rerun("chk_hist")
                # ── 자재 시리얼 토글 표시 ──
                if st.session_state.get(_hist_tog_key, False):
                    _hist_row_mats = _hist_bulk_mats[_hist_bulk_mats['메인시리얼'] == row['시리얼']] if not _hist_bulk_mats.empty else pd.DataFrame()
                    with st.container(border=True):
                        st.caption(f" 자재 시리얼 — `{row['시리얼']}`")
                        if not _hist_row_mats.empty:
                            for _hm in _hist_row_mats.to_dict('records'):
                                hmc1, hmc2 = st.columns([2, 4])
                                hmc1.caption(_hm.get('자재명', ''))
                                hmc2.caption(f"`{_hm.get('자재시리얼', '')}`")
                        else:
                            st.caption("등록된 자재 시리얼 없음")
                        # ── 라벨 시리얼 (완료 항목) ──
                        if row.get('상태') == '완료':
                            st.divider()
                            _cur_lsn = str(row.get('라벨시리얼', '') or '')
                            _lsn_edit_key = f"hist_lsn_{row['시리얼']}"
                            _lc1, _lc2 = st.columns([4, 1.5])
                            _lc1.markdown("<p style='font-size:0.78rem;font-weight:700;color:#5a4f45;margin:0 0 4px 0;'> 라벨 S/N</p>", unsafe_allow_html=True)
                            _new_lsn = _lc1.text_input(
                                "라벨 S/N",
                                value=_cur_lsn,
                                placeholder="라벨 시리얼 입력",
                                key=_lsn_edit_key,
                                label_visibility="collapsed",
                            )
                            if _lc2.button(" 저장", key=f"hist_lsn_save_{row['시리얼']}",
                                           use_container_width=True, type="primary",
                                           disabled=not bool(_new_lsn.strip())):
                                update_row(row['시리얼'], {'라벨시리얼': _new_lsn.strip()})
                                insert_audit_log(시리얼=row['시리얼'], 모델=row['모델'], 반=curr_g,
                                    이전상태='완료', 이후상태='완료', 작업자=st.session_state.user_id,
                                    비고=f"라벨시리얼 수정: {_new_lsn.strip()}")
                                _prod_update(row['시리얼'], {'라벨시리얼': _new_lsn.strip()})
                                st.rerun()
        else:
            st.info("해당 공정 내역이 없습니다.")

    # ── 기존 제품 자재 시리얼 조회 / 추가 ────────────────────────────
    MAT_NAME_OPTIONS_QL = st.session_state.get("dropdown_mat_name") or []
    with st.expander(" 자재 시리얼 조회 / 추가 등록", expanded=_xp("chk_mat"), key="_xp_chk_mat"):
        st.caption("메인 S/N을 입력하면 등록된 자재를 조회하고 누락된 자재를 추가할 수 있습니다.")

        _ql_sn_cnt_key = f"ql_sn_cnt_{curr_g}_{curr_l}"
        if _ql_sn_cnt_key not in st.session_state:
            st.session_state[_ql_sn_cnt_key] = 0
        ql_main_sn = st.text_input(" 메인 S/N 조회", placeholder="기존 등록된 메인 S/N 입력",
                                   key=f"ql_main_sn_{curr_g}_{curr_l}_{st.session_state[_ql_sn_cnt_key]}")

        if ql_main_sn.strip():
            _ql_exist = db_s[db_s['시리얼'] == ql_main_sn.strip()]
            if _ql_exist.empty:
                st.warning(f" 등록되지 않은 S/N입니다: **{ql_main_sn.strip()}**")
            else:
                _ql_er = _ql_exist.iloc[0]
                st.success(f" **{_ql_er['모델']}** `{_ql_er['품목코드']}` — 상태: {_ql_er['상태']}")

                _ql_mats = load_material_serials(ql_main_sn.strip())
                if not _ql_mats.empty:
                    st.markdown(f"<p style='font-size:0.78rem;color:#8a7f72;margin:6px 0 4px 0;'>기존 등록 자재: <b>{len(_ql_mats)}개</b></p>", unsafe_allow_html=True)
                    for _qm in _ql_mats.to_dict('records'):
                        qmc1, qmc2 = st.columns([2, 4])
                        qmc1.caption(_qm.get('자재명', ''))
                        qmc2.caption(f"`{_qm.get('자재시리얼', '')}`")
                else:
                    st.info("기존 등록된 자재 시리얼이 없습니다.")

                st.divider()
                st.markdown("<p style='font-size:0.85rem;font-weight:700;color:#5a4f45;margin:0 0 6px 0;'> 자재 추가 등록</p>", unsafe_allow_html=True)

                _ql_mat_list_key = f"ql_mat_list_{curr_g}_{curr_l}"
                if _ql_mat_list_key not in st.session_state:
                    st.session_state[_ql_mat_list_key] = []
                _ql_scan_cnt_key = f"ql_scan_cnt_{curr_g}_{curr_l}"
                if _ql_scan_cnt_key not in st.session_state:
                    st.session_state[_ql_scan_cnt_key] = 0

                qsc1, qsc2, qsc3 = st.columns([2, 3, 1])
                if MAT_NAME_OPTIONS_QL:
                    ql_sel_mat = qsc1.selectbox("자재명 선택", MAT_NAME_OPTIONS_QL,
                                                key=f"ql_mat_nm_sel_{curr_g}_{curr_l}")
                else:
                    ql_sel_mat = qsc1.text_input("자재명 입력", placeholder="예: PCB, 배터리",
                                                 key=f"ql_mat_nm_txt_{curr_g}_{curr_l}")
                _ql_scan_field_key = f"ql_scan_{curr_g}_{curr_l}_{st.session_state[_ql_scan_cnt_key]}"
                ql_scan_input = qsc2.text_input("자재 S/N 스캔",
                    placeholder="바코드 스캔 → 자동 추가 (Enter)",
                    key=_ql_scan_field_key)
                if st.session_state.pop("_autofocus_after_rerun", None) == _ql_scan_field_key:
                    _inject_autofocus("자재 S/N 스캔")
                qsc2.caption(" 스캐너로 스캔하면 Enter가 자동 입력됩니다")

                if ql_scan_input.strip():
                    if not any(m["자재시리얼"] == ql_scan_input.strip() for m in st.session_state[_ql_mat_list_key]):
                        st.session_state[_ql_mat_list_key].append({"자재명": ql_sel_mat, "자재시리얼": ql_scan_input.strip()})
                    else:
                        st.toast(f" 이미 추가된 S/N: {ql_scan_input.strip()}")
                    st.session_state["_autofocus_after_rerun"] = f"ql_scan_{curr_g}_{curr_l}_{st.session_state[_ql_scan_cnt_key] + 1}"
                    st.session_state[_ql_scan_cnt_key] += 1
                    _rerun("chk_mat")

                if qsc3.button(" 추가", key=f"ql_mat_add_{curr_g}_{curr_l}", use_container_width=True):
                    st.session_state[_ql_mat_list_key].append({"자재명": ql_sel_mat, "자재시리얼": ""})
                    _rerun("chk_mat")

                ql_mat_list_now = st.session_state[_ql_mat_list_key]
                if ql_mat_list_now:
                    st.markdown(f"<p style='font-size:0.78rem;color:#8a7f72;margin:6px 0 2px 0;'>추가 예정: <b>{len(ql_mat_list_now)}개</b></p>", unsafe_allow_html=True)
                    ql_updated = []
                    _ql_rerun = False
                    for qi, qmat in enumerate(ql_mat_list_now):
                        qlc1, qlc2, qlc3 = st.columns([2, 4, 1])
                        if MAT_NAME_OPTIONS_QL:
                            qn = qlc1.selectbox("", MAT_NAME_OPTIONS_QL,
                                index=MAT_NAME_OPTIONS_QL.index(qmat["자재명"]) if qmat["자재명"] in MAT_NAME_OPTIONS_QL else 0,
                                key=f"ql_nm_{curr_g}_{curr_l}_{qi}", label_visibility="collapsed")
                        else:
                            qn = qlc1.text_input("", value=qmat["자재명"],
                                key=f"ql_nm_txt_{curr_g}_{curr_l}_{qi}", label_visibility="collapsed")
                        qs = qlc2.text_input("", value=qmat["자재시리얼"],
                            key=f"ql_sv_{curr_g}_{curr_l}_{qi}", label_visibility="collapsed",
                            placeholder="S/N 직접 입력 또는 스캔")
                        if not qlc3.button("삭제", key=f"ql_del_{curr_g}_{curr_l}_{qi}", help="삭제"):
                            ql_updated.append({"자재명": qn, "자재시리얼": qs})
                        else:
                            _ql_rerun = True
                    st.session_state[_ql_mat_list_key] = ql_updated
                    if _ql_rerun:
                        _rerun("chk_mat")

                    _ql_btn_col, _ = st.columns([1, 2])
                    if _ql_btn_col.button(" 자재 시리얼 추가 저장", key=f"ql_save_{curr_g}_{curr_l}",
                                 type="primary", use_container_width=True):
                        valid_ql_mats = [m for m in st.session_state[_ql_mat_list_key] if m["자재시리얼"].strip()]
                        if valid_ql_mats:
                            if insert_material_serials(메인시리얼=ql_main_sn.strip(),
                                    모델=_ql_er['모델'], 반=curr_g,
                                    자재목록=valid_ql_mats, 작업자=st.session_state.user_id):
                                st.session_state[_ql_mat_list_key] = []
                                st.session_state[_ql_sn_cnt_key] += 1
                                st.session_state[_ql_scan_cnt_key] = 0
                                st.cache_data.clear()
                                st.toast(f" {len(valid_ql_mats)}개 자재 시리얼 추가 완료")
                                _rerun("chk_mat")
                        else:
                            st.warning("추가할 자재 시리얼을 입력해주세요.")
                else:
                    st.caption("자재 없음 — 스캔하거나  추가 버튼을 누르세요")

elif curr_l == "생산 현황 리포트":
    st.markdown("<h2 class='centered-title'> 생산 현황 리포트</h2>", unsafe_allow_html=True)

    _rpt_top1, _rpt_top2 = st.columns([2, 3])
    with _rpt_top1:
        v_group = st.radio("조회 범위", ["전체"] + PRODUCTION_GROUPS, horizontal=True, key="prod_report_grp")
    with _rpt_top2:
        _rpt_range = st.date_input(
            "조회 기간",
            value=(date.today() - timedelta(days=30), date.today()),
            key="prod_rpt_date_range"
        )
    if isinstance(_rpt_range, (list, tuple)) and len(_rpt_range) == 2:
        df_rpt = load_production_history(str(_rpt_range[0]), str(_rpt_range[1]))
    else:
        df_rpt = load_production_history(str(date.today()), str(date.today()))
    if v_group != "전체":
        df_rpt = df_rpt[df_rpt['반'] == v_group]

    if not df_rpt.empty:
        # ── KPI ──────────────────────────────────────────────────────
        _rpt_done    = df_rpt[(df_rpt['라인'] == '포장 라인') & (df_rpt['상태'] == '완료')]
        _rpt_ing     = df_rpt[df_rpt['상태'].isin(ACTIVE_STATES)]
        _rpt_defect  = df_rpt[df_rpt['상태'].str.contains('불량|부적합', na=False)]
        kp1, kp2, kp3, kp4 = st.columns(4)
        kp1.metric(" 총 투입",      f"{len(df_rpt)} EA")
        kp2.metric(" 최종 완료",    f"{len(_rpt_done)} EA")
        kp3.metric(" 진행 중",     f"{len(_rpt_ing)} EA")
        kp4.metric(" 불량/부적합", f"{len(_rpt_defect)} 건")
        st.divider()

        # ── 차트 행 1: 상태별 분포 + 모델별 비중 ─────────────────────
        cc1, cc2 = st.columns([1.8, 1.2])
        with cc1:
            _st_cnt = df_rpt.groupby('상태').size().reset_index(name='수량').sort_values('수량', ascending=False)
            _fig_st = px.bar(_st_cnt, x='상태', y='수량', color='상태',
                             title="<b>현재 상태별 제품 분포</b>", template="plotly_white",
                             text='수량')
            _fig_st.update_traces(textposition='outside', textfont_size=11)
            _fig_st.update_layout(showlegend=False, margin=dict(t=50, b=20))
            _fig_st.update_yaxes(dtick=5)
            st.plotly_chart(_fig_st, use_container_width=True)
        with cc2:
            _md_cnt = df_rpt.groupby('모델').size().reset_index(name='수량')
            _fig_md = px.pie(_md_cnt, values='수량', names='모델', hole=0.45,
                             title="<b>모델별 생산 비중</b>")
            _fig_md.update_layout(margin=dict(t=40, b=20))
            st.plotly_chart(_fig_md, use_container_width=True)

        # ── 차트 행 2: 반별/공정별 완료 + 일자별 투입 추이 ───────────
        cc3, cc4 = st.columns([1.2, 1.8])
        with cc3:
            if v_group == "전체":
                _ban_done = df_rpt[df_rpt['상태'] == '완료'].groupby('반').size().reset_index(name='완료')
                if not _ban_done.empty:
                    _fig_ban = px.bar(_ban_done, x='반', y='완료', color='반',
                                      title="<b>반별 완료 수량</b>", template="plotly_white")
                    _fig_ban.update_layout(showlegend=False, margin=dict(t=40, b=20))
                    st.plotly_chart(_fig_ban, use_container_width=True)
                else:
                    st.info("완료 데이터 없음")
            else:
                _ln_cnt = df_rpt.groupby('라인').size().reset_index(name='수량')
                _fig_ln = px.bar(_ln_cnt, x='라인', y='수량', color='라인',
                                 title=f"<b>{v_group} 공정별 현황</b>", template="plotly_white")
                _fig_ln.update_layout(showlegend=False, margin=dict(t=40, b=20))
                _fig_ln.update_yaxes(dtick=1)
                st.plotly_chart(_fig_ln, use_container_width=True)
        with cc4:
            try:
                # 투입 현황은 audit_log의 최초 등록(이전상태='-', 이후상태='조립중') 기준으로 집계
                # → production_db의 '시간'(마지막 상태변경 시간)을 쓰면 오늘 처리된
                #   검사/OQC/포장 제품까지 모두 포함되어 수치가 과도하게 집계되는 문제 방지
                _audit_all = load_audit_log()
                if v_group != "전체" and not _audit_all.empty:
                    _audit_all = _audit_all[_audit_all['반'] == v_group]
                if not _audit_all.empty and '이전상태' in _audit_all.columns and '이후상태' in _audit_all.columns:
                    _df_trend = _audit_all[
                        (_audit_all['이전상태'] == '-') & (_audit_all['이후상태'] == '조립중')
                    ].copy()
                else:
                    _df_trend = pd.DataFrame(columns=['시간'])
                _parsed = pd.to_datetime(_df_trend['시간'], errors='coerce')
                # timezone-aware면 KST로 변환, naive면 KST로 간주
                if not _parsed.empty and _parsed.dt.tz is not None:
                    _df_trend['_kst'] = _parsed.dt.tz_convert('Asia/Seoul')
                else:
                    _df_trend['_kst'] = _parsed
                _df_trend['_date'] = _df_trend['_kst'].dt.strftime('%Y-%m-%d')
                _df_trend['_hhmm'] = _df_trend['_kst'].dt.hour * 60 + _df_trend['_kst'].dt.minute

                # 오늘 날짜 기준으로 데이터 선택 → 없으면 가장 최근 근무일
                _today_str = get_now_kst_str()[:10]
                _today_data = _df_trend[_df_trend['_date'] == _today_str]
                if _today_data.empty:
                    # 오늘 데이터 없으면 가장 최근 날짜
                    _latest_date = _df_trend['_date'].dropna().max() if not _df_trend.empty else None
                    if _latest_date:
                        _today_data = _df_trend[_df_trend['_date'] == _latest_date]
                        _chart_date = _latest_date
                    else:
                        _chart_date = _today_str
                else:
                    _chart_date = _today_str

                # 근무시간 슬롯 필터: 08:30(510분) ~ 17:30(1050분)
                _work = _today_data[
                    _today_data['_hhmm'].notna() &
                    (_today_data['_hhmm'] >= 510) &
                    (_today_data['_hhmm'] <= 1050)
                ].copy()

                # 30분 단위 버킷 — 전체 근무 슬롯 고정 표시
                _slots = pd.date_range('2000-01-01 08:30', '2000-01-01 17:30', freq='30min').strftime('%H:%M').tolist()
                if not _work.empty:
                    _work['시간대'] = _work['_kst'].dt.floor('30min').dt.strftime('%H:%M')
                    _hourly = _work.groupby('시간대').size().reset_index(name='수량')
                else:
                    _hourly = pd.DataFrame(columns=['시간대', '수량'])
                _hourly = pd.DataFrame({'시간대': _slots}).merge(_hourly, on='시간대', how='left').fillna(0)
                _hourly['수량'] = _hourly['수량'].astype(int)

                _fig_tr = px.bar(_hourly, x='시간대', y='수량',
                                 title=f"<b>근무시간대별 생산 투입 현황</b> ({_chart_date})",
                                 template="plotly_white", text='수량')
                _fig_tr.update_traces(marker_color='#2471a3', textposition='outside', textfont_size=9)
                _fig_tr.update_xaxes(tickangle=-45)
                _fig_tr.update_yaxes(dtick=5)
                _fig_tr.update_layout(margin=dict(t=50, b=60))
                st.plotly_chart(_fig_tr, use_container_width=True)
            except Exception:
                st.info("추이 차트 데이터 처리 중 오류")

        st.divider()

        # ── 이력 테이블 ───────────────────────────────────────────────
        with st.expander(f" 전체 이력 테이블  ·  {len(df_rpt)}건", expanded=_xp("rpt_tbl"), key="_xp_rpt_tbl"):
            _RPT_PAGE_SIZE = 50
            _rpt_sorted = df_rpt.sort_values('시간', ascending=False).reset_index(drop=True)
            _rpt_total = len(_rpt_sorted)
            _rpt_total_pages = max(1, (_rpt_total + _RPT_PAGE_SIZE - 1) // _RPT_PAGE_SIZE)
            if "prod_rpt_page" not in st.session_state:
                st.session_state["prod_rpt_page"] = 1
            _pr_page = st.session_state["prod_rpt_page"]

            _pr1, _pr2, _pr3 = st.columns([1, 2, 1])
            if _pr1.button("◀ 이전", key="pr_prev", disabled=(_pr_page <= 1)):
                st.session_state["prod_rpt_page"] -= 1; _rerun("rpt_tbl")
            _pr2.markdown(
                f"<p style='text-align:center;font-size:0.82rem;color:#8a7f72;margin:6px 0;'>"
                f"페이지 <b>{_pr_page}</b> / {_rpt_total_pages}　"
                f"(전체 <b>{_rpt_total:,}</b>건, {_RPT_PAGE_SIZE}건/페이지)</p>",
                unsafe_allow_html=True)
            if _pr3.button("다음 ▶", key="pr_next", disabled=(_pr_page >= _rpt_total_pages)):
                st.session_state["prod_rpt_page"] += 1; _rerun("rpt_tbl")

            _pr_start = (_pr_page - 1) * _RPT_PAGE_SIZE
            st.dataframe(_rpt_sorted.iloc[_pr_start:_pr_start + _RPT_PAGE_SIZE],
                         use_container_width=True, hide_index=True)
    else:
        st.info("조회 가능한 데이터가 없습니다.")

# ── 검사 라인 ────────────────────────────────────────────────────
elif curr_l == "검사 라인":
    st.markdown(f"<h2 class='centered-title'> {curr_g} 검사 라인 현황</h2>", unsafe_allow_html=True)

    db_qc_all = st.session_state.production_db.copy()
    db_qc     = db_qc_all[db_qc_all['반'] == curr_g]
    DEFECT_CAUSES = st.session_state.get('dropdown_defect_cause', ['(선택)', '기타 (직접 입력)'])

    # ── KPI ─────────────────────────────────────────────────────────
    qc_wait = len(db_qc[db_qc['상태'].isin(['검사대기', '수리 완료(재투입)'])])
    qc_ing  = len(db_qc[db_qc['상태'] == '검사중'])
    qc_pass = len(db_qc[db_qc['상태'] == 'OQC대기'])
    qc_ng   = len(db_qc[db_qc['상태'] == '불량 처리 중'])

    k1, k2, k3, k4 = st.columns(4)
    k1.metric(" 검사 대기", f"{qc_wait}건")
    k2.metric(" 검사 중",   f"{qc_ing}건")
    k3.metric(" OQC 대기",  f"{qc_pass}건")
    k4.metric(" 불량",      f"{qc_ng}건")
    st.divider()

    # ── 검사 대기 목록 ───────────────────────────────────────────────
    st.markdown("<div class='section-title'> 검사 대기</div>", unsafe_allow_html=True)
    wait_df = db_qc[db_qc['상태'].isin(['검사대기', '수리 완료(재투입)'])].sort_values('시간', ascending=False)

    if not wait_df.empty:
        hh = st.columns([2, 2, 2, 1.5])
        for col, txt in zip(hh, ["시간", "모델", "시리얼", "검사 시작"]):
            col.markdown(f"<p style='font-size:0.72rem;font-weight:700;color:#8a7f72;margin:0;padding-bottom:3px;border-bottom:1px solid #e0d8c8;'>{txt}</p>", unsafe_allow_html=True)
        # 성능: iterrows → enumerate + to_dict('records')
        for idx, row in enumerate(wait_df.to_dict('records')):
            rr = st.columns([2, 2, 2, 1.5])
            rr[0].caption(str(row.get('시간', ''))[:16])
            rr[1].write(row.get('모델', ''))
            rr[2].markdown(f"`{row.get('시리얼', '')}`")
            if rr[3].button("▶ 검사 시작", key=f"qc_in_{idx}", use_container_width=True, type="primary"):
                _upd = {'상태': '검사중', '시간': get_now_kst_str(),
                    '라인': '검사 라인', '작업자': st.session_state.user_id}
                update_row(row['시리얼'], _upd)
                insert_audit_log(시리얼=row['시리얼'], 모델=row['모델'], 반=curr_g,
                    이전상태=row['상태'], 이후상태='검사중', 작업자=st.session_state.user_id)
                _prod_update(row['시리얼'], _upd)
                st.rerun()
    else:
        st.info("검사 대기 중인 제품이 없습니다.")

    st.divider()

    # ── 검사 중 → 판정 ───────────────────────────────────────────────
    st.markdown("<div class='section-title'> 검사 진행 중</div>", unsafe_allow_html=True)
    qc_ing_df = db_qc[db_qc['상태'] == '검사중'].sort_values('시간', ascending=False)

    if not qc_ing_df.empty:
        # 성능: iterrows → enumerate + to_dict('records')
        for idx, row in enumerate(qc_ing_df.to_dict('records')):
            with st.container(border=True):
                ic1, ic2, ic3 = st.columns([2, 2, 1.5])
                ic1.markdown(f"**{row.get('모델', '')}**")
                ic2.markdown(f"`{row.get('시리얼', '')}`")
                ic3.markdown("<span style='background:#ddeeff;color:#1a4a7a;padding:2px 8px;"
                             "border-radius:6px;font-size:0.8rem;font-weight:bold;'> 검사중</span>",
                             unsafe_allow_html=True)

                qc1, qc2 = st.columns([2, 1])
                with qc1:
                    cause_sel = st.selectbox("불량 원인 (불량 처리 시 선택)",
                        DEFECT_CAUSES, key=f"qc_cause_{idx}")
                    if cause_sel == "기타 (직접 입력)":
                        cause_txt = st.text_input("직접 입력", key=f"qc_cause_txt_{idx}", placeholder="불량 원인 입력")
                    elif cause_sel == "(선택)":
                        cause_txt = ""
                    else:
                        cause_txt = cause_sel
                with qc2:
                    btn_pass = st.button(" 합격 (OQC 대기)", key=f"qc_ok_{idx}", use_container_width=True, type="primary")
                    btn_ng   = st.button(" 불량 처리",       key=f"qc_ng_{idx}",  use_container_width=True)

                if btn_pass:
                    _upd = {'상태': 'OQC대기', '시간': get_now_kst_str()}
                    update_row(row['시리얼'], _upd)
                    insert_audit_log(시리얼=row['시리얼'], 모델=row['모델'], 반=curr_g,
                        이전상태='검사중', 이후상태='OQC대기', 작업자=st.session_state.user_id)
                    _prod_update(row['시리얼'], _upd)
                    st.rerun()
                if btn_ng:
                    if not cause_txt:
                        st.warning(" 불량 원인을 먼저 선택해주세요.")
                    else:
                        _upd = {'상태': '불량 처리 중', '시간': get_now_kst_str(), '증상': cause_txt}
                        update_row(row['시리얼'], _upd)
                        insert_audit_log(시리얼=row['시리얼'], 모델=row['모델'], 반=curr_g,
                            이전상태='검사중', 이후상태='불량 처리 중',
                            작업자=st.session_state.user_id, 비고=f"원인:{cause_txt}")
                        _prod_update(row['시리얼'], _upd)
                        st.rerun()
    else:
        st.info("검사 진행 중인 제품이 없습니다.")

    st.divider()

    _qc_hist_total = len(db_qc[db_qc['라인'] == '검사 라인'])
    with st.expander(f" 검사 이력  ·  전체 {_qc_hist_total}건", expanded=_xp("qc_hist"), key="_xp_qc_hist"):
        _qc_h1, _qc_h2 = st.columns([3, 1])
        _qc_drange = _qc_h1.date_input(
            "조회 기간",
            value=(date.today() - timedelta(days=30), date.today()),
            key="qc_hist_date_range"
        )
        _qc_state_f = _qc_h2.selectbox("상태 필터", ["전체", "검사대기", "검사중", "불량 처리 중"], key="qc_hist_state")
        if isinstance(_qc_drange, (list, tuple)) and len(_qc_drange) == 2:
            hist = load_production_history(str(_qc_drange[0]), str(_qc_drange[1]))
        else:
            hist = load_production_history(str(date.today()), str(date.today()))
        hist = hist[hist['라인'] == '검사 라인']
        if _qc_state_f != "전체":
            hist = hist[hist['상태'] == _qc_state_f]
        hist = hist.sort_values('시간', ascending=False)
        st.caption(f"조회 결과: {len(hist)}건")
        if not hist.empty:
            st.dataframe(hist[['시간', '모델', '시리얼', '상태', '증상', '작업자']].reset_index(drop=True),
                         use_container_width=True, hide_index=True)
        else:
            st.info("이력이 없습니다.")

# ── 포장 라인 ────────────────────────────────────────────────────
elif curr_l == "포장 라인":
    st.markdown(f"<h2 class='centered-title'> {curr_g} 포장 라인 현황</h2>", unsafe_allow_html=True)

    db_pk_all = st.session_state.production_db.copy()
    db_pk     = db_pk_all[db_pk_all['반'] == curr_g]

    # ── KPI ─────────────────────────────────────────────────────────
    pk_wait = len(db_pk[db_pk['상태'] == '출하승인'])
    pk_ing  = len(db_pk[db_pk['상태'] == '포장중'])
    pk_done = len(db_pk[db_pk['상태'] == '완료'])
    pk_ng   = len(db_pk[db_pk['상태'].str.contains('불량', na=False)])

    k1, k2, k3, k4 = st.columns(4)
    k1.metric(" 포장 대기", f"{pk_wait}건")
    k2.metric(" 포장 중",   f"{pk_ing}건")
    k3.metric(" 완료",      f"{pk_done}건")
    k4.metric(" 불량",      f"{pk_ng}건")
    st.divider()

    # ── 포장 대기 (OQC 합격 → 출하승인 상태) ────────────────────────
    st.markdown("<div class='section-title'> 포장 대기 (OQC 합격 제품)</div>", unsafe_allow_html=True)
    pk_wait_df = db_pk[db_pk['상태'] == '출하승인'].sort_values('시간', ascending=False)

    if not pk_wait_df.empty:
        hh = st.columns([2, 2, 2, 1.5])
        for col, txt in zip(hh, ["시간", "모델", "시리얼", "포장 시작"]):
            col.markdown(f"<p style='font-size:0.72rem;font-weight:700;color:#8a7f72;margin:0;padding-bottom:3px;border-bottom:1px solid #e0d8c8;'>{txt}</p>", unsafe_allow_html=True)
        # 성능: iterrows → enumerate + to_dict('records')
        for idx, row in enumerate(pk_wait_df.to_dict('records')):
            rr = st.columns([2, 2, 2, 1.5])
            rr[0].caption(str(row.get('시간', ''))[:16])
            rr[1].write(row.get('모델', ''))
            rr[2].markdown(f"`{row.get('시리얼', '')}`")
            if rr[3].button("▶ 포장 시작", key=f"pk_in_{idx}", use_container_width=True, type="primary"):
                _upd = {'상태': '포장중', '시간': get_now_kst_str(),
                    '라인': '포장 라인', '작업자': st.session_state.user_id}
                update_row(row['시리얼'], _upd)
                insert_audit_log(시리얼=row['시리얼'], 모델=row['모델'], 반=curr_g,
                    이전상태='출하승인', 이후상태='포장중', 작업자=st.session_state.user_id)
                _prod_update(row['시리얼'], _upd)
                st.rerun()
    else:
        st.info("포장 대기 중인 제품이 없습니다.")

    st.divider()

    # ── 포장 중 → 완료 ───────────────────────────────────────────────
    st.markdown("<div class='section-title'> 포장 진행 중</div>", unsafe_allow_html=True)
    pk_ing_df = db_pk[db_pk['상태'] == '포장중'].sort_values('시간', ascending=False)

    if not pk_ing_df.empty:
        hh = st.columns([1.5, 1.5, 2, 2.5, 1.5])
        for col, txt in zip(hh, ["시간", "모델", "제품 시리얼", "라벨 S/N 스캔", "포장 완료"]):
            col.markdown(f"<p style='font-size:0.72rem;font-weight:700;color:#8a7f72;margin:0;padding-bottom:3px;border-bottom:1px solid #e0d8c8;'>{txt}</p>", unsafe_allow_html=True)
        # 성능: iterrows → enumerate + to_dict('records')
        for idx, row in enumerate(pk_ing_df.to_dict('records')):
            rr = st.columns([1.5, 1.5, 2, 2.5, 1.5])
            rr[0].caption(str(row.get('시간', ''))[:16])
            rr[1].write(row.get('모델', ''))
            rr[2].markdown(f"`{row.get('시리얼', '')}`")
            _label_key = f"pk_label_{row['시리얼']}"
            label_sn = rr[3].text_input(
                "라벨 S/N",
                placeholder="바코드 스캔 또는 직접 입력",
                key=_label_key,
                label_visibility="collapsed",
            )
            _btn_disabled = not bool(label_sn.strip())
            if rr[4].button(" 완료", key=f"pk_done_{idx}", use_container_width=True,
                            type="primary", disabled=_btn_disabled):
                _upd = {'상태': '완료', '시간': get_now_kst_str(), '라벨시리얼': label_sn.strip()}
                update_row(row['시리얼'], _upd)
                insert_audit_log(시리얼=row['시리얼'], 모델=row['모델'], 반=curr_g,
                    이전상태='포장중', 이후상태='완료', 작업자=st.session_state.user_id)
                _prod_update(row['시리얼'], _upd)
                st.rerun()
    else:
        st.info("포장 진행 중인 제품이 없습니다.")

    st.divider()

    # ── 라벨 시리얼 누락 항목 ────────────────────────────────────────
    _pk_no_label = db_pk[
        (db_pk['상태'] == '완료') &
        (db_pk['라벨시리얼'].isna() | (db_pk['라벨시리얼'].astype(str).str.strip() == ''))
    ].sort_values('시간', ascending=False)

    if not _pk_no_label.empty:
        _nl_count = len(_pk_no_label)
        with st.expander(f" 라벨 시리얼 누락  ·  {_nl_count}건", expanded=True, key="_xp_pk_nolabel"):
            hh = st.columns([1.5, 1.5, 2, 2.5, 1.5])
            for col, txt in zip(hh, ["시간", "모델", "제품 시리얼", "라벨 S/N 입력", "저장"]):
                col.markdown(f"<p style='font-size:0.72rem;font-weight:700;color:#8a7f72;margin:0;padding-bottom:3px;border-bottom:1px solid #e0d8c8;'>{txt}</p>", unsafe_allow_html=True)
            for idx, row in enumerate(_pk_no_label.to_dict('records')):
                rr = st.columns([1.5, 1.5, 2, 2.5, 1.5])
                rr[0].caption(str(row.get('시간', ''))[:16])
                rr[1].write(row.get('모델', ''))
                rr[2].markdown(f"`{row.get('시리얼', '')}`")
                _add_label_key = f"pk_add_label_{row['시리얼']}"
                add_label_sn = rr[3].text_input(
                    "라벨 S/N",
                    placeholder="바코드 스캔 또는 직접 입력",
                    key=_add_label_key,
                    label_visibility="collapsed",
                )
                if rr[4].button(" 저장", key=f"pk_label_save_{idx}", use_container_width=True,
                                disabled=not bool(add_label_sn.strip())):
                    update_row(row['시리얼'], {'라벨시리얼': add_label_sn.strip()})
                    insert_audit_log(시리얼=row['시리얼'], 모델=row['모델'], 반=curr_g,
                        이전상태='완료', 이후상태='완료', 작업자=st.session_state.user_id,
                        비고=f"라벨시리얼 추가: {add_label_sn.strip()}")
                    _prod_update(row['시리얼'], {'라벨시리얼': add_label_sn.strip()})
                    st.rerun()
        st.divider()

    _pk_done_total = len(db_pk[db_pk['상태'] == '완료'])
    with st.expander(f" 완료 이력  ·  전체 {_pk_done_total}건", expanded=_xp("pk_hist"), key="_xp_pk_hist"):
        _pk_h1, _pk_h2 = st.columns([3, 1])
        _pk_drange = _pk_h1.date_input(
            "조회 기간",
            value=(date.today() - timedelta(days=30), date.today()),
            key="pk_hist_date_range"
        )
        _pk_model_f = _pk_h2.selectbox("모델 필터", ["전체"] + sorted(db_pk['모델'].dropna().unique().tolist()), key="pk_hist_model")
        if isinstance(_pk_drange, (list, tuple)) and len(_pk_drange) == 2:
            hist = load_production_history(str(_pk_drange[0]), str(_pk_drange[1]))
        else:
            hist = load_production_history(str(date.today()), str(date.today()))
        hist = hist[(hist['상태'] == '완료') & (hist['라인'] == '포장 라인')]
        if _pk_model_f != "전체":
            hist = hist[hist['모델'] == _pk_model_f]
        hist = hist.sort_values('시간', ascending=False)
        st.caption(f"조회 결과: {len(hist)}건  |  라벨 S/N 셀 클릭 후 직접 수정 가능")
        if not hist.empty:
            _pk_hist_cols = ['시간', '모델', '시리얼', '라벨시리얼', '작업자']
            _pk_hist_cols = [c for c in _pk_hist_cols if c in hist.columns]
            _hist_disp = hist[_pk_hist_cols].reset_index(drop=True)
            _edited = st.data_editor(
                _hist_disp,
                column_config={
                    '시간':     st.column_config.TextColumn('시간',     disabled=True),
                    '모델':     st.column_config.TextColumn('모델',     disabled=True),
                    '시리얼':   st.column_config.TextColumn('시리얼',   disabled=True),
                    '작업자':   st.column_config.TextColumn('작업자',   disabled=True),
                    '라벨시리얼': st.column_config.TextColumn('라벨 S/N', help="클릭하여 수정"),
                },
                use_container_width=True,
                hide_index=True,
                key="pk_hist_editor",
            )
            # 변경된 행 감지 → 저장
            _changed = _edited[_edited['라벨시리얼'] != _hist_disp['라벨시리얼']]
            if not _changed.empty:
                for _, _crow in _changed.iterrows():
                    _sn  = _crow['시리얼']
                    _lsn = str(_crow['라벨시리얼']).strip() if pd.notna(_crow['라벨시리얼']) else ''
                    update_row(_sn, {'라벨시리얼': _lsn})
                    insert_audit_log(시리얼=_sn, 모델=_crow.get('모델', ''), 반=curr_g,
                        이전상태='완료', 이후상태='완료', 작업자=st.session_state.user_id,
                        비고=f"라벨시리얼 수정: {_lsn}")
                    _prod_update(_sn, {'라벨시리얼': _lsn})
                st.rerun()
        else:
            st.info("완료된 제품이 없습니다.")

# ── 생산 지표 관리 ─────────────────────────────────────────────────
elif curr_l == "생산 지표 관리":
    render_kpi_dashboard()

# ── OQC 라인 ─────────────────────────────────────────────────────
elif curr_l == "OQC 라인":
    st.markdown("<h2 class='centered-title'> OQC 출하 품질 검사</h2>", unsafe_allow_html=True)

    # 부적합 사유 선택지
    OQC_DEFECT_REASONS = st.session_state.get('dropdown_oqc_defect', ['(선택)', '기타 (직접 입력)'])

    db_oqc_all = st.session_state.production_db

    # ── 반 선택 ──────────────────────────────────────────────────
    BAN_CLR  = {"제조1반": "#2471a3", "제조2반": "#1e8449", "제조3반": "#6c3483"}
    BAN_BG   = {"제조1반": "#ddeeff", "제조2반": "#d4f0e2", "제조3반": "#ede0f5"}
    oqc_ban  = st.radio("반 선택", ["전체"] + PRODUCTION_GROUPS, horizontal=True, key="oqc_ban_radio")
    db_oqc   = db_oqc_all[db_oqc_all['반'] == oqc_ban] if oqc_ban != "전체" else db_oqc_all

    # ── 요약 KPI (선택 반 기준) ───────────────────────────────────
    oqc_wait  = len(db_oqc[db_oqc['상태'] == 'OQC대기'])
    oqc_ing   = len(db_oqc[db_oqc['상태'] == 'OQC중'])
    oqc_pass  = len(db_oqc[db_oqc['상태'] == '출하승인'])
    oqc_fail  = len(db_oqc[db_oqc['상태'] == '부적합(OQC)'])


    # 반 색상 배지
    if oqc_ban != "전체":
        bc = BAN_CLR.get(oqc_ban, "#888"); bb = BAN_BG.get(oqc_ban, "#f0f0f0")
        st.markdown(f"<span style='background:{bb};color:{bc};padding:3px 12px;border-radius:8px;font-weight:bold;font-size:0.9rem;'> {oqc_ban}</span>", unsafe_allow_html=True)
        st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

    ok1,ok2,ok3,ok4 = st.columns(4)
    ok1.metric(" OQC 대기", f"{oqc_wait}건")
    ok2.metric(" 검사 중",  f"{oqc_ing}건")
    ok3.metric(" 출하 승인", f"{oqc_pass}건")
    ok4.metric(" 부적합",   f"{oqc_fail}건")
    st.divider()

    # ── 입고 대기 목록 (포장 완료 → OQC 대기 전환) ───────────────
    with st.expander(f" 입고 대기 (검사 합격 제품)  {oqc_wait}건", expanded=_xp("oqc_wait"), key="_xp_oqc_wait"):
        packing_done = db_oqc[
            db_oqc['상태'] == 'OQC대기'
        ].sort_values('시간', ascending=False).reset_index(drop=True)
    
        _oqc_in_ck_key = f"oqc_in_ck_{oqc_ban}"
        _oqc_in_sc_cnt = f"oqc_in_sc_cnt_{oqc_ban}"
        if _oqc_in_ck_key not in st.session_state: st.session_state[_oqc_in_ck_key] = {}
        if _oqc_in_sc_cnt not in st.session_state: st.session_state[_oqc_in_sc_cnt] = 0
    
        if not packing_done.empty:
            _oqc_in_sc_key = f"oqc_in_sc_{st.session_state[_oqc_in_sc_cnt]}"
            oi_c1, _ = st.columns([3, 3])
            _oqc_in_scan = oi_c1.text_input(" 시리얼 스캔/검색", placeholder="스캔 또는 입력 → 자동 체크",
                                             key=_oqc_in_sc_key)
            if st.session_state.pop("_autofocus_after_rerun", None) == _oqc_in_sc_key:
                _inject_autofocus(placeholder="스캔 또는 입력 → 자동 체크")
            if _oqc_in_scan.strip():
                _oqc_in_matched = packing_done[packing_done['시리얼'].str.contains(
                    _oqc_in_scan.strip(), case=False, na=False)]
                if not _oqc_in_matched.empty:
                    for _oi in _oqc_in_matched.index:
                        st.session_state[_oqc_in_ck_key][str(_oi)] = True
                    st.session_state["_autofocus_after_rerun"] = f"oqc_in_sc_{st.session_state[_oqc_in_sc_cnt] + 1}"
                    st.session_state[_oqc_in_sc_cnt] += 1  # 키 변경 → 체크박스 강제 재렌더
                    _rerun("oqc_wait")
                else:
                    oi_c1.warning(f"**'{_oqc_in_scan.strip()}'** — 입고 대기 목록에 없습니다.")
    
            _oqc_in_checked = [k for k, v in st.session_state[_oqc_in_ck_key].items() if v]
            if _oqc_in_checked:
                oib1, oib2, oib3 = st.columns([3, 1.5, 1])
                oib1.markdown(f"<span style='color:#2E75B6;font-weight:700;'> {len(_oqc_in_checked)}개 선택됨</span>",
                              unsafe_allow_html=True)
                if oib2.button("▶ 일괄 OQC 시작", key="oqc_bulk_in", type="primary", use_container_width=True):
                    _ops = []
                    for _oi in _oqc_in_checked:
                        _oi_int = int(_oi)
                        if _oi_int in packing_done.index:
                            _orow = packing_done.loc[_oi_int]
                            _upd = {'상태': 'OQC중', '시간': get_now_kst_str(), '라인': 'OQC 라인'}
                            _ops.append({"sn": _orow['시리얼'], "data": _upd,
                                "audit": {"시리얼": _orow['시리얼'], "모델": _orow['모델'], "반": _orow['반'],
                                         "이전상태": 'OQC대기', "이후상태": 'OQC중',
                                         "작업자": st.session_state.user_id}})
                    st.session_state[_oqc_in_ck_key] = {}
                    st.session_state[_oqc_in_sc_cnt] += 1  # 체크박스 키 리셋
                    _prod_bulk_update(_run_bulk_db_ops(_ops))
                    _rerun("oqc_wait")
                if oib3.button(" 해제", key="oqc_in_unck", use_container_width=True):
                    st.session_state[_oqc_in_ck_key] = {}
                    st.session_state[_oqc_in_sc_cnt] += 1  # 체크박스 키 리셋
                    _rerun("oqc_wait")
    
            hh = st.columns([0.4, 2, 2, 1.5, 2, 1.5])
            for col, txt in zip(hh, ["", "시간", "모델", "반", "시리얼", "OQC 시작"]):
                col.markdown(f"<p style='font-size:0.72rem;font-weight:700;color:#8a7f72;margin:0;padding-bottom:3px;border-bottom:1px solid #e0d8c8;'>{txt}</p>", unsafe_allow_html=True)
            _oqc_in_cb_ver = st.session_state[_oqc_in_sc_cnt]
            for idx, row in enumerate(packing_done.to_dict('records')):
                rr = st.columns([0.4, 2, 2, 1.5, 2, 1.5])
                _oqc_in_cb = rr[0].checkbox("", key=f"oqc_in_cb_{idx}_{_oqc_in_cb_ver}",
                    value=st.session_state[_oqc_in_ck_key].get(str(idx), False),
                    label_visibility="collapsed")
                st.session_state[_oqc_in_ck_key][str(idx)] = _oqc_in_cb
                rr[1].caption(str(row.get('시간',''))[:16])
                rr[2].write(row.get('모델',''))
                rr[3].write(row.get('반',''))
                rr[4].markdown(f"`{row.get('시리얼','')}`")
                if rr[5].button("▶ OQC 시작", key=f"oqc_in_{idx}", use_container_width=True, type="primary"):
                    _upd = {'상태': 'OQC중', '시간': get_now_kst_str(), '라인': 'OQC 라인'}
                    update_row(row['시리얼'], _upd)
                    insert_audit_log(시리얼=row['시리얼'], 모델=row['모델'], 반=row['반'],
                        이전상태='OQC대기', 이후상태='OQC중', 작업자=st.session_state.user_id)
                    st.session_state[_oqc_in_ck_key].pop(str(idx), None)
                    _prod_update(row['시리얼'], _upd)
                    _rerun("oqc_wait")
        else:
            st.info("OQC 대기 중인 제품이 없습니다.")
    
    st.divider()

    # ── OQC 검사 진행 ─────────────────────────────────────────────
    with st.expander(f" OQC 검사 진행  {oqc_ing}건", expanded=_xp("oqc_ing"), key="_xp_oqc_ing"):
        oqc_wait_list = db_oqc[db_oqc['상태'] == 'OQC중'].sort_values('시간', ascending=False).reset_index(drop=True)
    
        _oqc_ck_key = f"oqc_ck_{oqc_ban}"
        _oqc_sc_cnt = f"oqc_sc_cnt_{oqc_ban}"
        if _oqc_ck_key not in st.session_state: st.session_state[_oqc_ck_key] = {}
        if _oqc_sc_cnt not in st.session_state: st.session_state[_oqc_sc_cnt] = 0
    
        if not oqc_wait_list.empty:
            # 스캔 입력
            _oqc_sc_key = f"oqc_sc_{st.session_state[_oqc_sc_cnt]}"
            os1, _ = st.columns([3, 3])
            _oqc_scan = os1.text_input(" 시리얼 스캔/검색", placeholder="스캔 또는 입력 → 자동 체크",
                                        key=_oqc_sc_key)
            if st.session_state.pop("_autofocus_after_rerun", None) == _oqc_sc_key:
                _inject_autofocus(placeholder="스캔 또는 입력 → 자동 체크")
            if _oqc_scan.strip():
                _oqc_matched = oqc_wait_list[oqc_wait_list['시리얼'].str.contains(
                    _oqc_scan.strip(), case=False, na=False)]
                if not _oqc_matched.empty:
                    for _oi in _oqc_matched.index:
                        st.session_state[_oqc_ck_key][str(_oi)] = True
                    st.session_state["_autofocus_after_rerun"] = f"oqc_sc_{st.session_state[_oqc_sc_cnt] + 1}"
                    st.session_state[_oqc_sc_cnt] += 1  # 키 변경 → 체크박스 강제 재렌더
                    _rerun("oqc_ing")
                else:
                    os1.warning(f"**'{_oqc_scan.strip()}'** — OQC 검사 목록에 없습니다.")
    
            # 일괄 처리 버튼 (선택된 항목이 있을 때)
            _oqc_checked = [k for k, v in st.session_state[_oqc_ck_key].items() if v]
            if _oqc_checked:
                ob1, ob2, ob3, ob4 = st.columns([2, 1.2, 1.5, 0.8])
                ob1.markdown(f"<span style='color:#2E75B6;font-weight:700;'> {len(_oqc_checked)}개 선택됨</span>",
                             unsafe_allow_html=True)
                if ob2.button(" 일괄 합격", key="oqc_bulk_ok", type="primary", use_container_width=True,
                              disabled=not check_perm("OQC 라인", "write")):
                    _ops = []
                    for _oi in _oqc_checked:
                        _oi_int = int(_oi)
                        if _oi_int in oqc_wait_list.index:
                            _orow = oqc_wait_list.loc[_oi_int]
                            _upd = {'상태': '출하승인', '시간': get_now_kst_str(),
                                    'OQC판정': 'OQC합격'}
                            _ops.append({"sn": _orow['시리얼'], "data": _upd,
                                "audit": {"시리얼": _orow['시리얼'], "모델": _orow['모델'], "반": _orow['반'],
                                         "이전상태": 'OQC중', "이후상태": '출하승인',
                                         "작업자": st.session_state.user_id}})
                    st.session_state[_oqc_ck_key] = {}
                    st.session_state[_oqc_sc_cnt] += 1  # 체크박스 키 리셋
                    _prod_bulk_update(_run_bulk_db_ops(_ops))
                    _rerun("oqc_ing")
                _bulk_defect = ob3.selectbox("부적합 사유", OQC_DEFECT_REASONS,
                                             key="oqc_bulk_defect", label_visibility="collapsed")
                if ob4.button(" 부적합", key="oqc_bulk_ng", use_container_width=True,
                              disabled=not check_perm("OQC 라인", "write")):
                    _dflt = _bulk_defect if _bulk_defect not in ["(선택)", "", None] else ""
                    if not _dflt:
                        st.warning(" 부적합 사유를 먼저 선택해주세요.")
                    else:
                        _ops = []
                        for _oi in _oqc_checked:
                            _oi_int = int(_oi)
                            if _oi_int in oqc_wait_list.index:
                                _orow = oqc_wait_list.loc[_oi_int]
                                _upd = {
                                    '상태': '불량 처리 중',
                                    '시간': get_now_kst_str(),
                                    'OQC판정': f"OQC 부적합 - 사유: {_dflt}"
                                }
                                _ops.append({"sn": _orow['시리얼'], "data": _upd,
                                    "audit": {"시리얼": _orow['시리얼'], "모델": _orow['모델'], "반": _orow['반'],
                                             "이전상태": 'OQC중', "이후상태": '불량 처리 중',
                                             "작업자": st.session_state.user_id,
                                             "비고": f"OQC 부적합 - 사유: {_dflt}"}})
                        st.session_state[_oqc_ck_key] = {}
                        st.session_state[_oqc_sc_cnt] += 1  # 체크박스 키 리셋
                        _prod_bulk_update(_run_bulk_db_ops(_ops))
                        _rerun("oqc_ing")

            st.markdown("<hr style='margin:8px 0;border-color:#e0d8c8;'>", unsafe_allow_html=True)

            # 개별 항목 목록 (체크박스 + 개별 판정)
            _oqc_cb_ver = st.session_state[_oqc_sc_cnt]
            # 자재 시리얼 일괄 조회 (N+1 방지)
            _oqc_bulk_sns = tuple(oqc_wait_list['시리얼'].unique().tolist())
            _oqc_bulk_mats = load_material_serials_bulk(_oqc_bulk_sns) if _oqc_bulk_sns else pd.DataFrame()
            for idx, row in enumerate(oqc_wait_list.to_dict('records')):
                with st.container(border=True):
                    ic1, ic2, ic3, ic4, ic5 = st.columns([0.4, 2, 1.5, 1.5, 1.5])
                    _oqc_cb = ic1.checkbox("", key=f"oqc_cb_{idx}_{_oqc_cb_ver}",
                        value=st.session_state[_oqc_ck_key].get(str(idx), False),
                        label_visibility="collapsed")
                    st.session_state[_oqc_ck_key][str(idx)] = _oqc_cb
                    ic2.markdown(f"**{row.get('모델','')}**  `{row.get('시리얼','')}`")
                    ic3.write(row.get('반',''))
                    ic4.caption(str(row.get('시간',''))[:16])
                    ic5.markdown("<span style='background:#ddeeff;color:#1a4a7a;padding:2px 8px;border-radius:6px;font-size:0.8rem;font-weight:bold;'> 검사중</span>",
                                 unsafe_allow_html=True)
    
                    oq1, oq2 = st.columns([2, 1])
                    with oq1:
                        defect_sel = st.selectbox("부적합 사유 (부적합 처리 시 선택)",
                            st.session_state.get("dropdown_oqc_defect", OQC_DEFECT_REASONS),
                            key=f"oqc_reason_{idx}")
                        if defect_sel == "기타 (직접 입력)":
                            defect_txt = st.text_input("직접 입력", key=f"oqc_reason_txt_{idx}",
                                placeholder="부적합 사유 입력")
                        elif defect_sel == "(선택)":
                            defect_txt = ""
                        else:
                            defect_txt = defect_sel
                    with oq2:
                        btn1 = st.button(" 합격 (출하 승인)", key=f"oqc_ok_{idx}",
                                         use_container_width=True, type="primary")
                        btn2 = st.button(" 부적합", key=f"oqc_ng_{idx}",
                                         use_container_width=True)
                    if btn1:
                        _upd = {'상태': '출하승인', '시간': get_now_kst_str(),
                                'OQC판정': 'OQC합격'}
                        update_row(row['시리얼'], _upd)
                        insert_audit_log(시리얼=row['시리얼'], 모델=row['모델'], 반=row['반'],
                            이전상태='OQC중', 이후상태='출하승인', 작업자=st.session_state.user_id)
                        st.session_state[_oqc_ck_key].pop(str(idx), None)
                        st.session_state[_oqc_sc_cnt] += 1  # 체크박스 키 리셋
                        _prod_update(row['시리얼'], _upd)
                        _rerun("oqc_ing")
                    if btn2:
                        if not defect_txt:
                            st.warning(" 부적합 사유를 먼저 선택해주세요.")
                        else:
                            _upd = {
                                '상태': '불량 처리 중',
                                '시간': get_now_kst_str(),
                                'OQC판정': f"OQC 부적합 - 사유: {defect_txt}"
                            }
                            update_row(row['시리얼'], _upd)
                            insert_audit_log(시리얼=row['시리얼'], 모델=row['모델'], 반=row['반'],
                                이전상태='OQC중', 이후상태='불량 처리 중',
                                작업자=st.session_state.user_id, 비고=f"OQC 부적합 - 사유: {defect_txt}")
                            st.session_state[_oqc_ck_key].pop(str(idx), None)
                            st.session_state[_oqc_sc_cnt] += 1  # 체크박스 키 리셋
                            _prod_update(row['시리얼'], _upd)
                            _rerun("oqc_ing")
    
                    # ── 자재 시리얼 인라인 표시 ──
                    _sn_key = row.get('시리얼', '')
                    _row_mats = _oqc_bulk_mats[_oqc_bulk_mats['메인시리얼'] == _sn_key] if not _oqc_bulk_mats.empty else pd.DataFrame()
                    _mat_label = f" 자재 시리얼 {len(_row_mats)}개 등록됨" if not _row_mats.empty else " 자재 시리얼 미등록 "
                    with st.expander(_mat_label, expanded=False):
                        if not _row_mats.empty:
                            for _rm in _row_mats.to_dict('records'):
                                rmc1, rmc2 = st.columns([2, 4])
                                rmc1.caption(_rm.get('자재명', ''))
                                rmc2.caption(f"`{_rm.get('자재시리얼', '')}`")
                        else:
                            st.caption("등록된 자재 시리얼이 없습니다.")
        else:
            st.info("OQC 검사 대기 중인 제품이 없습니다.")
    
    st.divider()

    # ── OQC 결과 이력 ─────────────────────────────────────────────
    # 부적합 판정 후 불량 처리 중으로 이관된 항목도 포함
    # 구 방식: 수리 컬럼에 'OQC 부적합 판정' / 신규 방식: OQC판정 컬럼에 'OQC 부적합'
    _oqc_fail_transferred = db_oqc[
        (db_oqc['상태'] == '불량 처리 중') & (
            db_oqc['수리'].str.contains('OQC 부적합 판정', na=False) |
            db_oqc['OQC판정'].str.contains('OQC 부적합', na=False)
        )
    ]
    oqc_done = pd.concat([
        db_oqc[db_oqc['상태'].isin(['출하승인', '부적합(OQC)'])],
        _oqc_fail_transferred
    ]).drop_duplicates(subset=['시리얼']).sort_values('시간', ascending=False)
    with st.expander(f" OQC 결과 이력  ·  {len(oqc_done)}건", expanded=_xp("oqc_hist"), key="_xp_oqc_hist"):
    
        if not oqc_done.empty:
            oqc_sn_filter = st.text_input(" S/N 검색", key="oqc_sn_filter", placeholder="시리얼 일부 입력")
            if oqc_sn_filter.strip():
                oqc_done = oqc_done[oqc_done['시리얼'].str.contains(oqc_sn_filter.strip(), case=False, na=False)]
    
            STATE_CLR2 = STATUS_BG  # 전역 상수 재사용 (중복 정의 제거)
    
            rh = st.columns([1.8, 2, 1.5, 2.2, 1.5, 2.5, 1])
            for col, txt in zip(rh, ["시간", "모델", "반", "시리얼", "결과", "비고", "이력"]):
                col.markdown(f"<p style='font-size:0.72rem;font-weight:700;color:#8a7f72;margin:0;padding-bottom:3px;border-bottom:1px solid #e0d8c8;'>{txt}</p>", unsafe_allow_html=True)
    
            # 자재 시리얼 일괄 조회 (OQC 결과 이력)
            _oqc_done_sns = tuple(oqc_done['시리얼'].unique().tolist())
            _oqc_done_mats = load_material_serials_bulk(_oqc_done_sns) if _oqc_done_sns else pd.DataFrame()
            # 성능: iterrows → enumerate + to_dict('records') (idx2 → 순번 _i 로 교체)
            for _i, row in enumerate(oqc_done.to_dict('records')):
                rr2 = st.columns([1.8, 2, 1.5, 2.2, 1.5, 2.5, 1])
                rr2[0].caption(str(row.get('시간',''))[:16])
                rr2[1].write(row.get('모델',''))
                rr2[2].write(row.get('반',''))
                _oqc_done_sn = row.get('시리얼','')
                _oqc_done_mc = len(_oqc_done_mats[_oqc_done_mats['메인시리얼'] == _oqc_done_sn]) if not _oqc_done_mats.empty else 0
                _oqc_done_tog = f"mat_tog_oqcdone_{_oqc_done_sn}_{_i}"
                _oqc_done_badge = f"  {_oqc_done_mc}" if _oqc_done_mc > 0 else "  "
                if rr2[3].button(f"{_oqc_done_sn}{_oqc_done_badge}", key=f"sntog_oqcd_{_i}",
                                 use_container_width=True, help="클릭하여 자재 시리얼 조회"):
                    st.session_state[_oqc_done_tog] = not st.session_state.get(_oqc_done_tog, False)
                결과 = row.get('상태','')
                _is_oqc_transferred = (결과 == '불량 처리 중' and
                    'OQC 부적합 판정' in str(row.get('수리','')))
                if 결과 == '출하승인':
                    rr2[4].markdown("<span style='background:#d4f0e2;color:#1f6640;padding:2px 8px;border-radius:5px;font-size:0.8rem;font-weight:bold;'> 출하승인</span>", unsafe_allow_html=True)
                else:
                    if _is_oqc_transferred:
                        rr2[4].markdown("<span style='background:#fde8e7;color:#7a2e2a;padding:2px 8px;border-radius:5px;font-size:0.8rem;font-weight:bold;'> 부적합 · 이관완료</span>", unsafe_allow_html=True)
                    else:
                        rr2[4].markdown("<span style='background:#fde8e7;color:#7a2e2a;padding:2px 8px;border-radius:5px;font-size:0.8rem;font-weight:bold;'> 부적합</span>", unsafe_allow_html=True)
                    if 결과 == '부적합(OQC)':
                        if rr2[4].button(" 불량 공정 이관", key=f"oqc_send_defect_{_i}",
                                         use_container_width=True, help="불량 공정으로 이관하여 수리/교체 처리"):
                            _oqc_판정 = str(row.get('OQC판정', ''))
                            if '사유: ' in _oqc_판정:
                                _reason = _oqc_판정.split('사유: ', 1)[-1].strip()
                            else:
                                _reason = _oqc_판정 or str(row.get('증상', ''))
                            _upd = {
                                '상태': '불량 처리 중',
                                '시간': get_now_kst_str(),
                                'OQC판정': f"OQC 부적합 이관 - 사유: {_reason}"
                            }
                            update_row(row['시리얼'], _upd)
                            insert_audit_log(시리얼=row['시리얼'], 모델=row['모델'], 반=row['반'],
                                이전상태='부적합(OQC)', 이후상태='불량 처리 중',
                                작업자=st.session_state.user_id, 비고=f"OQC 부적합 이관 - 사유: {_reason}")
                            _prod_update(row['시리얼'], _upd)
                            st.rerun()
    
                # 이력 버튼 → 해당 행 아래 인라인 expander로 표시
                _hist_key = f"oqc_hist_open_{_i}"
                if rr2[6].button("이력", key=f"oqc_hist_{_i}", help="이력 조회"):
                    st.session_state[_hist_key] = not st.session_state.get(_hist_key, False)
    
                if st.session_state.get(_hist_key, False):
                    sn = row.get('시리얼','')
                    with st.container(border=True):
                        hc1, hc2 = st.columns([8, 1])
                        hc1.markdown(f" **제품 전체 이력** — `{sn}`")
                        if hc2.button(" 닫기", key=f"oqc_hist_close_{_i}"):
                            st.session_state[_hist_key] = False
                            st.rerun()
    
                        db_all_h = st.session_state.production_db
                        sn_rows = db_all_h[db_all_h['시리얼'] == sn]
                        if not sn_rows.empty:
                            r0 = sn_rows.iloc[0]
                            st.caption(f"반: {r0.get('반','')}　|　모델: {r0.get('모델','')}　|　품목코드: {r0.get('품목코드','')}")
                        st.markdown("---")
    
                        # 상태 변경 이력
                        st.markdown("** 상태 변경 이력**")
                        try:
                            res = get_supabase().table("audit_log").select("*").eq("시리얼", sn).order("시간").execute()
                            if res.data:
                                aud_df = pd.DataFrame(res.data).drop(columns=['id'], errors='ignore')
                                ah = st.columns([1.8, 1.5, 1.5, 1.2, 3])
                                for col, txt in zip(ah, ["시간","이전상태","이후상태","작업자","비고"]):
                                    col.markdown(f"<p style='font-size:0.7rem;font-weight:700;color:#8a7f72;margin:0;border-bottom:1px solid #e0d8c8;'>{txt}</p>", unsafe_allow_html=True)
                                # 성능: iterrows → to_dict('records')
                                for ar in aud_df.to_dict('records'):
                                    ac = st.columns([1.8, 1.5, 1.5, 1.2, 3])
                                    ac[0].caption(str(ar.get('시간',''))[:16])
                                    prev_c = STATE_CLR2.get(ar.get('이전상태',''), '#f5f2ec')
                                    next_c = STATE_CLR2.get(ar.get('이후상태',''), '#f5f2ec')
                                    ac[1].markdown(f"<span style='background:{prev_c};padding:1px 6px;border-radius:4px;font-size:0.75rem;'>{ar.get('이전상태','')}</span>", unsafe_allow_html=True)
                                    ac[2].markdown(f"<span style='background:{next_c};padding:1px 6px;border-radius:4px;font-size:0.75rem;font-weight:bold;'>{ar.get('이후상태','')}</span>", unsafe_allow_html=True)
                                    ac[3].caption(ar.get('작업자',''))
                                    ac[4].caption(ar.get('비고',''))
                            else:
                                st.info("상태 변경 이력 없음")
                        except Exception as e:
                            st.warning(f"이력 조회 실패: {e}")
    
                        # 자재 시리얼
                        st.markdown("---")
                        st.markdown("** 연결된 자재 시리얼**")
                        mat_df = load_material_serials(sn)
                        if not mat_df.empty:
                            # 성능: iterrows → to_dict('records')
                            for mr in mat_df.to_dict('records'):
                                st.markdown(f"- **{mr.get('자재명','')}** : `{mr.get('자재시리얼','')}`　<span style='color:#aaa;font-size:0.75rem;'>{mr.get('작업자','')}</span>", unsafe_allow_html=True)
                        else:
                            st.info("등록된 자재 시리얼 없음")
                # ── 시리얼 클릭 자재 토글 표시 ──
                if st.session_state.get(_oqc_done_tog, False):
                    _oqcd_row_mats = _oqc_done_mats[_oqc_done_mats['메인시리얼'] == _oqc_done_sn] if not _oqc_done_mats.empty else pd.DataFrame()
                    with st.container(border=True):
                        st.caption(f" 자재 시리얼 — `{_oqc_done_sn}`")
                        if not _oqcd_row_mats.empty:
                            for _dm in _oqcd_row_mats.to_dict('records'):
                                dmc1, dmc2 = st.columns([2, 4])
                                dmc1.caption(_dm.get('자재명', ''))
                                dmc2.caption(f"`{_dm.get('자재시리얼', '')}`")
                        else:
                            st.caption("등록된 자재 시리얼 없음")
        else:
            st.info("OQC 결과 이력이 없습니다.")
    

    st.divider()

    # ── OQC 전용 차트 ─────────────────────────────────────────────
    st.markdown("<div class='section-title'> OQC 분석 차트</div>", unsafe_allow_html=True)

    db_oqc_chart = db_oqc_all.copy()  # 전체 반 기준
    # 부적합 판정 후 불량 처리 중인 항목을 부적합(OQC)로 정규화하여 포함
    # 구 방식: 수리 컬럼에 'OQC 부적합 판정' / 신규 방식: OQC판정 컬럼에 'OQC 부적합'
    _chart_transferred = db_oqc_chart[
        (db_oqc_chart['상태'] == '불량 처리 중') & (
            db_oqc_chart['수리'].str.contains('OQC 부적합 판정', na=False) |
            db_oqc_chart['OQC판정'].str.contains('OQC 부적합', na=False)
        )
    ].copy()
    _chart_transferred['상태'] = '부적합(OQC)'
    oqc_chart_df = pd.concat([
        db_oqc_chart[db_oqc_chart['상태'].isin(['OQC대기','OQC중','출하승인','부적합(OQC)'])],
        _chart_transferred
    ]).drop_duplicates(subset=['시리얼'])

    if not oqc_chart_df.empty:
        import plotly.graph_objects as go
        import pandas as _pd2

        # ── 부적합 사유 추출 (벡터화) ────────────────────────────────
        _fail_nc = oqc_chart_df[oqc_chart_df['상태'] == '부적합(OQC)'].copy()
        if not _fail_nc.empty:
            _oqc_col  = _fail_nc['OQC판정'].astype(str)
            _rep_col  = _fail_nc['수리'].astype(str)
            _sym_col  = _fail_nc['증상'].astype(str)
            _c1 = _oqc_col.str.contains('OQC 부적합', na=False) & _oqc_col.str.contains('사유: ', na=False)
            _c2 = _rep_col.str.contains('OQC 부적합 판정', na=False) & _rep_col.str.contains('사유: ', na=False)
            _c3 = _sym_col.str.contains('부적합사유:', na=False)
            _from_oqc = _oqc_col.str.split('사유: ', n=1).str[-1].str.strip()
            _from_rep = _rep_col.str.split('사유: ', n=1).str[-1].str.strip()
            _from_sym = _sym_col.str.split('부적합사유:', n=1).str[-1].str.strip().str.rstrip(')')
            _fallback = _sym_col.replace({'': '미기재', 'nan': '미기재'})
            _fail_nc['부적합 사유'] = _from_oqc.where(_c1, _from_rep.where(_c2, _from_sym.where(_c3, _fallback)))

        # ── 부적합 판정 이력 (OQC) — audit_log 기준 ──────────────────
        # 서버 필터로 'OQC 부적합 - 사유:' 비고만 조회 (행 수 제한 없음)
        _fail_audit = load_oqc_fail_audit_log().copy()

        # 비고에서 사유 파싱: "OQC 부적합 - 사유: {reason}" 형식 (벡터화)
        _bigo = _fail_audit['비고'].astype(str)
        _has_reason = _bigo.str.contains('사유:', na=False)
        _fail_audit['부적합 사유'] = (
            _bigo.str.split('사유:').str[-1].str.strip().where(
                _has_reason,
                _bigo.replace({'': '미기재', 'nan': '미기재'})
            )
        )

        with st.expander(f" 부적합 판정 이력 (OQC)  ·  {len(_fail_audit)}건",
                         expanded=_xp("oqc_nonconf"), key="_xp_oqc_nonconf"):
            if not _fail_audit.empty:
                _nf_date = st.date_input(
                    "조회 기간",
                    value=(date.today() - timedelta(days=30), date.today()),
                    key="oqc_nc_date_range"
                )
                _nf1, _nf2, _nf3 = st.columns([1.5, 2, 2])
                _nc_ban    = _nf1.selectbox("반 필터",     ["전체"] + PRODUCTION_GROUPS, key="oqc_nc_ban")
                _nc_sn     = _nf2.text_input("S/N 검색",  placeholder="시리얼 일부 입력", key="oqc_nc_sn")
                _nc_reason_opts = ["전체"] + sorted(_fail_audit['부적합 사유'].dropna().unique().tolist())
                _nc_reason = _nf3.selectbox("부적합 사유", _nc_reason_opts, key="oqc_nc_reason")

                _nc_view = _fail_audit.copy()
                if isinstance(_nf_date, (list, tuple)) and len(_nf_date) == 2:
                    _nc_view = _nc_view[_nc_view['시간'].astype(str).str[:10].between(str(_nf_date[0]), str(_nf_date[1]))]
                if _nc_ban != "전체":
                    _nc_view = _nc_view[_nc_view['반'] == _nc_ban]
                if _nc_sn.strip():
                    _nc_view = _nc_view[_nc_view['시리얼'].str.contains(_nc_sn.strip(), case=False, na=False)]
                if _nc_reason != "전체":
                    _nc_view = _nc_view[_nc_view['부적합 사유'] == _nc_reason]

                nk1, nk2, nk3 = st.columns(3)
                nk1.metric("전체 부적합 판정", f"{len(_fail_audit):,} 건")
                nk2.metric("조회 결과",        f"{len(_nc_view):,} 건")
                _top_nc = _fail_audit['부적합 사유'].value_counts().idxmax()
                nk3.metric("최다 발생 사유",   _top_nc)

                if not _nc_view.empty:
                    _cca, _ccb = st.columns(2)
                    with _cca:
                        _r_cnt = _nc_view.groupby('부적합 사유').size().reset_index(name='건수')
                        st.plotly_chart(px.bar(_r_cnt, x='부적합 사유', y='건수',
                                               title="부적합 사유별 현황"), use_container_width=True)
                    with _ccb:
                        _m_cnt = _nc_view.groupby('모델').size().reset_index(name='건수')
                        st.plotly_chart(px.pie(_m_cnt, values='건수', names='모델', hole=0.4,
                                               title="모델별 부적합 비중"), use_container_width=True)
                    _nc_disp = _nc_view[['시간', '시리얼', '모델', '반', '작업자', '부적합 사유']].reset_index(drop=True)
                    st.dataframe(_nc_disp, use_container_width=True, hide_index=True)
                else:
                    st.info("조건에 맞는 부적합 판정 이력이 없습니다.")
            else:
                st.info("부적합 판정 이력이 없습니다.")

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        # ── 요약 KPI 3종 (전체 / 진행 중 / 누적 부적합) ──────────────
        _oqc_total   = len(oqc_chart_df)
        _oqc_active  = len(oqc_chart_df[oqc_chart_df['상태'].isin(['OQC대기','OQC중'])])
        _oqc_fail_tot= len(oqc_chart_df[oqc_chart_df['상태']=='부적합(OQC)'])
        _oqc_pass_tot= len(oqc_chart_df[oqc_chart_df['상태']=='출하승인'])
        _fail_rate   = round(_oqc_fail_tot / max(_oqc_pass_tot + _oqc_fail_tot, 1) * 100, 1)
        ks1, ks2, ks3, ks4 = st.columns(4)
        ks1.metric(" 전체 OQC 수량",    f"{_oqc_total}건")
        ks2.metric(" 현재 진행 중",     f"{_oqc_active}건",
                   help="OQC대기 + OQC중")
        ks3.metric(" 부적합 누적",      f"{_oqc_fail_tot}건",
                   delta=f"부적합률 {_fail_rate}%", delta_color="inverse")
        ks4.metric(" 출하승인 누적",    f"{_oqc_pass_tot}건")
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        cc1, cc3 = st.columns(2)

        # ① 일별 OQC 전체 현황 (전체 수량·진행 중·부적합 누적 포함)
        with cc1:
            if '시간' in oqc_chart_df.columns:
                _chart_df = oqc_chart_df.copy()
                _chart_df['날짜'] = _chart_df['시간'].str[:10]
                from datetime import timedelta as _td
                _cutoff = (date.today() - _td(days=29)).isoformat()
                _chart_df = _chart_df[_chart_df['날짜'] >= _cutoff]

                _daily = _chart_df.groupby(['날짜','상태']).size().reset_index(name='건수')
                _all_dates = _pd2.DataFrame({
                    '날짜': _pd2.date_range(_cutoff, date.today()).strftime('%Y-%m-%d')
                })
                _state_cols = ['출하승인', '부적합(OQC)', 'OQC중', 'OQC대기']
                for _s in _state_cols:
                    _sub = _daily[_daily['상태']==_s][['날짜','건수']].rename(columns={'건수':_s})
                    _all_dates = _all_dates.merge(_sub, on='날짜', how='left')
                _all_dates = _all_dates.fillna(0)

                # 일별 전체 = 4개 상태 합산
                _all_dates['전체'] = _all_dates[_state_cols].sum(axis=1)
                # 누적
                _all_dates['누적_전체']  = _all_dates['전체'].cumsum()
                _all_dates['누적_부적합']= _all_dates['부적합(OQC)'].cumsum()

                _CLR = {'출하승인':'#4da875','부적합(OQC)':'#e8706a',
                        'OQC중':'#7eb8e8','OQC대기':'#f0c878'}
                fig1 = go.Figure()
                # 쌓기 막대: OQC대기 → OQC중 → 출하승인 → 부적합 순
                for _s in ['OQC대기','OQC중','출하승인','부적합(OQC)']:
                    fig1.add_trace(go.Bar(
                        name=_s, x=_all_dates['날짜'], y=_all_dates[_s],
                        marker_color=_CLR[_s], opacity=0.85, yaxis='y'
                    ))
                # 누적 라인 (오른쪽 y축)
                fig1.add_trace(go.Scatter(
                    name='누적 전체', x=_all_dates['날짜'], y=_all_dates['누적_전체'],
                    mode='lines', line=dict(color='#34495e', width=2),
                    yaxis='y2'
                ))
                fig1.add_trace(go.Scatter(
                    name='누적 부적합', x=_all_dates['날짜'], y=_all_dates['누적_부적합'],
                    mode='lines', line=dict(color='#c0392b', width=1.5, dash='dot'),
                    yaxis='y2'
                ))
                # x축 날짜 중 실제 데이터 있는 날만 틱으로 표시
                _tick_dates = _all_dates[_all_dates['전체'] > 0]['날짜'].tolist()
                fig1.update_layout(
                    title=dict(text="일별 OQC 처리 현황 & 누적 추이 (최근 30일)",
                               font=dict(size=13), x=0),
                    template='plotly_white', height=380,
                    barmode='stack',
                    margin=dict(t=40, b=100, l=45, r=55),
                    legend=dict(
                        orientation='h', yanchor='top', y=-0.25,
                        xanchor='center', x=0.5, font=dict(size=10),
                        bgcolor='rgba(255,255,255,0.9)',
                        bordercolor='#e0e0e0', borderwidth=1,
                        traceorder='normal'
                    ),
                    yaxis=dict(
                        title=dict(text='건수', font=dict(size=11)),
                        side='left', gridcolor='#f0f0f0', zeroline=False
                    ),
                    yaxis2=dict(
                        title=dict(text='누적', font=dict(size=11)),
                        side='right', overlaying='y', showgrid=False, zeroline=False
                    ),
                    xaxis=dict(
                        tickvals=_tick_dates,
                        ticktext=[d[5:] for d in _tick_dates],  # MM-DD 형식
                        tickangle=-30, tickfont=dict(size=9),
                        showgrid=False
                    ),
                    plot_bgcolor='white'
                )
                st.plotly_chart(fig1, use_container_width=True)
            else:
                st.info("OQC 처리 이력 데이터 없음")

        # ③ 합격률 추이 (월별)
        with cc3:
            oqc_done_chart = oqc_chart_df[oqc_chart_df['상태'].isin(['출하승인','부적합(OQC)'])].copy()
            if not oqc_done_chart.empty and '시간' in oqc_done_chart.columns:
                oqc_done_chart['월'] = oqc_done_chart['시간'].str[:7]
                # oqc_done_chart는 이미 ['출하승인','부적합(OQC)'] 필터됨 → 전체=투입, 출하승인=합격
                _m_total = oqc_done_chart.groupby('월').size()
                _m_pass  = oqc_done_chart[oqc_done_chart['상태'] == '출하승인'].groupby('월').size()
                _m_pass  = _m_pass.reindex(_m_total.index, fill_value=0)
                monthly  = (_m_pass / _m_total.clip(lower=1) * 100).round(1).reset_index()
                monthly.columns = ['월', '합격률(%)']
                _y_min = max(0, monthly['합격률(%)'].min() - 10) if not monthly.empty else 0
                # 월 레이블: "2026-03" → "26년 3월" 형식 (벡터화)
                _m_parts = monthly['월'].str.split('-')
                monthly = monthly.copy()
                monthly['월_표시'] = _m_parts.str[0].str[2:] + '년 ' + _m_parts.str[1].str.lstrip('0') + '월'
                fig3 = go.Figure()
                fig3.add_trace(go.Scatter(
                    x=monthly['월_표시'], y=monthly['합격률(%)'],
                    mode='lines+markers+text',
                    line=dict(color='#4da875', width=2.5),
                    marker=dict(size=9, color='#4da875',
                                line=dict(color='white', width=2)),
                    text=(monthly['합격률(%)'].astype(str) + '%'),
                    textposition='top center',
                    textfont=dict(size=11, color='#2d6a4f'),
                    hovertemplate='%{x}<br>합격률: %{y}%<extra></extra>'
                ))
                fig3.add_hline(
                    y=100, line_dash="dash", line_color="#e57373", line_width=1.2,
                    annotation_text="목표 100%",
                    annotation_position="top right",
                    annotation_font=dict(size=9, color="#e57373")
                )
                fig3.update_layout(
                    title=dict(text="월별 합격률 추이", font=dict(size=13), x=0),
                    template='plotly_white', height=320,
                    margin=dict(t=40, b=30, l=45, r=20),
                    xaxis=dict(
                        type='category', showgrid=False,
                        tickfont=dict(size=10)
                    ),
                    yaxis=dict(
                        range=[_y_min, 108],
                        title=dict(text='%', font=dict(size=11)),
                        gridcolor='#f0f0f0', zeroline=False
                    ),
                    plot_bgcolor='white'
                )
                st.plotly_chart(fig3, use_container_width=True)
            else:
                st.info("이력 데이터 없음")

        # ④ 모델별 부적합률 테이블
        if not oqc_done_chart.empty:
            _mg_total = oqc_done_chart.groupby('모델').size().rename('전체')
            _mg_pass  = oqc_done_chart[oqc_done_chart['상태'] == '출하승인'].groupby('모델').size().rename('출하승인')
            _mg_fail  = oqc_done_chart[oqc_done_chart['상태'] == '부적합(OQC)'].groupby('모델').size().rename('부적합')
            model_grp = pd.concat([_mg_total, _mg_pass, _mg_fail], axis=1).fillna(0).astype(int).reset_index()
            model_grp['부적합률(%)'] = (model_grp['부적합'] / model_grp['전체'].clip(lower=1) * 100).round(1)
            model_grp = model_grp.sort_values('부적합률(%)', ascending=False)
            st.dataframe(model_grp, use_container_width=True, hide_index=True)
    else:
        st.info("OQC 데이터가 쌓이면 차트가 표시됩니다.")

    st.divider()
    st.divider()
    st.markdown("<div class='db-section' style='background:#7a5c3a;'> 자재 시리얼 관리</div>", unsafe_allow_html=True)

    mat_tab1, mat_tab2, mat_tab3 = st.tabs([" 자재 S/N 검색 (역추적)", " 자재 S/N 등록", " 엑셀 업로드"])

    with mat_tab1:
        st.caption("메인 S/N 또는 자재 S/N으로 검색합니다.")
        
        #  개선: 검색 타입 선택 추가
        search_col1, search_col2 = st.columns(2)
        
        with search_col1:
            st.markdown("####  메인 S/N → 자재 조회")
            main_search = st.text_input("메인 S/N 입력", placeholder="메인 시리얼 입력", key="main_sn_search_mat")
            
            if main_search.strip():
                # 메인 시리얼로 자재 목록 조회
                mat_list = load_material_serials(main_search.strip())
                
                if not mat_list.empty:
                    st.success(f" {len(mat_list)}건 발견")
                    st.markdown(f"**메인 S/N: `{main_search.strip()}`에 사용된 자재 목록**")
                    mh1 = st.columns([2, 2.5, 2.5, 1.5])
                    for col, txt in zip(mh1, ["등록시간", "자재명", "자재 S/N", "작업자"]):
                        col.markdown(f"<p style='font-size:0.72rem;font-weight:700;color:#8a7f72;margin:0;border-bottom:1px solid #e0d8c8;'>{txt}</p>", unsafe_allow_html=True)
                    
                    # 성능: iterrows → to_dict('records')
                    for mr in mat_list.to_dict('records'):
                        mc1 = st.columns([2, 2.5, 2.5, 1.5])
                        mc1[0].caption(str(mr.get('시간',''))[:16])
                        mc1[1].write(mr.get('자재명',''))
                        mc1[2].markdown(f"`{mr.get('자재시리얼','')}`")
                        mc1[3].caption(mr.get('작업자',''))
                else:
                    st.info("해당 메인 S/N의 자재 기록이 없습니다.")
        
        with search_col2:
            st.markdown("####  자재 S/N → 메인 역추적")
            mat_search = st.text_input("자재 S/N 입력", placeholder="자재 시리얼 입력", key="mat_sn_search_reverse")
            
            if mat_search.strip():
                # 자재 시리얼로 역추적
                found = search_material_by_sn(mat_search.strip())
                
                if not found.empty:
                    st.success(f" {len(found)}건 발견")
                    st.markdown(f"**자재 S/N: `{mat_search.strip()}`이 사용된 제품**")
                    mh2 = st.columns([1.8, 2, 1.5, 2, 1.5])
                    for col, txt in zip(mh2, ["등록시간","메인 S/N","반","모델","작업자"]):
                        col.markdown(f"<p style='font-size:0.72rem;font-weight:700;color:#8a7f72;margin:0;border-bottom:1px solid #e0d8c8;'>{txt}</p>", unsafe_allow_html=True)
                    
                    # 성능: iterrows → to_dict('records')
                    for mr in found.to_dict('records'):
                        mc2 = st.columns([1.8, 2, 1.5, 2, 1.5])
                        mc2[0].caption(str(mr.get('시간',''))[:16])
                        mc2[1].markdown(f"**`{mr.get('메인시리얼','')}`**")
                        mc2[2].write(mr.get('반',''))
                        mc2[3].write(mr.get('모델',''))
                        mc2[4].caption(mr.get('작업자',''))
                else:
                    st.info("검색 결과 없음")

    with mat_tab2:
        st.caption("메인 S/N에 자재 S/N을 수동으로 추가 등록합니다.")
        db_now = st.session_state.production_db
        sn_list = db_now['시리얼'].dropna().unique().tolist() if not db_now.empty else []
        mt1, mt2 = st.columns(2)
        main_sn_sel = mt1.selectbox("메인 S/N 선택", ["직접 입력"] + sn_list, key="mat_reg_sn_sel")
        if main_sn_sel == "직접 입력":
            main_sn_val = mt2.text_input("메인 S/N 직접 입력", key="mat_reg_sn_txt")
        else:
            main_sn_val = main_sn_sel
            # 모델/반 자동 표시
            row_info = db_now[db_now['시리얼']==main_sn_sel]
            if not row_info.empty:
                ri = row_info.iloc[0]
                mt2.caption(f"모델: {ri.get('모델','')} / 반: {ri.get('반','')}")

        mat_add_count = st.number_input("추가할 자재 수", min_value=1, max_value=20, step=1, value=1, key="mat_add_count")
        mat_add_list = []
        for mi in range(int(mat_add_count)):
            ac1, ac2 = st.columns(2)
            an = ac1.text_input(f"자재명 #{mi+1}", key=f"mat_add_name_{mi}", placeholder="예: PCB")
            as_ = ac2.text_input(f"자재 S/N #{mi+1}", key=f"mat_add_sn_{mi}", placeholder="자재 시리얼")
            mat_add_list.append({"자재명": an, "자재시리얼": as_})

        if st.button(" 자재 S/N 저장", type="primary", key="mat_save_btn"):
            sn_final = main_sn_val.strip() if main_sn_sel == "직접 입력" else main_sn_sel
            valid = [m for m in mat_add_list if m["자재시리얼"].strip()]
            if sn_final and valid:
                row_m = db_now[db_now['시리얼']==sn_final]
                m_model = row_m.iloc[0]['모델'] if not row_m.empty else ""
                m_ban   = row_m.iloc[0]['반']   if not row_m.empty else ""
                if insert_material_serials(sn_final, m_model, m_ban, valid, st.session_state.user_id):
                    load_material_serials.clear()
                    st.success(f" {sn_final} → {len(valid)}개 자재 S/N 저장 완료")
            else:
                st.warning("메인 S/N과 자재 S/N을 입력해주세요.")

    with mat_tab3:
        st.caption("엑셀 파일로 자재 시리얼을 일괄 등록합니다.")
        # 양식 다운로드
        try:
            import openpyxl as _xl2; import io as _io2
            def _mat_template():
                wb = _xl2.Workbook(); ws = wb.active; ws.title = "자재시리얼"
                headers = ["메인시리얼","모델","반","자재명","자재시리얼"]
                for ci, h in enumerate(headers, 1):
                    ws.cell(1, ci, h)
                ws.append(["MAIN-001","G3014 KBD","제조1반","PCB","PCB-2024-001"])
                ws.append(["MAIN-001","G3014 KBD","제조1반","배터리","BAT-2024-001"])
                buf = _io2.BytesIO(); wb.save(buf); buf.seek(0); return buf
            st.download_button(" 양식 다운로드", data=_mat_template(),
                               file_name="자재시리얼_양식.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="mat_tmpl_dl")
        except Exception: st.caption("양식 다운로드 기능 준비 중")

        mat_file = st.file_uploader("자재 시리얼 엑셀 업로드", type=["xlsx","xls"], key="mat_xl_upload")
        if mat_file:
            try:
                import openpyxl as _xl3
                wb2 = _xl3.load_workbook(mat_file)
                ws2 = wb2.active
                rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
                ok_cnt = 0; err_cnt = 0
                for r2 in rows2:
                    if len(r2) >= 5 and r2[0] and r2[4]:
                        res2 = insert_material_serials(
                            메인시리얼=str(r2[0]).strip(),
                            모델=str(r2[1] or "").strip(),
                            반=str(r2[2] or "").strip(),
                            자재목록=[{"자재명": str(r2[3] or "").strip(), "자재시리얼": str(r2[4]).strip()}],
                            작업자=st.session_state.user_id
                        )
                        if res2: ok_cnt += 1
                        else:    err_cnt += 1
                load_material_serials.clear()  # 자재 캐시만 초기화
                st.success(f" 업로드 완료: {ok_cnt}건 성공 / {err_cnt}건 실패")
            except Exception as e:
                st.error(f"파일 처리 오류: {e}")



# ── 불량 공정 ────────────────────────────────────────────────────
elif curr_l == "불량 공정":
    st.markdown("<h2 class='centered-title'> 불량 분석 및 수리 조치</h2>", unsafe_allow_html=True)
    db = st.session_state.production_db

    # 사용자가 접근 가능한 반 계산
    # "조립 라인" 등 전체 권한이면 모든 반, "조립 라인::제조2반" 등 반별 권한이면 해당 반만
    _has_global_line = any(p in allowed_nav for p in ["조립 라인", "검사 라인", "포장 라인"])
    if _has_global_line:
        _accessible_groups = PRODUCTION_GROUPS
    else:
        _accessible_groups = [g for g in PRODUCTION_GROUPS if any(f"::{g}" in nav for nav in allowed_nav)]
        if not _accessible_groups:
            _accessible_groups = PRODUCTION_GROUPS  # master/admin 등 fallback

    # 반 선택 UI: 접근 가능한 반이 1개면 선택 UI 숨기고 자동 적용
    if len(_accessible_groups) == 1:
        target_groups = _accessible_groups
        st.caption(f" 조회 반: **{_accessible_groups[0]}**")
    else:
        sel_group = st.radio("조회 반 선택", ["전체"] + _accessible_groups, horizontal=True,
                             key="defect_group_radio")
        target_groups = _accessible_groups if sel_group == "전체" else [sel_group]

    # 요약 카드 (선택된 반별)
    card_cols = st.columns(len(target_groups))
    for ci, g in enumerate(target_groups):
        w = len(db[(db['반']==g)&(db['상태']=="불량 처리 중")])
        d = len(db[(db['반']==g)&(db['상태']=='수리 완료(재투입)')])
        with card_cols[ci]:
            st.markdown(
                f"<div style='background:#fffdf8; border:1px solid #e0d8c8; border-radius:12px; padding:14px; margin-bottom:8px;'>"
                f"<div style='font-weight:bold; color:#3d3530; margin-bottom:10px; font-size:1rem;'> {g}</div>"
                f"<div style='display:flex; gap:8px;'>"
                f"<div style='flex:1; background:#fde8e7; border-radius:8px; padding:10px 4px; text-align:center;'>"
                f"<div style='font-size:0.72rem; color:#7a2e2a; font-weight:bold;'> 분석 대기</div>"
                f"<div style='font-size:1.8rem; color:#c8605a; font-weight:bold;'>{w}</div></div>"
                f"<div style='flex:1; background:#d4f0e2; border-radius:8px; padding:10px 4px; text-align:center;'>"
                f"<div style='font-size:0.72rem; color:#1f6640; font-weight:bold;'> 조치 완료</div>"
                f"<div style='font-size:1.8rem; color:#4da875; font-weight:bold;'>{d}</div></div>"
                f"</div></div>",
                unsafe_allow_html=True
            )

    st.divider()

    # ── 시리얼 / 모델 검색 ────────────────────────────────────────
    _def_search = st.text_input(
        " 시리얼 / 모델 검색",
        placeholder="시리얼 또는 모델명 일부 입력",
        key="defect_sn_search"
    )

    # 수리 조치 선택지
    REPAIR_ACTIONS = st.session_state.get('dropdown_repair_action', ['(선택)', '기타 (직접 입력)'])

    # 처리 대기 목록
    has_any = False
    for g in target_groups:
        wait = db[(db['반']==g)&(db['상태']=="불량 처리 중")]
        # 검색 필터 적용
        if _def_search.strip():
            _ds = _def_search.strip().lower()
            wait = wait[
                wait['시리얼'].str.lower().str.contains(_ds, na=False) |
                wait['모델'].str.lower().str.contains(_ds, na=False)
            ]
        if wait.empty: continue
        has_any = True
        with st.expander(f" {g} 불량 처리 대기 ({len(wait)}건)", expanded=_xp(f"def_wait_{g}"), key=f"_xp_def_wait_{g}"):
            for row in wait.to_dict('records'):
                sn_key = row['시리얼']  # idx 대신 실제 시리얼을 키로 사용 (목록 변경 시 키 밀림 방지)
                with st.container(border=True):
                    # 불량 입고 출처 파싱
                    _증상_raw = str(row.get('증상', ''))
                    _oqc판정_raw = str(row.get('OQC판정', ''))
                    _from_line = ''
                    if '불량입고출처:' in _증상_raw:
                        _from_line = _증상_raw.split('불량입고출처:')[-1].strip().split()[0]
                    elif 'OQC 부적합' in _oqc판정_raw:
                        _from_line = 'OQC 라인'

                    ic1, ic2, ic3, ic4, ic5 = st.columns([2, 1.3, 1.5, 1.5, 1.2])
                    ic1.markdown(f"**{row['모델']}**")
                    ic2.markdown(f"`{row['품목코드']}`")
                    ic3.markdown(f"`{row['시리얼']}`")
                    ic4.caption(f" {str(row.get('시간',''))[:16]}")
                    if _from_line:
                        ic5.markdown(
                            f"<div style='background:#fff3d4;color:#7a5c00;padding:2px 6px;"
                            f"border-radius:5px;font-size:0.72rem;font-weight:700;text-align:center;"
                            f"border:1px solid #f0c878;'> {_from_line}</div>",
                            unsafe_allow_html=True)
                    else:
                        ic5.caption("출처 미기록")

                    # OQC 부적합 사유 표시 (신규: OQC판정 / 구 방식: 증상 호환)
                    _oqc_reason = ''
                    if 'OQC 부적합' in _oqc판정_raw and '사유: ' in _oqc판정_raw:
                        _oqc_reason = _oqc판정_raw.split('사유: ', 1)[-1].strip()
                    elif '부적합사유:' in _증상_raw:
                        _oqc_reason = _증상_raw.split('부적합사유:')[-1].strip().rstrip(')')
                    if _oqc_reason:
                        st.markdown(
                            f"<div style='background:#fde8e8;color:#7a1a1a;padding:4px 10px;"
                            f"border-radius:5px;font-size:0.78rem;margin:4px 0 2px 0;"
                            f"border-left:3px solid #e87878;'> OQC 부적합 사유: <b>{_oqc_reason}</b></div>",
                            unsafe_allow_html=True)
                    # 불량 원인 표시 (조립/검사/포장 라인에서 입력한 경우)
                    if '불량원인:' in _증상_raw:
                        _ng_cause_disp = _증상_raw.split('불량원인:')[-1].strip().split('|')[0].strip()
                        st.markdown(
                            f"<div style='background:#fff3d4;color:#7a5c00;padding:4px 10px;"
                            f"border-radius:5px;font-size:0.78rem;margin:4px 0 2px 0;"
                            f"border-left:3px solid #f0c878;'> 불량 원인: <b>{_ng_cause_disp}</b></div>",
                            unsafe_allow_html=True)

                    action_sel = st.selectbox("수리 조치", REPAIR_ACTIONS, key=f"as_{sn_key}")
                    if action_sel == "기타 (직접 입력)":
                        v_a = st.text_input("직접 입력", key=f"a_{sn_key}", placeholder="조치 직접 입력")
                    elif action_sel == "(선택)":
                        v_a = ""
                    else:
                        v_a = action_sel

                    rep_c1, rep_c2 = st.columns(2)
                    target_sn = rep_c1.text_input(
                        " 기존 시리얼",
                        value=row['시리얼'],
                        placeholder="교체 대상 S/N 스캔/입력",
                        key=f"target_{sn_key}"
                    )
                    replace_sn = rep_c2.text_input(
                        " 교체 시리얼",
                        placeholder="새 S/N 스캔/입력 (없으면 비워두세요)",
                        key=f"rep_{sn_key}"
                    )

                    # 등록된 자재 시리얼 표시 (기존 시리얼란에 자재 S/N 입력 가능 안내)
                    _def_mats = load_material_serials(row['시리얼'])
                    if not _def_mats.empty:
                        with st.expander(f" 등록된 자재 시리얼 ({len(_def_mats)}개) — 자재 교체 시 기존 시리얼란에 입력", expanded=_xp(f"def_mat_{g}"), key=f"_xp_def_mat_{g}"):
                            for _, _dm in _def_mats.iterrows():
                                _dmc1, _dmc2 = st.columns([1, 1])
                                _dmc1.caption(f"**{_dm.get('자재명', '')}**")
                                _dmc2.caption(f"`{_dm.get('자재시리얼', '')}`")

                    _btn_col, _ = st.columns([1, 2])
                    _is_submitting = st.session_state.get(f"_def_submit_{sn_key}", False)
                    if _btn_col.button(" 확정", key=f"b_{sn_key}", type="primary",
                                       use_container_width=True, disabled=_is_submitting):
                        if v_a:
                            st.session_state[f"_def_submit_{sn_key}"] = True
                            _target_sn = target_sn.strip() or row['시리얼']
                            _rep_sn = replace_sn.strip()
                            if _rep_sn:
                                # 기존 시리얼이 자재 시리얼인지 확인
                                _def_mats_check = load_material_serials(row['시리얼'])
                                _is_mat_sn = (
                                    not _def_mats_check.empty and
                                    _target_sn != row['시리얼'] and
                                    _target_sn in _def_mats_check['자재시리얼'].values
                                )
                                if _is_mat_sn:
                                    # 자재 시리얼 교체
                                    _mat_info = _def_mats_check[_def_mats_check['자재시리얼'] == _target_sn].iloc[0]
                                    _mat_name = _mat_info.get('자재명', '')
                                    if update_material_serial_sn(row['시리얼'], _target_sn, _rep_sn):
                                        load_material_serials.clear()
                                        _upd = {
                                            '상태': "수리 완료(재투입)", '시간': get_now_kst_str(),
                                            '수리': f"자재교체({_mat_name}:{_target_sn}→{_rep_sn})"
                                        }
                                        update_row(row['시리얼'], _upd)
                                        insert_audit_log(
                                            시리얼=row['시리얼'], 모델=row['모델'], 반=row['반'],
                                            이전상태="불량 처리 중", 이후상태="수리 완료(재투입)",
                                            작업자=st.session_state.user_id,
                                            비고=f"자재교체:{_mat_name} {_target_sn}→{_rep_sn}"
                                        )
                                        _prod_update(row['시리얼'], _upd)
                                        st.toast(f" 자재 시리얼 교체 완료: {_target_sn} → {_rep_sn}")
                                        _rerun(f"def_wait_{g}")
                                else:
                                    # 메인 시리얼 교체 (기존 로직)
                                    _fresh_db = load_realtime_ledger()
                                    _rep_exist = _fresh_db[_fresh_db['시리얼'] == _rep_sn]
                                    if not _rep_exist.empty:
                                        st.warning(f" 교체 시리얼이 이미 등록되어 있습니다: **{_rep_sn}**")
                                    else:
                                        _upd = {
                                            '상태': "교체됨", '시간': get_now_kst_str(),
                                            '수리': f"교체처리({_rep_sn})"
                                        }
                                        update_row(_target_sn, _upd)
                                        insert_audit_log(
                                            시리얼=_target_sn, 모델=row['모델'], 반=row['반'],
                                            이전상태="불량 처리 중", 이후상태="교체됨",
                                            작업자=st.session_state.user_id,
                                            비고=f"교체S/N:{_rep_sn}"
                                        )
                                        _prod_update(_target_sn, _upd)
                                        insert_row({
                                            '시간': get_now_kst_str(), '반': row['반'],
                                            '라인': row.get('라인', '조립 라인'),
                                            '모델': row['모델'], '품목코드': row['품목코드'],
                                            '시리얼': _rep_sn, '상태': '조립중',
                                            '증상': f"교체투입(구S/N:{_target_sn})",
                                            '수리': '', '작업자': st.session_state.user_id
                                        })
                                        insert_audit_log(
                                            시리얼=_rep_sn, 모델=row['모델'], 반=row['반'],
                                            이전상태="-", 이후상태="조립중",
                                            작업자=st.session_state.user_id,
                                            비고=f"교체투입 (구S/N:{_target_sn})"
                                        )
                                        st.session_state.production_db = load_realtime_ledger()
                                        st.toast(f" 교체 완료: {_target_sn} → {_rep_sn}")
                                        _rerun(f"def_wait_{g}")
                            else:
                                _upd = {
                                    '상태': "수리 완료(재투입)", '시간': get_now_kst_str(),
                                    '수리': v_a
                                }
                                update_row(row['시리얼'], _upd)
                                insert_audit_log(
                                    시리얼=row['시리얼'], 모델=row['모델'], 반=row['반'],
                                    이전상태="불량 처리 중", 이후상태="수리 완료(재투입)",
                                    작업자=st.session_state.user_id,
                                    비고=f"조치:{v_a}"
                                )
                                _prod_update(row['시리얼'], _upd)
                                _rerun(f"def_wait_{g}")
                        else:
                            st.warning("수리 조치를 선택해주세요.")
    if not has_any:
        if _def_search.strip():
            st.info(f" '{_def_search.strip()}' 에 해당하는 처리 대기 항목이 없습니다.")
        else:
            st.success("현재 처리 대기 중인 불량 이슈가 없습니다.")

# ── 수리 현황 리포트 ─────────────────────────────────────────────
elif curr_l == "수리 현황 리포트":
    st.markdown("<h2 class='centered-title'> 품질 분석 및 수리 이력 리포트</h2>", unsafe_allow_html=True)

    # ── 날짜 / 반 / 상태 필터 ────────────────────────────────────
    _rp_f1, _rp_f2, _rp_f3 = st.columns([3, 1.2, 1.5])
    _rp_drange = _rp_f1.date_input(
        "조회 기간",
        value=(date.today() - timedelta(days=30), date.today()),
        key="repair_rpt_date_range"
    )
    _rp_ban   = _rp_f2.selectbox("반 필터", ["전체"] + PRODUCTION_GROUPS, key="repair_rpt_ban")
    _rp_state = _rp_f3.selectbox("상태 필터", ["전체", "불량 처리 중", "수리 완료(재투입)", "부적합(OQC)"], key="repair_rpt_state")

    if isinstance(_rp_drange, (list, tuple)) and len(_rp_drange) == 2:
        _rp_from, _rp_to = str(_rp_drange[0]), str(_rp_drange[1])
    else:
        _rp_from = _rp_to = str(date.today())

    hist_df    = load_production_history(_rp_from, _rp_to)
    _audit_rp  = load_audit_log_by_date(_rp_from, _rp_to)

    # 수리 이력 필터 (수리 컬럼 비어있지 않은 행)
    _repair_col = hist_df['수리'].astype(str).str.strip()
    hist_df = hist_df[_repair_col != ""]
    if _rp_ban != "전체":
        hist_df = hist_df[hist_df['반'] == _rp_ban]
    if _rp_state != "전체":
        hist_df = hist_df[hist_df['상태'] == _rp_state]

    # audit_rp 반 필터
    if _rp_ban != "전체" and not _audit_rp.empty:
        _audit_rp = _audit_rp[_audit_rp['반'] == _rp_ban]

    # ══════════════════════════════════════════════════════════════
    # 누적 수리 KPI (audit_log 기반)
    # ══════════════════════════════════════════════════════════════
    st.markdown("<div class='section-title'> 기간 누적 수리 지표</div>", unsafe_allow_html=True)
    if not _audit_rp.empty:
        _ng_total   = len(_audit_rp[_audit_rp['이후상태'] == '불량 처리 중'])
        _repair_done = len(_audit_rp[_audit_rp['이후상태'] == '수리 완료(재투입)'])
        _oqc_fail   = len(_audit_rp[_audit_rp['이후상태'].isin(['부적합(OQC)', '불량 처리 중']) &
                                    _audit_rp['이전상태'].isin(['OQC중', 'OQC대기'])])
        # 반복 수리: 동일 시리얼에서 '수리 완료(재투입)' 이벤트 2회 이상
        _repeat_sn = (
            _audit_rp[_audit_rp['이후상태'] == '수리 완료(재투입)']
            .groupby('시리얼').size()
        )
        _repeat_cnt = int((_repeat_sn >= 2).sum()) if not _repeat_sn.empty else 0

        _kk = st.columns(4)
        _kk[0].metric(" 불량 발생 (누적)", f"{_ng_total:,} 건",
                      help="기간 내 '불량 처리 중' 상태로 전환된 총 횟수")
        _kk[1].metric(" 수리 완료 (누적)", f"{_repair_done:,} 건",
                      help="기간 내 수리 후 재투입된 총 횟수")
        _kk[2].metric(" OQC 부적합 (누적)", f"{_oqc_fail:,} 건",
                      help="OQC 단계에서 부적합 판정된 총 횟수")
        _kk[3].metric(" 반복 수리 S/N", f"{_repeat_cnt:,} 건",
                      help="동일 시리얼이 2회 이상 수리된 건수")
    else:
        st.info("해당 기간 감사 로그 데이터가 없습니다.")

    st.divider()

    # ══════════════════════════════════════════════════════════════
    # 날짜별 수리 추이 (audit_log 기반)
    # ══════════════════════════════════════════════════════════════
    if not _audit_rp.empty:
        _trend_df = _audit_rp[_audit_rp['이후상태'].isin(
            ['불량 처리 중', '수리 완료(재투입)', '부적합(OQC)'])].copy()
        if not _trend_df.empty:
            _trend_df['날짜'] = _trend_df['시간'].astype(str).str[:10]
            _trend_grp = (
                _trend_df.groupby(['날짜', '이후상태'])
                .size().reset_index(name='건수')
            )
            _trend_color = {
                '불량 처리 중':       '#c0392b',
                '수리 완료(재투입)':  '#27ae60',
                '부적합(OQC)':       '#e67e22',
            }
            st.plotly_chart(
                px.bar(_trend_grp, x='날짜', y='건수', color='이후상태',
                       barmode='group', title="날짜별 불량·수리 발생 추이",
                       color_discrete_map=_trend_color),
                use_container_width=True
            )

    # ══════════════════════════════════════════════════════════════
    # 공정별 이슈 빈도 + 모델별 비중 차트
    # ══════════════════════════════════════════════════════════════
    if not hist_df.empty:
        st.caption(f"수리 이력 조회 결과: {len(hist_df)}건")
        c_l, c_r = st.columns([1.8, 1.2])
        with c_l:
            _line_order = ['조립 라인', '검사 라인', 'OQC 라인']
            _issue_df = hist_df.groupby('라인').size().reset_index(name='수량')
            _issue_df['라인'] = pd.Categorical(_issue_df['라인'], categories=_line_order, ordered=True)
            _issue_df = _issue_df.sort_values('라인')
            st.plotly_chart(px.bar(_issue_df, x='라인', y='수량', title="공정별 이슈 빈도",
                category_orders={'라인': _line_order}), use_container_width=True)
        with c_r:
            st.plotly_chart(px.pie(hist_df.groupby('모델').size().reset_index(name='수량'),
                values='수량', names='모델', hole=0.4, title="모델별 불량 비중"), use_container_width=True)

        # ── 수리 이력 테이블 (페이지네이션) ───────────────────────
        _HIST_PAGE_SIZE = 50
        _hist_total = len(hist_df)
        _hist_total_pages = max(1, (_hist_total + _HIST_PAGE_SIZE - 1) // _HIST_PAGE_SIZE)
        _rh_filter_key = f"{_rp_from}|{_rp_to}|{_rp_ban}|{_rp_state}"
        if st.session_state.get("_repair_filter_key") != _rh_filter_key:
            st.session_state["repair_hist_page"] = 1
            st.session_state["_repair_filter_key"] = _rh_filter_key
        if "repair_hist_page" not in st.session_state:
            st.session_state["repair_hist_page"] = 1
        _rh_page = min(st.session_state["repair_hist_page"], _hist_total_pages)
        _rh_start = (_rh_page - 1) * _HIST_PAGE_SIZE
        hist_page_df = hist_df.iloc[_rh_start:_rh_start + _HIST_PAGE_SIZE]

        _rh1, _rh2, _rh3 = st.columns([1, 2, 1])
        if _rh1.button("◀ 이전", key="rh_prev", disabled=(_rh_page <= 1)):
            st.session_state["repair_hist_page"] -= 1; st.rerun()
        _rh2.markdown(
            f"<p style='text-align:center;font-size:0.82rem;color:#8a7f72;margin:6px 0;'>"
            f"페이지 <b>{_rh_page}</b> / {_hist_total_pages}　"
            f"(전체 <b>{_hist_total:,}</b>건, {_HIST_PAGE_SIZE}건/페이지)</p>",
            unsafe_allow_html=True)
        if _rh3.button("다음 ▶", key="rh_next", disabled=(_rh_page >= _hist_total_pages)):
            st.session_state["repair_hist_page"] += 1; st.rerun()

        st.dataframe(hist_page_df, use_container_width=True, hide_index=True)
    else:
        st.info("기록된 이슈 내역이 없습니다.")

    # ── 감사 로그 조회 ────────────────────────────────────────────
    st.divider()
    with st.expander(" 감사 로그 (상태 변경 이력)", expanded=_xp("repair_audit"), key="_xp_repair_audit"):
        # 필터
        af1, af2, af3, af4 = st.columns([1.5, 1.5, 2, 1])
        a_ban    = af1.selectbox("반 필터", ["전체"] + PRODUCTION_GROUPS, key="audit_ban")
        a_state  = af2.selectbox("이후 상태", ["전체", "검사대기", "검사중", "포장대기", "포장중",
                                               "완료", "불량 처리 중", "수리 완료(재투입)", "부적합(OQC)"], key="audit_state")
        a_sn     = af3.text_input("S/N 검색", placeholder="시리얼 일부 입력", key="audit_sn")
        if af4.button(" 새로고침", key="audit_refresh", use_container_width=True):
            _clear_audit_cache(); st.rerun()

        audit_df = load_audit_log()

        if not audit_df.empty:
            if a_ban   != "전체":  audit_df = audit_df[audit_df['반'] == a_ban]
            if a_state != "전체":  audit_df = audit_df[audit_df['이후상태'] == a_state]
            if a_sn.strip():       audit_df = audit_df[audit_df['시리얼'].str.contains(a_sn.strip(), case=False, na=False)]

            # 요약 KPI
            k1, k2, k3, k4, k5 = st.columns(5)
            k1.metric("전체 기록",   f"{len(audit_df):,} 건")
            k2.metric("완료 처리",   f"{len(audit_df[audit_df['이후상태']=='완료']):,} 건")
            k3.metric("불량 발생",   f"{len(audit_df[audit_df['이후상태']=='불량 처리 중']):,} 건")
            k4.metric("수리 완료",   f"{len(audit_df[audit_df['이후상태']=='수리 완료(재투입)']):,} 건")
            k5.metric("부적합(OQC)", f"{len(audit_df[audit_df['이후상태']=='부적합(OQC)']):,} 건")

            st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

            # 상태별 색상 (전역 STATUS_BG 재사용)
            STATE_CLR = STATUS_BG

            # ── 페이지네이션 ──────────────────────────────────────────
            _AUDIT_PAGE_SIZE = 50
            _audit_total = len(audit_df)
            _audit_total_pages = max(1, (_audit_total + _AUDIT_PAGE_SIZE - 1) // _AUDIT_PAGE_SIZE)
            if "audit_page" not in st.session_state:
                st.session_state["audit_page"] = 1
            # 필터 변경 시 첫 페이지로 리셋
            _audit_filter_key = f"{a_ban}|{a_state}|{a_sn.strip()}"
            if st.session_state.get("_audit_filter_key") != _audit_filter_key:
                st.session_state["audit_page"] = 1
                st.session_state["_audit_filter_key"] = _audit_filter_key

            _audit_page = st.session_state["audit_page"]
            _audit_start = (_audit_page - 1) * _AUDIT_PAGE_SIZE
            _audit_end   = _audit_start + _AUDIT_PAGE_SIZE
            audit_page_df = audit_df.iloc[_audit_start:_audit_end]

            # 페이지 네비게이션
            _pn1, _pn2, _pn3 = st.columns([1, 2, 1])
            if _pn1.button("◀ 이전", key="audit_prev", disabled=(_audit_page <= 1)):
                st.session_state["audit_page"] -= 1; st.rerun()
            _pn2.markdown(
                f"<p style='text-align:center;font-size:0.82rem;color:#8a7f72;margin:6px 0;'>"
                f"페이지 <b>{_audit_page}</b> / {_audit_total_pages}　"
                f"(전체 <b>{_audit_total:,}</b>건, {_AUDIT_PAGE_SIZE}건/페이지)</p>",
                unsafe_allow_html=True)
            if _pn3.button("다음 ▶", key="audit_next", disabled=(_audit_page >= _audit_total_pages)):
                st.session_state["audit_page"] += 1; st.rerun()

            # 테이블 헤더
            th = st.columns([1.8, 1.5, 2.2, 1.2, 1.5, 1.5, 1.2, 2.5])
            for col, txt in zip(th, ["시간", "시리얼", "모델", "반", "이전 상태", "이후 상태", "작업자", "비고"]):
                col.markdown(f"<p style='font-size:0.72rem;font-weight:700;color:#8a7f72;margin:0;padding-bottom:3px;border-bottom:1px solid #e0d8c8;'>{txt}</p>", unsafe_allow_html=True)

            for row in audit_page_df.to_dict('records'):
                tr = st.columns([1.8, 1.5, 2.2, 1.2, 1.5, 1.5, 1.2, 2.5])
                tr[0].caption(str(row.get('시간',''))[:16])
                tr[1].markdown(f"`{row.get('시리얼','')}`")
                tr[2].write(row.get('모델',''))
                tr[3].write(row.get('반',''))
                prev_s   = row.get('이전상태','')
                prev_clr = STATE_CLR.get(prev_s, '#f5f2ec')
                tr[4].markdown(f"<span style='background:{prev_clr};padding:1px 6px;border-radius:4px;font-size:0.75rem;'>{html_mod.escape(str(prev_s))}</span>", unsafe_allow_html=True)
                next_s   = row.get('이후상태','')
                next_clr = STATE_CLR.get(next_s, '#f5f2ec')
                tr[5].markdown(f"<span style='background:{next_clr};padding:1px 6px;border-radius:4px;font-size:0.75rem;font-weight:bold;'>{html_mod.escape(str(next_s))}</span>", unsafe_allow_html=True)
                tr[6].caption(row.get('작업자',''))
                tr[7].caption(row.get('비고',''))
        else:
            st.info("감사 로그가 없습니다. 상태 변경 시 자동으로 기록됩니다.")

# ── 마스터 관리 ──────────────────────────────────────────────────
elif curr_l == "마스터 관리":
    st.markdown("<h2 class='centered-title'>⚙️ 시스템 마스터 데이터 관리</h2>", unsafe_allow_html=True)
    try:
        _master_url = st.secrets["master_admin_url"]
    except Exception:
        try:
            _master_url = st.secrets["theme"]["master_admin_url"]
        except Exception:
            _master_url = "https://smart-logistics-dashboard-byzejtmxwn3arn5kd9pmgy.streamlit.app/"
    if _master_url:
        st.info("마스터 관리는 별도 앱에서 운영됩니다.")
        st.link_button("⚙️ 마스터 관리 앱 열기", _master_url, use_container_width=False, type="primary")
    else:
        st.warning("마스터 관리 앱 URL이 설정되지 않았습니다. secrets.toml에 master_admin_url을 추가하세요.")

# =================================================================
# [ PMS v1.0.0 종료 ]
# =================================================================

# ── 사용 설명서 ──────────────────────────────────────────────────
elif curr_l == "작업자 매뉴얼":
    from modules.manual_worker import render_worker_manual
    render_worker_manual()

elif curr_l == "관리자 매뉴얼":
    from modules.manual_admin import render_admin_manual
    render_admin_manual()

elif curr_l == "생산 중단 일지":
    # ══════════════════════════════════════════════════════════
    #  생산 중단 일지
    # ══════════════════════════════════════════════════════════
    st.markdown("<h2 class='centered-title'> 생산 중단 일지</h2>", unsafe_allow_html=True)

    _is_admin_stop = st.session_state.user_role in ["master", "admin"]
    _is_readonly_stop = not _is_admin_stop  # 관리자·마스터만 작성 가능

    _STOP_TYPES = ["설비 고장", "자재 부족", "품질 이슈", "안전 사고", "정전/유틸리티", "작업자 부족", "작업 불량", "관리자 작업지도 불량", "기타"]
    _STOP_LINES = ["조립 라인", "검사 라인", "OQC 라인", "포장 라인", "전체 라인"]

    _st_tab1, _st_tab2 = st.tabs(["📝 일지 등록", "📋 일지 조회"])

    with _st_tab1:
        if _is_readonly_stop:
            st.info("생산 중단 일지 등록은 관리자·마스터만 가능합니다. 조회 탭에서 내역을 확인하세요.")
        else:
            st.markdown("<div class='section-title'> 생산 중단 일지 등록</div>", unsafe_allow_html=True)
            with st.form("stoppage_form", clear_on_submit=True):
                _sf1, _sf2 = st.columns(2)
                _stop_date = _sf1.date_input("중단 날짜", value=date.today(), key="stop_date_input")
                _stop_ban  = _sf2.selectbox("반", PRODUCTION_GROUPS, key="stop_ban_sel")

                _sf3, _sf4 = st.columns(2)
                _stop_line = _sf3.selectbox("라인", _STOP_LINES, key="stop_line_sel")
                _stop_type = _sf4.selectbox("중단 유형", _STOP_TYPES, key="stop_type_sel")

                _sf5, _sf6 = st.columns(2)
                _stop_start = _sf5.text_input("중단 시작 시간 (HH:MM)", placeholder="예: 09:30", key="stop_start_t")
                _stop_end   = _sf6.text_input("중단 종료 시간 (HH:MM)", placeholder="예: 10:15  (진행 중이면 공란)", key="stop_end_t")

                _stop_cause  = st.text_area("중단 원인", placeholder="구체적인 중단 원인을 입력하세요.", height=80, key="stop_cause_ta")
                _stop_action = st.text_area("조치 사항", placeholder="취해진 조치 또는 향후 조치 계획을 입력하세요.", height=80, key="stop_action_ta")

                st.markdown("<p style='font-size:0.85rem;font-weight:700;color:#5a4f45;margin:8px 0 4px 0;'> 조치 수량</p>", unsafe_allow_html=True)
                _sf7, _sf8 = st.columns(2)
                _stop_semi  = _sf7.number_input("반제품 수량", min_value=0, value=0, step=1, key="stop_semi_qty")
                _stop_final = _sf8.number_input("완제품 수량", min_value=0, value=0, step=1, key="stop_final_qty")

                st.markdown("<p style='font-size:0.85rem;font-weight:700;color:#7a5f3a;margin:12px 0 2px 0;'> 재작업 예정 일정 <span style=\"font-weight:400;font-size:0.78rem;color:#aaa;\">(미완료 시 작성)</span></p>", unsafe_allow_html=True)
                _sf9, _sf10 = st.columns(2)
                _stop_rw_date = _sf9.date_input("재작업 예정일", value=None, key="stop_rw_date")
                _stop_rw_note = _sf10.text_input("재작업 비고", placeholder="예: 부품 입고 후 재작업", key="stop_rw_note")
                _sf11, _sf12 = st.columns(2)
                _stop_rw_semi  = _sf11.number_input("재작업 반제품 예정", min_value=0, value=0, step=1, key="stop_rw_semi")
                _stop_rw_final = _sf12.number_input("재작업 완제품 예정", min_value=0, value=0, step=1, key="stop_rw_final")

                _submitted_stop = st.form_submit_button(" 등록", type="primary", use_container_width=True)

            if _submitted_stop:
                if not _stop_cause.strip():
                    st.warning("중단 원인을 입력하세요.")
                elif not _stop_start.strip():
                    st.warning("중단 시작 시간을 입력하세요.")
                else:
                    _stop_row = {
                        "날짜":       str(_stop_date),
                        "반":         _stop_ban,
                        "라인":       _stop_line,
                        "중단유형":   _stop_type,
                        "시작시간":   _stop_start.strip(),
                        "종료시간":   _stop_end.strip() if _stop_end.strip() else None,
                        "중단원인":   _stop_cause.strip(),
                        "조치사항":   _stop_action.strip(),
                        "반제품수량":   int(_stop_semi),
                        "완제품수량":   int(_stop_final),
                        "재작업예정일": str(_stop_rw_date) if _stop_rw_date else None,
                        "재작업반제품": int(_stop_rw_semi),
                        "재작업완제품": int(_stop_rw_final),
                        "재작업비고":   _stop_rw_note.strip(),
                        "재작업완료":   "",
                        "작성자":       st.session_state.user_id,
                        "등록시간":     get_now_kst_str(),
                    }
                    if insert_stoppage_log(_stop_row):
                        st.success(" 생산 중단 일지가 등록되었습니다.")
                        st.rerun()

    with _st_tab2:
        st.markdown("<div class='section-title'> 중단 일지 조회</div>", unsafe_allow_html=True)

        # 기간 프리셋
        _sv_preset_opts = ["직접 입력", "오늘", "이번 주", "이번 달", "지난 달", "전체"]
        _svp1, _svp2, _svp3 = st.columns([2, 1.5, 1.5])
        _sv_preset = _svp1.selectbox("조회 기간", _sv_preset_opts, key="stop_view_preset")
        _sv_ban    = _svp2.selectbox("반 필터", ["전체"] + PRODUCTION_GROUPS, key="stop_view_ban")
        _sv_type   = _svp3.selectbox("유형 필터", ["전체"] + _STOP_TYPES, key="stop_view_type")

        _today = date.today()
        if _sv_preset == "오늘":
            _sv_from = _sv_to = str(_today)
        elif _sv_preset == "이번 주":
            _sv_from = str(_today - timedelta(days=_today.weekday()))
            _sv_to   = str(_today)
        elif _sv_preset == "이번 달":
            _sv_from = _today.strftime("%Y-%m-01")
            _sv_to   = str(_today)
        elif _sv_preset == "지난 달":
            _first_this = _today.replace(day=1)
            _last_prev  = _first_this - timedelta(days=1)
            _sv_from = _last_prev.strftime("%Y-%m-01")
            _sv_to   = str(_last_prev)
        elif _sv_preset == "전체":
            _sv_from = ""
            _sv_to   = ""
        else:  # 직접 입력
            _sv_drange = st.date_input(
                "날짜 범위 선택",
                value=(_today - timedelta(days=30), _today),
                key="stop_view_date"
            )
            if isinstance(_sv_drange, (list, tuple)) and len(_sv_drange) == 2:
                _sv_from, _sv_to = str(_sv_drange[0]), str(_sv_drange[1])
            else:
                _sv_from = _sv_to = str(_today)

        _stop_df = load_stoppage_log(_sv_from, _sv_to)

        if not _stop_df.empty:
            if _sv_ban != "전체":
                _stop_df = _stop_df[_stop_df["반"] == _sv_ban]
            if _sv_type != "전체":
                _stop_df = _stop_df[_stop_df["중단유형"] == _sv_type]

        # KPI 요약
        _sv_kpi1, _sv_kpi2, _sv_kpi3, _sv_kpi4 = st.columns(4)
        _sv_total = len(_stop_df) if not _stop_df.empty else 0
        if not _stop_df.empty and "종료시간" in _stop_df.columns:
            _sv_ongoing = len(_stop_df[_stop_df["종료시간"].isna() | (_stop_df["종료시간"] == "")])
        else:
            _sv_ongoing = 0
        _sv_equip = len(_stop_df[_stop_df["중단유형"] == "설비 고장"]) if not _stop_df.empty else 0
        _sv_mat   = len(_stop_df[_stop_df["중단유형"] == "자재 부족"]) if not _stop_df.empty else 0
        _sv_kpi1.markdown(f"<div class='stat-box'><div class='stat-label'> 총 건수</div><div class='stat-value'>{_sv_total}</div></div>", unsafe_allow_html=True)
        _sv_kpi2.markdown(f"<div class='stat-box'><div class='stat-label'> 진행 중</div><div class='stat-value' style='color:{'#c0392b' if _sv_ongoing>0 else '#888'};'>{_sv_ongoing}</div></div>", unsafe_allow_html=True)
        _sv_kpi3.markdown(f"<div class='stat-box'><div class='stat-label'> 설비 고장</div><div class='stat-value'>{_sv_equip}</div></div>", unsafe_allow_html=True)
        _sv_kpi4.markdown(f"<div class='stat-box'><div class='stat-label'> 자재 부족</div><div class='stat-value'>{_sv_mat}</div></div>", unsafe_allow_html=True)

        st.divider()

        if _stop_df.empty:
            st.info("조회된 중단 일지가 없습니다.")
        else:
            for _si, _sr in _stop_df.reset_index(drop=True).iterrows():
                _is_ongoing = not _sr.get("종료시간") or str(_sr.get("종료시간", "")).strip() == ""
                _badge_color = "#c0392b" if _is_ongoing else "#27ae60"
                _badge_txt   = "진행 중" if _is_ongoing else "완료"
                _duration_txt = f"{_sr.get('시작시간','')} ~ {_sr.get('종료시간','')}" if not _is_ongoing else f"{_sr.get('시작시간','')} ~ (진행 중)"

                with st.expander(
                    f"[{_sr.get('날짜','')}] {_sr.get('반','')} · {_sr.get('라인','')} · {_sr.get('중단유형','')}  ({_duration_txt})",
                    expanded=False
                ):
                    _dc1, _dc2 = st.columns([3, 1])
                    with _dc1:
                        st.markdown(
                            f"<span style='background:{_badge_color};color:#fff;border-radius:6px;padding:2px 10px;font-size:0.78rem;'>{_badge_txt}</span>"
                            f"&nbsp;&nbsp;<span style='color:#888;font-size:0.82rem;'>작성자: {html_mod.escape(str(_sr.get('작성자','')))} &nbsp;·&nbsp; 등록: {str(_sr.get('등록시간',''))[:16]}</span>",
                            unsafe_allow_html=True
                        )
                        st.markdown(f"**중단 원인**")
                        st.write(_sr.get("중단원인", ""))
                        st.markdown(f"**조치 사항**")
                        st.write(_sr.get("조치사항", "") or "—")
                        def _safe_int(v):
                            try: return int(v) if v and str(v) not in ("nan", "None") else 0
                            except: return 0
                        _semi_q  = _safe_int(_sr.get("반제품수량"))
                        _final_q = _safe_int(_sr.get("완제품수량"))
                        if _semi_q > 0 or _final_q > 0:
                            st.markdown(
                                f"<p style='font-size:0.82rem;color:#5a4f45;margin:6px 0 0 0;'>"
                                f" <b>조치 수량</b> &nbsp;·&nbsp; 반제품 <b>{_semi_q}</b>개 &nbsp;/&nbsp; 완제품 <b>{_final_q}</b>개</p>",
                                unsafe_allow_html=True
                            )
                        _rw_date = _sr.get("재작업예정일", "") or ""
                        if _rw_date:
                            _rw_done   = str(_sr.get("재작업완료", "")).strip() == "Y"
                            _rw_bc     = "#27ae60" if _rw_done else "#e67e22"
                            _rw_bt     = "재작업 완료" if _rw_done else "재작업 예정"
                            _rw_semi   = _safe_int(_sr.get("재작업반제품"))
                            _rw_final  = _safe_int(_sr.get("재작업완제품"))
                            _rw_note   = html_mod.escape(str(_sr.get("재작업비고", "") or ""))
                            _rw_qty_txt = f"&nbsp;·&nbsp; 반제품 <b>{_rw_semi}</b>개 / 완제품 <b>{_rw_final}</b>개" if (_rw_semi > 0 or _rw_final > 0) else ""
                            _rw_note_txt = f"&nbsp;·&nbsp; {_rw_note}" if _rw_note else ""
                            st.markdown(
                                f"<div style='background:#fff8f0;border:1.5px solid #f0c070;border-radius:8px;padding:8px 12px;margin-top:10px;'>"
                                f"<span style='background:{_rw_bc};color:#fff;border-radius:6px;padding:2px 9px;font-size:0.75rem;font-weight:bold;'>{_rw_bt}</span>"
                                f"&nbsp;&nbsp;<span style='font-size:0.82rem;color:#5a4f45;'>예정일: <b>{_rw_date}</b>{_rw_qty_txt}{_rw_note_txt}</span>"
                                f"</div>",
                                unsafe_allow_html=True
                            )

                    with _dc2:
                        if _is_admin_stop:
                            _rid_raw = _sr.get("id")
                            if _rid_raw is not None:
                                _rid = int(_rid_raw)
                                _edit_sk = f"stop_edit_mode_{_rid}"
                                if _edit_sk not in st.session_state:
                                    st.session_state[_edit_sk] = False
                                # 종료 시간 입력 (진행 중인 경우)
                                if _is_ongoing:
                                    _end_key = f"stop_end_edit_{_rid}"
                                    _new_end = st.text_input("종료 시간 (HH:MM)", key=_end_key, placeholder="10:30")
                                    if st.button(" 종료 처리", key=f"stop_close_{_rid}", use_container_width=True):
                                        if _new_end.strip():
                                            if update_stoppage_log(_rid, {"종료시간": _new_end.strip()}):
                                                st.success("종료 처리되었습니다.")
                                                st.rerun()
                                        else:
                                            st.warning("종료 시간을 입력하세요.")
                                _rw_date_dc2 = _sr.get("재작업예정일", "") or ""
                                _rw_done_dc2 = str(_sr.get("재작업완료", "")).strip() == "Y"
                                if _rw_date_dc2 and not _rw_done_dc2:
                                    if st.button(" 재작업 완료", key=f"stop_rw_done_{_rid}", use_container_width=True, type="primary"):
                                        if update_stoppage_log(_rid, {"재작업완료": "Y"}):
                                            st.success("재작업 완료 처리되었습니다.")
                                            st.rerun()
                                if st.button(" 수정", key=f"stop_edit_btn_{_rid}", use_container_width=True):
                                    st.session_state[_edit_sk] = not st.session_state[_edit_sk]
                                    st.rerun()
                                if st.button(" 삭제", key=f"stop_del_{_rid}", use_container_width=True):
                                    if delete_stoppage_log_row(_rid):
                                        st.success("삭제되었습니다.")
                                        st.rerun()

                    # 수정 폼 (토글)
                    _rid_raw2 = _sr.get("id")
                    if _is_admin_stop and _rid_raw2 is not None and st.session_state.get(f"stop_edit_mode_{int(_rid_raw2)}", False):
                        _rid = int(_rid_raw2)
                        st.divider()
                        with st.form(f"stop_edit_form_{_rid}"):
                            st.markdown("**✏️ 일지 수정**")
                            _ef1, _ef2 = st.columns(2)
                            try:
                                _e_date_val = date.fromisoformat(_sr["날짜"]) if _sr.get("날짜") else date.today()
                            except Exception:
                                _e_date_val = date.today()
                            _e_date = _ef1.date_input("날짜", value=_e_date_val, key=f"e_date_{_rid}")
                            _e_ban  = _ef2.selectbox("반", PRODUCTION_GROUPS,
                                index=PRODUCTION_GROUPS.index(_sr.get("반")) if _sr.get("반") in PRODUCTION_GROUPS else 0,
                                key=f"e_ban_{_rid}")
                            _ef3, _ef4 = st.columns(2)
                            _e_line = _ef3.selectbox("라인", _STOP_LINES,
                                index=_STOP_LINES.index(_sr.get("라인")) if _sr.get("라인") in _STOP_LINES else 0,
                                key=f"e_line_{_rid}")
                            _e_type = _ef4.selectbox("중단 유형", _STOP_TYPES,
                                index=_STOP_TYPES.index(_sr.get("중단유형")) if _sr.get("중단유형") in _STOP_TYPES else 0,
                                key=f"e_type_{_rid}")
                            _ef5, _ef6 = st.columns(2)
                            _e_start  = _ef5.text_input("시작시간", value=_sr.get("시작시간", ""), key=f"e_start_{_rid}")
                            _e_end    = _ef6.text_input("종료시간", value=_sr.get("종료시간", "") or "", key=f"e_end_{_rid}")
                            _e_cause  = st.text_area("중단 원인", value=_sr.get("중단원인", ""), height=70, key=f"e_cause_{_rid}")
                            _e_action = st.text_area("조치 사항", value=_sr.get("조치사항", "") or "", height=70, key=f"e_action_{_rid}")
                            _ef7, _ef8 = st.columns(2)
                            _e_semi  = _ef7.number_input("반제품 수량", min_value=0, value=_safe_int(_sr.get("반제품수량")), step=1, key=f"e_semi_{_rid}")
                            _e_final = _ef8.number_input("완제품 수량", min_value=0, value=_safe_int(_sr.get("완제품수량")), step=1, key=f"e_final_{_rid}")
                            st.markdown("<p style='font-size:0.82rem;font-weight:700;color:#7a5f3a;margin:10px 0 2px 0;'> 재작업 예정 일정</p>", unsafe_allow_html=True)
                            _ef9, _ef10 = st.columns(2)
                            _e_rw_date_val = _sr.get("재작업예정일", None)
                            try:
                                _e_rw_date_val = date.fromisoformat(str(_e_rw_date_val)) if _e_rw_date_val and str(_e_rw_date_val) not in ("None", "") else None
                            except Exception:
                                _e_rw_date_val = None
                            _e_rw_date = _ef9.date_input("재작업 예정일", value=_e_rw_date_val, key=f"e_rw_date_{_rid}")
                            _e_rw_note = _ef10.text_input("재작업 비고", value=_sr.get("재작업비고", "") or "", key=f"e_rw_note_{_rid}")
                            _ef11, _ef12 = st.columns(2)
                            _e_rw_semi  = _ef11.number_input("재작업 반제품 예정", min_value=0, value=_safe_int(_sr.get("재작업반제품")), step=1, key=f"e_rw_semi_{_rid}")
                            _e_rw_final = _ef12.number_input("재작업 완제품 예정", min_value=0, value=_safe_int(_sr.get("재작업완제품")), step=1, key=f"e_rw_final_{_rid}")
                            _esave, _ecancel = st.columns(2)
                            _submitted_edit = _esave.form_submit_button(" 저장", type="primary", use_container_width=True)
                            _cancel_edit    = _ecancel.form_submit_button(" 취소", use_container_width=True)
                        if _submitted_edit:
                            _upd = {
                                "날짜":       str(_e_date),
                                "반":         _e_ban,
                                "라인":       _e_line,
                                "중단유형":   _e_type,
                                "시작시간":   _e_start.strip(),
                                "종료시간":   _e_end.strip() if _e_end.strip() else None,
                                "중단원인":   _e_cause.strip(),
                                "조치사항":   _e_action.strip(),
                                "반제품수량":   int(_e_semi),
                                "완제품수량":   int(_e_final),
                                "재작업예정일": str(_e_rw_date) if _e_rw_date else None,
                                "재작업반제품": int(_e_rw_semi),
                                "재작업완제품": int(_e_rw_final),
                                "재작업비고":   _e_rw_note.strip(),
                            }
                            if update_stoppage_log(_rid, _upd):
                                st.session_state[f"stop_edit_mode_{_rid}"] = False
                                st.success("수정되었습니다.")
                                st.rerun()
                        if _cancel_edit:
                            st.session_state[f"stop_edit_mode_{_rid}"] = False
                            st.rerun()

elif curr_l == "플로우차트":
    # ══════════════════════════════════════════════════════════
    #  시스템 플로우차트
    # ══════════════════════════════════════════════════════════
    st.markdown("<h2 class='centered-title'> 시스템 플로우차트</h2>", unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    _fc_tab1, _fc_tab2 = st.tabs(["전체 상세 플로우", "작업자 흐름도"])

    with _fc_tab1:
        _fc_path1 = "assets/전체 상세 플로우 차트.html"
        if os.path.exists(_fc_path1):
            with open(_fc_path1, "r", encoding="utf-8") as _f:
                st.components.v1.html(_f.read(), height=900, scrolling=True)
        else:
            st.warning(" 파일을 찾을 수 없습니다: assets/전체 상세 플로우 차트.html")

    with _fc_tab2:
        _fc_path2 = "assets/작업자 흐름도 플로우 차트.html"
        if os.path.exists(_fc_path2):
            with open(_fc_path2, "r", encoding="utf-8") as _f:
                st.components.v1.html(_f.read(), height=900, scrolling=True)
        else:
            st.warning(" 파일을 찾을 수 없습니다: assets/작업자 흐름도 플로우 차트.html")


# =================================================================
# 코드 품질 체크리스트 (Pull Request 전 확인)
# =================================================================
# 
# 새 코드 작성 시 확인사항:
# □ 함수가 100라인 이하인가?
# □ 타입 힌팅을 추가했는가?
# □ Docstring을 작성했는가?
# □ 에러 처리가 구체적인가? (Exception 대신 ValueError 등)
# □ SQL 쿼리에 사용자 입력이 직접 들어가지 않는가?
# □ iterrows() 대신 벡터화 연산을 사용했는가?
# □ st.rerun() 사용이 꼭 필요한가?
# □ 매직 넘버를 상수로 정의했는가?
# □ 단위 테스트를 작성했는가?
# □ 주석이 코드의 '왜'를 설명하는가? (무엇이 아닌)
#
# 코드 리뷰 시 확인사항:
# □ 비즈니스 로직이 명확한가?
# □ 엣지 케이스를 고려했는가?
# □ 성능 병목이 없는가?
# □ 보안 이슈가 없는가?
# □ 일관된 네이밍 규칙을 따르는가?
#
# =================================================================



# =================================================================
# 빠른 참조 가이드
# =================================================================
#
#  주요 상수:
#   AUTO_REFRESH_INTERVAL_MS = 30000  # 자동 새로고침 간격
#   PDF_VIEWER_HEIGHT_PX = 900         # PDF 뷰어 높이
#   MAX_FUNCTION_LINES = 200           # 함수 최대 권장 라인
#
#  색상 상수:
#   COLOR_SUCCESS = "#28a745"
#   COLOR_ERROR = "#dc3545"
#   COLOR_WARNING = "#ffc107"
#   COLOR_INFO = "#17a2b8"
#
#  데이터베이스:
#   DEFAULT_PAGE_SIZE = 100
#   MAX_QUERY_RESULTS = 1000
#
#  보안:
#   - SQL 쿼리 시 re.sub()로 입력 검증
#   - 비밀번호는 hash_pw()로 해싱
#   - 역할 확인: CALENDAR_EDIT_ROLES
#
#  성능:
#   - iterrows() → 벡터화 연산 사용
#   - st.cache_data / st.cache_resource 활용
#   - st.rerun() 최소화
#
#  테스트:
#   pytest tests/ --cov=. --cov-report=html
#
# =================================================================
