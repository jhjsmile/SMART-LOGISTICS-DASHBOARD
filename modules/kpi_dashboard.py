import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta, timezone
import calendar
import re
import io

from modules.database import (
    get_supabase,
    _clear_plan_cache, _clear_schedule_cache,
    load_production_history, load_plan_change_log, save_production_plan,
    load_schedule, insert_schedule, delete_schedule,
    insert_plan_change_log,
)
from modules.auth import check_perm
from modules.calendar_view import _xp, _rerun

# ── 상수 ──────────────────────────────────────────────────────────────
KST = timezone(timedelta(hours=9))
PRODUCTION_GROUPS   = ["제조1반", "제조2반", "제조3반"]
CALENDAR_EDIT_ROLES = ["master", "admin", "control_tower", "schedule_manager"]
ACTIVE_STATES = [
    '조립중', '검사대기', '검사중', 'OQC대기', 'OQC중',
    '출하승인', '포장대기', '포장중', '수리 완료(재투입)', '불량 처리 중',
]
PLAN_CATEGORIES = ["조립계획", "포장계획", "출하계획"]
SCHEDULE_COLORS = {
    "조립계획": "#7eb8e8",
    "포장계획": "#7ec8a0",
    "출하계획": "#f0c878",
    "특이사항": "#e8908a",
    "기타":     "#b49fd4",
}
SCH_CHANGE_REASONS = [
    "(선택 필수)",
    "영업 요구량 변경 (주문 취소)",
    "영업 요구량 변경 (물량 증가)",
    "긴급 주문 (Rush Order)",
    "자재 수급 문제 (입고 지연)",
    "자재 수급 문제 (불량 자재)",
    "설비 고장 / 유지보수",
    "인력 변동 (부족/결원)",
    "품질 문제 (불량 발생)",
    "계획 오입력 수정",
    "기타 (직접 입력)",
]


def render_kpi_dashboard():
    """생산 지표 관리 대시보드 렌더링"""

    # ── CSS: 전광판 스타일 ─────────────────────────────────────────
    st.markdown("""
<style>
.db-title   { font-size:1.35rem; font-weight:800; color:#1a242c; letter-spacing:0.04em;
              text-transform:uppercase; margin:0 0 2px 0; }
.db-section { display:flex; align-items:center; gap:8px; font-size:0.78rem; font-weight:700;
              color:#fff; padding:5px 14px; border-radius:3px; margin:14px 0 8px 0;
              width:fit-content; letter-spacing:0.4px; }
.kpi-card   { background:#f5f6f7; border:1px solid #90a4ae; border-radius:4px;
              padding:12px 16px 10px 16px; text-align:center; }
.kpi-lbl    { font-size:0.68rem; font-weight:700; color:#546e7a;
              text-transform:uppercase; letter-spacing:0.7px; margin-bottom:2px; }
.kpi-val    { font-size:2.2rem; font-weight:800; line-height:1.1; color:#1a242c; }
.kpi-sub    { font-size:0.68rem; color:#78909c; margin-top:2px; }
.kpi-green  { color:#2e7d32; }
.kpi-red    { color:#c62828; }
.kpi-blue   { color:#1565c0; }
.kpi-amber  { color:#f4922a; }
.ban-card   { border-radius:4px; padding:10px 14px 8px 14px; margin-bottom:2px; }
.ban-name   { font-size:0.72rem; font-weight:700; letter-spacing:0.5px;
              text-transform:uppercase; margin-bottom:4px; }
.ban-pct    { font-size:2.6rem; font-weight:900; line-height:1.05; }
.ban-sub    { font-size:0.65rem; color:#78909c; margin-top:1px; }
.ban-row    { display:flex; gap:6px; margin-top:8px; }
.ban-chip   { flex:1; border-radius:3px; padding:5px 0; text-align:center; }
.ban-chip-lbl { font-size:0.6rem; font-weight:700; color:#78909c; text-transform:uppercase; }
.ban-chip-val { font-size:1.3rem; font-weight:800; line-height:1.1; }
.proc-card  { border-radius:4px; padding:10px 14px 10px 14px; }
.proc-name  { font-size:0.75rem; font-weight:700; letter-spacing:0.4px;
              text-transform:uppercase; margin-bottom:6px; }
.proc-row   { display:flex; gap:6px; }
.proc-chip  { flex:1; background:#eceff1; border-radius:3px; padding:6px 4px; text-align:center; }
.proc-chip-lbl { font-size:0.58rem; font-weight:700; color:#78909c; text-transform:uppercase; }
.proc-chip-val { font-size:1.5rem; font-weight:800; line-height:1.1; color:#1a242c; }
.proc-arrow { display:flex; align-items:center; justify-content:center;
              font-size:1.6rem; color:#90a4ae; padding:0 2px; }
.ng-row     { display:flex; gap:8px; align-items:center; padding:6px 0;
              border-bottom:1px solid #cfd8dc; }
.ng-model   { flex:2; font-size:0.82rem; font-weight:600; color:#1a242c; }
.ng-bar-wrap{ flex:3; background:#cfd8dc; border-radius:2px; height:7px; overflow:hidden; }
.ng-bar     { height:100%; border-radius:2px; }
.ng-pct     { flex:1; font-size:0.82rem; font-weight:700; text-align:right; }
.ng-cnt     { flex:1; font-size:0.72rem; color:#78909c; text-align:right; }
.rt-row     { display:flex; gap:0; padding:5px 0; border-bottom:1px solid #cfd8dc;
              align-items:center; font-size:0.78rem; }
.rt-chip    { font-size:0.65rem; font-weight:600; border-radius:3px;
              padding:1px 6px; margin-right:6px; }
</style>""", unsafe_allow_html=True)

    db_all    = st.session_state.production_db
    sch_all   = st.session_state.schedule_db
    today_str = datetime.now(KST).strftime('%Y-%m-%d')

    # ── 상단 필터 (한 줄, 컴팩트) ─────────────────────────────────
    st.markdown("<div class='db-title'> 생산 지표 관리</div>", unsafe_allow_html=True)
    fc1, fc2, _sp = st.columns([2, 2.5, 3])
    period     = fc1.radio("기간", ["월간","주간","현재"], horizontal=True, key="dash_period")
    ban_filter = fc2.radio("반", ["전체"] + PRODUCTION_GROUPS, horizontal=True, key="dash_ban")

    _today = date.today()
    if period == "현재":
        date_from = date_to_d = today_str
        plan_date_to = today_str
    elif period == "주간":
        _mon = _today - timedelta(days=_today.weekday())
        _fri = _mon + timedelta(days=4)
        date_from = _mon.strftime('%Y-%m-%d'); date_to_d = today_str
        plan_date_to = _fri.strftime('%Y-%m-%d')
    else:
        import calendar as _cal
        date_from = today_str[:7] + "-01"; date_to_d = today_str
        _last_day = _cal.monthrange(_today.year, _today.month)[1]
        plan_date_to = f"{today_str[:7]}-{_last_day:02d}"

    if not db_all.empty:
        _t = db_all['시간'].str[:10]
        _mask = (_t >= date_from) & (_t <= date_to_d)
        if ban_filter != "전체": _mask &= (db_all['반'] == ban_filter)
        db_f = db_all[_mask]
    else:
        db_f = db_all

    # KPI/실적용: production_history 포함 (완료 후 아카이브된 제품까지 반영)
    db_kpi = load_production_history(date_from, date_to_d)
    db_kpi_f = db_kpi[db_kpi['반'] == ban_filter] if (not db_kpi.empty and ban_filter != "전체") else db_kpi

    if not sch_all.empty:
        sch_f = sch_all[(sch_all['날짜'] >= date_from) & (sch_all['날짜'] <= plan_date_to)]
        if ban_filter != "전체": sch_f = sch_f[sch_f['반'] == ban_filter]
    else:
        sch_f = sch_all

    # 계획 수량은 조립계획 기준만 사용
    # (조립계획 → 포장계획 → 출하계획은 같은 제품이 단계를 거치는 것이므로 중복 합산 금지)
    sch_f_asm = sch_f[sch_f['카테고리'] == '조립계획'] if not sch_f.empty else sch_f

    def _qty(df, col='조립수'):
        if df.empty: return 0
        return int(pd.to_numeric(df[col], errors='coerce').fillna(0).sum())

    total_in   = len(db_kpi_f) if not db_kpi_f.empty else 0
    total_done = len(db_kpi_f[(db_kpi_f['라인']=='포장 라인') & (db_kpi_f['상태']=='완료')]) if not db_kpi_f.empty else 0
    total_wip  = len(db_kpi_f[db_kpi_f['상태'].isin(ACTIVE_STATES)]) if not db_kpi_f.empty else 0
    # 불량 기준: 현재 불량/부적합 상태 OR 수리 이력 있는 제품 (모델별 불량 분석과 동일 기준)
    total_ng   = len(db_kpi_f[
        db_kpi_f['상태'].str.contains('불량|부적합', na=False) |
        (db_kpi_f['수리'].astype(str).str.strip() != '')
    ]) if not db_kpi_f.empty else 0
    plan_qty   = _qty(sch_f_asm)
    achieve_pct = round(total_done / plan_qty * 100, 1) if plan_qty > 0 else 0
    defect_pct  = round(total_ng / total_in * 100, 1) if total_in > 0 else 0

    # ══════════════════════════════════════════════════════════════
    # [A] KPI 5개 — 한 줄
    # ══════════════════════════════════════════════════════════════
    st.markdown("<div class='db-section' style='background:#4a4540;'>▪ 핵심 지표</div>", unsafe_allow_html=True)
    k = st.columns(5)
    kpi_data = [
        ("계획",    f"{plan_qty:,}", "대", "#5b6abf"),   # 인디고 — 고정 목표
        ("생산 완료", f"{total_done:,}", "대", "#1e8449"),  # 초록
        ("달성률",  f"{achieve_pct}", "%", "#d68910"),
        ("진행 중", f"{total_wip:,}", "대", "#0891b2"),   # 청록 — 계획(인디고)과 구분
        ("불량률",  f"{defect_pct}", "%",
            "#c0392b" if defect_pct > 3 else "#d68910" if defect_pct > 0 else "#6b7280"),  # 0%=회색(달성률 초록과 구분)
    ]
    for col, (lbl, val, unit, color) in zip(k, kpi_data):
        _sub = f"{date_from} ~ {plan_date_to}" if lbl == "계획" else f"{date_from} ~ {date_to_d}"
        col.markdown(f"""
<div class='kpi-card'>
  <div class='kpi-lbl'>{lbl}</div>
  <div class='kpi-val' style='color:{color};'>{val}<span style='font-size:1rem;font-weight:600;color:#aaa;margin-left:2px;'>{unit}</span></div>
  <div class='kpi-sub'>{_sub}</div>
</div>""", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════
    # [B+C] 반별 달성률 + 공정 병목 — 한 줄 (4+3 비율)
    # ══════════════════════════════════════════════════════════════
    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
    left_col, right_col = st.columns([4, 3])

    BAN_COLORS_D = {"제조1반": "#2471a3", "제조2반": "#1e8449", "제조3반": "#6c3483"}

    with left_col:
        st.markdown("<div class='db-section' style='background:#2471a3;'> 반별 달성률</div>", unsafe_allow_html=True)
        # ban_filter와 무관하게 전체 반 계획 수량 필요 → 날짜만 필터링한 별도 변수 사용
        _sch_f_all = sch_all[(sch_all['날짜'] >= date_from) & (sch_all['날짜'] <= plan_date_to)] if not sch_all.empty else sch_all
        _sch_f_asm_all = _sch_f_all[_sch_f_all['카테고리'] == '조립계획'] if not _sch_f_all.empty else _sch_f_all
        bc = st.columns(3)
        for bi, ban in enumerate(PRODUCTION_GROUPS):
            bdb  = db_kpi[db_kpi['반']==ban] if not db_kpi.empty else pd.DataFrame()
            bsch = _sch_f_asm_all[_sch_f_asm_all['반']==ban] if not _sch_f_asm_all.empty else pd.DataFrame()
            b_plan = _qty(bsch)
            b_done = len(bdb[(bdb['라인']=='포장 라인')&(bdb['상태']=='완료')]) if not bdb.empty else 0
            b_wip  = len(bdb[bdb['상태'].isin(ACTIVE_STATES)]) if not bdb.empty else 0
            b_ng   = len(bdb[bdb['상태'].str.contains('불량|부적합',na=False)]) if not bdb.empty else 0
            b_pct  = round(b_done / b_plan * 100, 1) if b_plan > 0 else 0
            clr    = BAN_COLORS_D.get(ban, "#888")
            bar_w  = min(int(b_pct), 100)
            pct_clr = "#d68910"

            with bc[bi]:
                st.markdown(f"""
<div class='ban-card' style='background:{clr}0d; border:1.5px solid {clr}44;'>
  <div class='ban-name' style='color:{clr};'>{ban}</div>
  <div style='background:#e8e2d8;border-radius:99px;height:5px;margin-bottom:6px;overflow:hidden;'>
    <div style='background:{clr};width:{bar_w}%;height:100%;border-radius:99px;'></div>
  </div>
  <div class='ban-pct' style='color:{pct_clr};'>{b_pct}<span style='font-size:1.1rem;'>%</span></div>
  <div class='ban-sub'>계획 {b_plan}대 → 완료 {b_done}대</div>
  <div class='ban-row'>
    <div class='ban-chip' style='background:#ddeeff;'>
      <div class='ban-chip-lbl'>진행</div>
      <div class='ban-chip-val' style='color:#2471a3;'>{b_wip}</div>
    </div>
    <div class='ban-chip' style='background:#d4f0e2;'>
      <div class='ban-chip-lbl'>완료</div>
      <div class='ban-chip-val' style='color:#1e8449;'>{b_done}</div>
    </div>
    <div class='ban-chip' style='background:{"#fde8e7" if b_ng>0 else "#f5f5f5"};'>
      <div class='ban-chip-lbl'>불량</div>
      <div class='ban-chip-val' style='color:{"#c0392b" if b_ng>0 else "#aaa"};'>{b_ng}</div>
    </div>
  </div>
</div>""", unsafe_allow_html=True)

    with right_col:
        st.markdown("<div class='db-section' style='background:#7a6f65;'> 공정 흐름 (병목 감지)</div>", unsafe_allow_html=True)
        lines_info = [
            ("", "조립", "#7eb8e8"),
            ("", "검사", "#7ec8a0"),
            ("", "포장", "#c8a07e"),
        ]
        line_names_full = ["조립 라인", "검사 라인", "포장 라인"]

        # 각 라인 데이터 선계산
        _proc_data = []
        for _pi, (_em, _nm, _cl) in enumerate(lines_info):
            _ldf   = db_f[db_f['라인'] == line_names_full[_pi]] if not db_f.empty else pd.DataFrame()
            _l_tot  = len(_ldf)
            _l_done = len(_ldf[_ldf['상태'] == '완료']) if not _ldf.empty else 0
            _l_wip  = len(_ldf[_ldf['상태'].isin(['조립중','검사중','포장중','수리 완료(재투입)'])]) if not _ldf.empty else 0
            _l_ng   = len(_ldf[_ldf['상태'].str.contains('불량|부적합', na=False)]) if not _ldf.empty else 0
            _l_wait = len(db_f[(db_f['라인'] == line_names_full[_pi-1]) & (db_f['상태'] == '완료')]) if _pi > 0 and not db_f.empty else 0
            _proc_data.append((_em, _nm, _cl, _l_tot, _l_done, _l_wip, _l_ng, _l_wait))

        # 최대 wip 기준 병목 강도 계산
        _max_wip = max(d[5] for d in _proc_data) if _proc_data else 1

        proc_html = "<div style='display:flex;align-items:stretch;gap:0;'>"
        for _pi, (_em, _nm, _cl, _l_tot, _l_done, _l_wip, _l_ng, _l_wait) in enumerate(_proc_data):
            _wip_ratio = _l_wip / max(_max_wip, 1)
            if _wip_ratio >= 0.7 and _l_wip > 0:
                btl_flag = " 병목!"
                btl_hdr  = "#c0392b"
                wip_clr  = "#c0392b"
                _card_bg = "#fde8e730"
                _card_bd = "#c0392b88"
            elif _wip_ratio >= 0.4 and _l_wip > 0:
                btl_flag = " 주의"
                btl_hdr  = "#d68910"
                wip_clr  = "#d68910"
                _card_bg = f"{_cl}18"
                _card_bd = "#d6891088"
            else:
                btl_flag = ""
                btl_hdr  = _cl
                wip_clr  = "#2471a3"
                _card_bg = f"{_cl}18"
                _card_bd = f"{_cl}55"

            proc_html += f"""
<div class='proc-card' style='flex:1;background:{_card_bg};border:1.5px solid {_card_bd};border-radius:10px;'>
  <div class='proc-name' style='color:{btl_hdr};font-weight:700;'>{_em} {_nm}{btl_flag}</div>
  <div class='proc-row'>
    <div class='proc-chip'><div class='proc-chip-lbl'>투입</div><div class='proc-chip-val'>{_l_tot}</div></div>
    <div class='proc-chip'><div class='proc-chip-lbl'>완료</div><div class='proc-chip-val' style='color:#1e8449;'>{_l_done}</div></div>
  </div>
  <div class='proc-row' style='margin-top:4px;'>
    <div class='proc-chip'><div class='proc-chip-lbl'>진행</div><div class='proc-chip-val' style='color:{wip_clr};font-weight:{"700" if _wip_ratio>=0.4 else "400"};'>{_l_wip}</div></div>
    <div class='proc-chip'><div class='proc-chip-lbl'>불량</div><div class='proc-chip-val' style='color:{"#c0392b" if _l_ng>0 else "#aaa"};'>{_l_ng}</div></div>
  </div>
  {"<div style='font-size:0.6rem;color:#888;margin-top:4px;'> 대기 "+str(_l_wait)+"대</div>" if _pi>0 else ""}
</div>"""
            if _pi < 2:
                proc_html += "<div class='proc-arrow'>▶</div>"
        proc_html += "</div>"
        st.markdown(proc_html, unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════
    # [D+E] 불량 분석 + 실시간 — 한 줄 (3+4 비율)
    # ══════════════════════════════════════════════════════════════
    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
    ng_col, rt_col = st.columns([3, 4])

    with ng_col:
        st.markdown("<div class='db-section' style='background:#c0392b;'> 모델별 불량 분석</div>", unsafe_allow_html=True)
        if not db_kpi_f.empty:
            # 불량 판단 기준:
            #  ① 현재 상태가 불량/부적합 이거나
            #  ② 수리 컬럼이 채워진 경우 (수리 이력 = 불량이 있었던 제품)
            _ng_mask = (
                db_kpi_f['상태'].str.contains('불량|부적합', na=False) |
                (db_kpi_f['수리'].astype(str).str.strip() != '')
            )
            _ng_df_src = db_kpi_f.copy()
            _ng_df_src['_has_ng'] = _ng_mask
            ng_df = _ng_df_src.groupby('모델').agg(
                투입=('시리얼','count'),
                불량=('_has_ng', 'sum')
            ).reset_index()
            ng_df['불량'] = ng_df['불량'].astype(int)
            ng_df['불량률'] = (ng_df['불량'] / ng_df['투입'] * 100).round(1)
            ng_df = ng_df[ng_df['불량'] > 0].sort_values('불량률', ascending=False)
            if not ng_df.empty:
                import plotly.graph_objects as go
                _ng_sorted = ng_df.sort_values('불량', ascending=False).reset_index(drop=True)
                _ng_sorted['누적비중(%)'] = (_ng_sorted['불량'].cumsum() / _ng_sorted['불량'].sum() * 100).round(1)
                _bar_clrs = ['#c0392b' if v > 10 else '#d68910' if v > 5 else '#e8c97a' for v in _ng_sorted['불량률']]
                fig_ng = go.Figure()
                fig_ng.add_trace(go.Bar(
                    x=_ng_sorted['모델'], y=_ng_sorted['불량률'],
                    marker_color=_bar_clrs, name='불량률(%)',
                    text=(_ng_sorted['불량률'].astype(str) + '%'),
                    textposition='outside', textfont=dict(size=9),
                    yaxis='y'
                ))
                fig_ng.add_trace(go.Scatter(
                    x=_ng_sorted['모델'], y=_ng_sorted['누적비중(%)'],
                    mode='lines+markers', name='누적 비중(%)',
                    line=dict(color='#2c3e50', width=2),
                    marker=dict(size=6),
                    yaxis='y2'
                ))
                fig_ng.add_hline(y=80, line_dash='dash', line_color='#aaa',
                                 line_width=1, yref='y2',
                                 annotation_text='80%', annotation_font_size=9,
                                 annotation_position='top right')
                fig_ng.update_layout(
                    height=240, template='plotly_white',
                    margin=dict(t=10, b=30, l=30, r=40),
                    legend=dict(orientation='h', y=-0.25, font=dict(size=9)),
                    yaxis=dict(title='불량률(%)', side='left', showgrid=False),
                    yaxis2=dict(title='누적(%)', side='right', overlaying='y',
                                range=[0, 115], showgrid=False),
                    plot_bgcolor='white'
                )
                st.plotly_chart(fig_ng, use_container_width=True)
            else:
                st.success(" 불량 없음")
        else:
            st.info("데이터 없음")

    with rt_col:
        rt_df = st.session_state.production_db.copy()
        if ban_filter != "전체": rt_df = rt_df[rt_df['반'] == ban_filter]
        rt_wip = rt_df[rt_df['상태'].isin(ACTIVE_STATES)].sort_values('시간', ascending=False) if not rt_df.empty else pd.DataFrame()

        with st.expander(f" 실시간 진행 중 ({len(rt_wip)}건)", expanded=_xp("idx_wip"), key="_xp_idx_wip"):
            if not rt_wip.empty:
                BAN_BG = {"제조1반":"#ddeeff","제조2반":"#d4f0e2","제조3반":"#ede0f5"}
                BAN_CL = {"제조1반":"#2471a3","제조2반":"#1e8449","제조3반":"#6c3483"}
                LINE_BG = {"조립 라인":"#fff3d4","검사 라인":"#d4f0e2","포장 라인":"#fde8d4","OQC 라인":"#e8d4f0"}
                STATE_BG = {"불량 처리 중":"#fde8e7"}
                rt_html = "<div style='font-size:0.7rem;font-weight:600;color:#aaa;display:flex;gap:0;padding:0 0 4px 0;border-bottom:2px solid #e8e2d8;margin-bottom:2px;'><span style='flex:1.2;'>반</span><span style='flex:1.5;'>라인</span><span style='flex:2.5;'>모델</span><span style='flex:2;'>시리얼</span><span style='flex:1.8;'>상태</span></div>"
                for row in rt_wip.to_dict('records'):
                    ban_v   = row.get('반','')
                    line_v  = row.get('라인','')
                    state_v = row.get('상태','')
                    bbg = BAN_BG.get(ban_v, "#f0f0f0"); bcl = BAN_CL.get(ban_v, "#666")
                    lbg = LINE_BG.get(line_v, "#f0f0f0")
                    row_bg = "background:#fdf5f5;" if state_v == '불량 처리 중' else ""
                    state_color = "#c0392b" if state_v == '불량 처리 중' else "#5a5048"
                    rt_html += f"""
<div class='rt-row' style='{row_bg}'>
  <span style='flex:1.2;'><span class='rt-chip' style='background:{bbg};color:{bcl};'>{ban_v[:3]}</span></span>
  <span style='flex:1.5;'><span class='rt-chip' style='background:{lbg};color:#555;'>{line_v[:2]}</span></span>
  <span style='flex:2.5;font-weight:600;'>{row.get('모델','')}</span>
  <span style='flex:2;color:#5a5048;font-family:monospace;'>{row.get('시리얼','')}</span>
  <span style='flex:1.8;color:{state_color};font-size:0.68rem;'>{state_v}</span>
</div>"""
                st.markdown(rt_html, unsafe_allow_html=True)
            else:
                st.info("현재 진행 중인 작업 없음")

    # ══════════════════════════════════════════════════════════════
    # [F] 계획 수량 입력 + 월별 달성률 그래프
    # ══════════════════════════════════════════════════════════════
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    st.markdown("<div class='db-section' style='background:#5a4f8a;'> 월별 계획 수량 관리</div>", unsafe_allow_html=True)

    # 변경 사유 선택지
    PLAN_CHANGE_REASONS = [
        "신규 계획 등록",
        "영업 요구량 변경 (주문 취소)",
        "영업 요구량 변경 (물량 증가)",
        "긴급 주문 (Rush Order)",
        "자재 수급 문제 (입고 지연)",
        "자재 수급 문제 (불량 자재)",
        "설비 고장 / 유지보수",
        "인력 변동 (부족/결원)",
        "품질 문제 (불량 발생)",
        "기타 (직접 입력)",
    ]

    # 입력 폼 (관리자 전용)
    if st.session_state.user_role in CALENDAR_EDIT_ROLES and check_perm("생산 지표 관리", "write"):
        with st.expander(" 월 계획 수량 입력 / 변경", expanded=_xp("idx_plan"), key="_xp_idx_plan"):
            st.caption("반/월별 목표 수량을 입력합니다. 변경 시 사유를 선택하면 이력이 자동 기록됩니다.")
            pl1, pl2, pl3 = st.columns([1.5, 1.5, 1.2])
            p_ban  = pl1.selectbox("반", PRODUCTION_GROUPS, key="plan_ban")
            # Bug fix: 28일 단위 대신 월 단위로 정확히 역산 (Python modulo 활용)
            _today2 = date.today()
            _months = []
            for i in range(6):
                _mo = (_today2.month - 1 - i) % 12 + 1
                _yr = _today2.year + (_today2.month - 1 - i) // 12
                _months.append(f"{_yr}-{_mo:02d}")
            _months = sorted(set(_months), reverse=True)[:6]
            p_month = pl2.selectbox("월", _months, key="plan_month")
            p_qty   = pl3.number_input("계획 수량 (대)", min_value=0, step=10, key="plan_qty")

            # 현재 저장된 수량 표시
            _cur_qty = st.session_state.production_plan.get(f"{p_ban}_{p_month}", 0)
            if _cur_qty > 0:
                st.caption(f" 현재 저장된 수량: **{_cur_qty:,}대** → 변경 후: **{int(p_qty):,}대** "
                           f"({'▲' if int(p_qty) > _cur_qty else '▼'} {abs(int(p_qty)-_cur_qty):,}대)")

            pr1, pr2 = st.columns([2, 2])
            p_reason = pr1.selectbox("변경 사유 *", PLAN_CHANGE_REASONS, key="plan_reason")
            p_detail = pr2.text_input("상세 내용 (선택)", placeholder="예: 고객사 요청, 부품 수급 지연 등", key="plan_detail")

            if st.button(" 저장", type="primary", key="plan_save_btn"):
                _reason_final = p_detail.strip() if p_reason == "기타 (직접 입력)" and p_detail.strip() else p_reason
                if save_production_plan(p_ban, p_month, int(p_qty)):
                    # 변경 로그 기록
                    insert_plan_change_log(
                        반=p_ban, 월=p_month,
                        이전수량=_cur_qty, 변경수량=int(p_qty),
                        변경사유=_reason_final, 사유상세=p_detail.strip(),
                        작업자=st.session_state.user_id
                    )
                    plan_key = f"{p_ban}_{p_month}"
                    st.session_state.production_plan[plan_key] = int(p_qty)
                    _clear_plan_cache()
                    st.toast(f" {p_ban} / {p_month} → {p_qty:,}대 저장 완료")
                    _rerun("idx_plan")

    # ── 월별 달성률 그래프 ────────────────────────────────────────
    plan_map_now = st.session_state.production_plan  # {반_YYYY-MM: 계획수량}

    # 최근 6개월 목록 (Bug fix: 28일 단위 대신 월 단위로 정확히 역산)
    _td3 = date.today()
    months_list = []
    for i in range(5, -1, -1):
        _mo3 = (_td3.month - 1 - i) % 12 + 1
        _yr3 = _td3.year + (_td3.month - 1 - i) // 12
        months_list.append(f"{_yr3}-{_mo3:02d}")
    months_list = sorted(set(months_list))[-6:]

    # 반별 월별 실적 집계 (6개월 데이터 → load_production_history로 DB 레벨 범위 지정)
    _six_ago = (date.today().replace(day=1) - timedelta(days=5*28)).strftime('%Y-%m-%d')
    db_raw = load_production_history(_six_ago, str(date.today()), limit=10000)
    chart_rows = []
    for ban in PRODUCTION_GROUPS:
        for mo in months_list:
            plan_v = plan_map_now.get(f"{ban}_{mo}", 0)
            # 해당 월 포장 완료 건수
            if not db_raw.empty:
                actual_v = len(db_raw[
                    (db_raw['반'] == ban) &
                    (db_raw['상태'] == '완료') &
                    (db_raw['라인'] == '포장 라인') &
                    (db_raw['시간'].str[:7] == mo)
                ])
            else:
                actual_v = 0
            pct = round(actual_v / plan_v * 100, 1) if plan_v > 0 else 0
            chart_rows.append({
                '월': mo, '반': ban,
                '계획': plan_v, '실적': actual_v, '달성률(%)': pct
            })

    chart_df = pd.DataFrame(chart_rows)

    if not chart_df.empty and chart_df['계획'].sum() > 0:
        import plotly.graph_objects as go

        BAN_CLR = {"제조1반": "#2471a3", "제조2반": "#1e8449", "제조3반": "#6c3483"}
        BAN_CLR_LIGHT = {"제조1반": "#aad4f5", "제조2반": "#a8dfc4", "제조3반": "#caaee8"}

        gc1, gc2 = st.columns([3, 2])

        with gc1:
            # 그룹 막대 차트: 계획 vs 실적
            fig_plan = go.Figure()
            for ban in PRODUCTION_GROUPS:
                bdf = chart_df[chart_df['반'] == ban]
                fig_plan.add_trace(go.Bar(
                    name=f"{ban} 계획",
                    x=bdf['월'], y=bdf['계획'],
                    marker_color=BAN_CLR_LIGHT.get(ban, "#ccc"),
                    opacity=0.6, offsetgroup=ban,
                    text=bdf['계획'].map(lambda v: f"{v:,}" if v > 0 else ""),
                    textposition='outside', textfont=dict(size=9)
                ))
                fig_plan.add_trace(go.Bar(
                    name=f"{ban} 실적",
                    x=bdf['월'], y=bdf['실적'],
                    marker_color=BAN_CLR.get(ban, "#888"),
                    offsetgroup=ban + "_실적",
                    text=bdf['실적'].map(lambda v: f"{v:,}" if v > 0 else ""),
                    textposition='outside', textfont=dict(size=9)
                ))
            fig_plan.update_layout(
                title="월별 계획 vs 실적 (반별)",
                barmode='group',
                template='plotly_white',
                height=320,
                margin=dict(t=40, b=40, l=20, r=20),
                legend=dict(orientation='h', y=-0.2, font=dict(size=10)),
                yaxis_title="수량 (대)"
            )
            st.plotly_chart(fig_plan, use_container_width=True)

        with gc2:
            # 달성률 라인 차트
            fig_pct = go.Figure()
            for ban in PRODUCTION_GROUPS:
                bdf = chart_df[chart_df['반'] == ban]
                fig_pct.add_trace(go.Scatter(
                    name=ban, x=bdf['월'], y=bdf['달성률(%)'],
                    mode='lines+markers+text',
                    line=dict(color=BAN_CLR.get(ban, "#888"), width=2),
                    marker=dict(size=7),
                    text=bdf['달성률(%)'].map(lambda v: f"{v}%" if v > 0 else ""),
                    textposition='top center', textfont=dict(size=9)
                ))
            # 100% 기준선
            fig_pct.add_hline(y=100, line_dash="dash", line_color="#e8908a",
                              annotation_text="목표 100%", annotation_font_size=10)
            fig_pct.update_layout(
                title="월별 달성률 추이 (%)",
                template='plotly_white',
                height=320,
                margin=dict(t=40, b=40, l=20, r=20),
                legend=dict(orientation='h', y=-0.2, font=dict(size=10)),
                yaxis=dict(title="달성률 (%)", range=[0, max(120, (chart_df['달성률(%)'].max() + 10) if not chart_df.empty else 0)])
            )
            st.plotly_chart(fig_pct, use_container_width=True)

        # 요약 테이블
        summary = chart_df.groupby('반').agg(
            총계획=('계획','sum'), 총실적=('실적','sum')
        ).reset_index()
        summary['전체달성률(%)'] = (summary['총실적'] / summary['총계획'] * 100).round(1).where(summary['총계획'] > 0, 0)
        st.dataframe(summary, use_container_width=True, hide_index=True)
    else:
        st.info("계획 수량을 입력하면 달성률 그래프가 표시됩니다.")

    # ── 변경 이력 로그 (탭 구분) ─────────────────────────────────
    st.divider()
    st.markdown("<div class='db-section' style='background:#5a4f8a;'> 변경 이력 로그</div>", unsafe_allow_html=True)

    REASON_COLOR = {
        "신규 계획 등록":              "#ddeeff",
        "영업 요구량 변경 (주문 취소)": "#fde8e7",
        "영업 요구량 변경 (물량 증가)": "#d4f0e2",
        "긴급 주문 (Rush Order)":      "#fff3d4",
        "자재 수급 문제 (입고 지연)":   "#fde8d4",
        "자재 수급 문제 (불량 자재)":   "#fde8d4",
        "설비 고장 / 유지보수":         "#fde8e7",
        "인력 변동 (부족/결원)":        "#ede0f5",
        "품질 문제 (불량 발생)":        "#fde8e7",
        "계획 오입력 수정":             "#f5f2ec",
    }

    log_tab1, log_tab2 = st.tabs([" 월별 계획 수량 변경", " 일정 수정 이력"])

    # ── 탭1: 월별 계획 수량 변경 이력 ───────────────────────────
    with log_tab1:
        lf1, lf2, lf3 = st.columns([1.5, 2, 1])
        log_ban    = lf1.selectbox("반 필터", ["전체"] + PRODUCTION_GROUPS, key="plog_ban")
        log_reason = lf2.selectbox("사유 필터", ["전체"] + PLAN_CHANGE_REASONS, key="plog_reason")
        if lf3.button(" 새로고침", key="plog_refresh", use_container_width=True):
            _clear_plan_cache(); st.rerun()

        plog_df = load_plan_change_log()
        if not plog_df.empty:
            if log_ban    != "전체": plog_df = plog_df[plog_df['반'] == log_ban]
            if log_reason != "전체": plog_df = plog_df[plog_df['변경사유'].str.contains(log_reason, na=False)]

            pk1, pk2, pk3, pk4 = st.columns(4)
            pk1.metric("전체 변경 건수", f"{len(plog_df)}건")
            inc = plog_df[plog_df['증감'] > 0]['증감'].sum()
            dec = plog_df[plog_df['증감'] < 0]['증감'].sum()
            pk2.metric("총 증가량", f"+{int(inc):,}대")
            pk3.metric("총 감소량", f"{int(dec):,}대")
            pk4.metric("순 변동량", f"{int(inc+dec):+,}대")

            th = st.columns([1.8, 1.0, 1.0, 1.0, 1.0, 0.9, 2.5, 2.0, 1.2])
            for col, txt in zip(th, ["시간","반","월","이전","변경","증감","변경 사유","상세 내용","작업자"]):
                col.markdown(f"<p style='font-size:0.72rem;font-weight:700;color:#8a7f72;margin:0;padding-bottom:3px;border-bottom:1px solid #e0d8c8;'>{txt}</p>", unsafe_allow_html=True)
            # 성능: iterrows → to_dict('records')
            for row in plog_df.to_dict('records'):
                tr = st.columns([1.8, 1.0, 1.0, 1.0, 1.0, 0.9, 2.5, 2.0, 1.2])
                tr[0].caption(str(row.get('시간',''))[:16])
                tr[1].write(row.get('반',''))
                tr[2].write(row.get('월',''))
                tr[3].caption(f"{int(row.get('이전수량',0)):,}대")
                tr[4].write(f"**{int(row.get('변경수량',0)):,}대**")
                증감 = int(row.get('증감', 0))
                clr = "#1e8449" if 증감 > 0 else "#c0392b" if 증감 < 0 else "#888"
                tr[5].markdown(f"<span style='color:{clr};font-weight:bold;font-size:0.85rem;'>{증감:+,}</span>", unsafe_allow_html=True)
                reason_v = str(row.get('변경사유',''))
                rbg = REASON_COLOR.get(reason_v, "#f5f2ec")
                reason_esc = html_mod.escape(reason_v)
                tr[6].markdown(f"<span style='background:{rbg};padding:1px 6px;border-radius:4px;font-size:0.72rem;'>{reason_esc}</span>", unsafe_allow_html=True)
                tr[7].caption(row.get('사유상세',''))
                tr[8].caption(row.get('작업자',''))
        else:
            st.info("계획 변경 이력이 없습니다. 계획 수량 저장 시 자동 기록됩니다.")

    # ── 탭2: 일정 수정 이력 ──────────────────────────────────────
    with log_tab2:
        @st.cache_data(ttl=30)
        def load_schedule_change_log(limit: int = 200) -> pd.DataFrame:
            try:
                sb  = get_supabase()
                res = sb.table("schedule_change_log").select("*").order("시간", desc=True).limit(limit).execute()
                if res.data:
                    return pd.DataFrame(res.data).drop(columns=['id'], errors='ignore')
                return pd.DataFrame(columns=['시간','일정id','날짜','반','모델명','이전내용','변경내용','변경사유','사유상세','작업자'])
            except Exception:
                return pd.DataFrame(columns=['시간','일정id','날짜','반','모델명','이전내용','변경내용','변경사유','사유상세','작업자'])

        sf1, sf2, sf3 = st.columns([1.5, 2, 1])
        s_ban    = sf1.selectbox("반 필터", ["전체"] + PRODUCTION_GROUPS, key="slog_ban")
        s_reason = sf2.selectbox("사유 필터", ["전체"] + SCH_CHANGE_REASONS[1:], key="slog_reason")
        if sf3.button(" 새로고침", key="slog_refresh", use_container_width=True):
            _clear_schedule_cache(); st.rerun()

        slog_df = load_schedule_change_log()
        if not slog_df.empty:
            if s_ban    != "전체": slog_df = slog_df[slog_df['반'] == s_ban]
            if s_reason != "전체": slog_df = slog_df[slog_df['변경사유'].str.contains(s_reason, na=False)]

            sk1, sk2, sk3 = st.columns(3)
            sk1.metric("전체 수정 건수", f"{len(slog_df)}건")
            sk2.metric("수정 관여 작업자", f"{slog_df['작업자'].nunique()}명")
            sk3.metric("수정된 날짜 수", f"{slog_df['날짜'].nunique()}일")

            st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
            th2 = st.columns([1.6, 1.0, 1.0, 1.2, 2.0, 2.0, 2.2, 1.8, 1.2])
            for col, txt in zip(th2, ["수정 시간","날짜","반","모델","이전 내용","변경 내용","변경 사유","상세","작업자"]):
                col.markdown(f"<p style='font-size:0.72rem;font-weight:700;color:#8a7f72;margin:0;padding-bottom:3px;border-bottom:1px solid #e0d8c8;'>{txt}</p>", unsafe_allow_html=True)
            # 성능: iterrows → to_dict('records')
            for row in slog_df.to_dict('records'):
                tr2 = st.columns([1.6, 1.0, 1.0, 1.2, 2.0, 2.0, 2.2, 1.8, 1.2])
                tr2[0].caption(str(row.get('시간',''))[:16])
                tr2[1].caption(str(row.get('날짜',''))[:10])
                tr2[2].write(row.get('반',''))
                tr2[3].write(row.get('모델명',''))
                tr2[4].caption(row.get('이전내용',''))
                tr2[5].caption(row.get('변경내용',''))
                reason_v2 = str(row.get('변경사유',''))
                rbg2 = REASON_COLOR.get(reason_v2, "#f5f2ec")
                tr2[6].markdown(f"<span style='background:{rbg2};padding:1px 6px;border-radius:4px;font-size:0.72rem;'>{reason_v2}</span>", unsafe_allow_html=True)
                tr2[7].caption(row.get('사유상세',''))
                tr2[8].caption(row.get('작업자',''))
        else:
            st.info("일정 수정 이력이 없습니다. 캘린더에서 일정 수정 시 자동 기록됩니다.")

    # ══════════════════════════════════════════════════════════════
    # [G] 생산 일정 관리
    # ══════════════════════════════════════════════════════════════
    st.divider()
    st.markdown("<div class='section-title'> 생산 일정 관리</div>", unsafe_allow_html=True)
    # 등록/삭제 결과 toast 표시
    _sch_toast = st.session_state.pop("_sch_add_toast", None)
    if _sch_toast:
        st.success(_sch_toast)

    sch_tab1, sch_tab2, sch_tab3 = st.tabs([" 직접 입력", " 엑셀 일괄 업로드", " 등록된 일정 관리"])

    with sch_tab2:
        st.markdown("<p style='color:#2a2420;'>생산계획 엑셀 파일을 업로드하면 일정에 자동 등록됩니다.</p>", unsafe_allow_html=True)

        # 양식 다운로드
        dl1, dl2 = st.columns([1, 2])
        with dl1:
            try:
                import openpyxl as _xl
                import io as _tmpio
                from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align, Border as _Border, Side as _Side
                from openpyxl.worksheet.datavalidation import DataValidation as _DV

                def _make_template():
                    _GROUP_COLORS = {
                        "제조1반": ("2471A3", "D4ECF7"),
                        "제조2반": ("1E8449", "D5F5E3"),
                        "제조3반": ("6C3483", "E8DAEF"),
                    }
                    _TAB_COLORS = {"제조1반":"2471A3","제조2반":"1E8449","제조3반":"6C3483"}
                    _EXAMPLES = {
                        "제조1반": ["2026-03-05","조립계획","TMP1115TI405","AM-1115 BLACK","32","3/15 32" ,"정상",""],
                        "제조2반": ["2026-03-05","조립계획","","모델명 예시","0","" ,"",""],
                        "제조3반": ["2026-03-05","포장계획","TMS9150008","T20i (i3-12세대) DUAL","20","3/20 20" ,"정상",""],
                    }

                    def _hf(color="FFFFFF", sz=10):
                        return _Font(name="맑은 고딕", bold=True, size=sz, color=color)
                    def _bf():
                        return _Font(name="맑은 고딕", size=10, color="2A2420")
                    def _fl(c): return _Fill("solid", fgColor=c)
                    def _bd(c="C8B89A"):
                        s = _Side(style="thin", color=c)
                        return _Border(left=s, right=s, top=s, bottom=s)
                    def _ca(): return _Align(horizontal="center", vertical="center", wrap_text=True)
                    def _la(): return _Align(horizontal="left",   vertical="center", wrap_text=True)

                    _wb = _xl.Workbook()
                    _wb.remove(_wb.active)

                    for _g in ["제조1반","제조2반","제조3반"]:
                        _hdr_col, _inp_bg = _GROUP_COLORS[_g]
                        _ws = _wb.create_sheet(_g)
                        _ws.sheet_properties.tabColor = _TAB_COLORS[_g]

                        # 1행 타이틀
                        _ws.merge_cells("A1:H1")
                        _ws["A1"].value = f"  {_g}  생산 일정 업로드 양식  |  시트명 = 반 이름 (자동 인식)"
                        _ws["A1"].font  = _Font(name="맑은 고딕", bold=True, size=12, color="FFFFFF")
                        _ws["A1"].fill  = _fl(_hdr_col)
                        _ws["A1"].alignment = _ca()
                        _ws.row_dimensions[1].height = 30

                        # 2행 안내
                        _ws.merge_cells("A2:H2")
                        _ws["A2"].value = "  날짜: YYYY-MM-DD  |  카테고리: 드롭다운 선택  |  처리수: 숫자만  |  5행부터 입력 (4행 예시는 자동 스킵)"
                        _ws["A2"].font  = _Font(name="맑은 고딕", size=9, color="2A2420")
                        _ws["A2"].fill  = _fl("FFF3CD")
                        _ws["A2"].alignment = _la()
                        _ws.row_dimensions[2].height = 18

                        # 3행 헤더 (반 컬럼 없음 - 시트명이 곧 반)
                        _headers = ["날짜 *", "카테고리 *", "P/N", "모델명 *", "처리수", "출하계획", "특이사항", "비고"]
                        for _ci, _h in enumerate(_headers, 1):
                            _c = _ws.cell(3, _ci)
                            _c.value = _h
                            _c.font  = _hf(color="FFFFFF")
                            _c.fill  = _fl(_hdr_col)
                            _c.alignment = _ca()
                            _c.border = _bd(_hdr_col)
                        _ws.row_dimensions[3].height = 26

                        # 4행 예시
                        for _ci, _v in enumerate(_EXAMPLES[_g], 1):
                            _c = _ws.cell(4, _ci)
                            _c.value = _v
                            _c.font  = _Font(name="맑은 고딕", size=9, color="8A7F72", italic=True)
                            _c.fill  = _fl("EEEBE4")
                            _c.alignment = _ca()
                            _c.border = _bd()
                        _ws.row_dimensions[4].height = 20

                        # 5~204행 입력 영역
                        for _r in range(5, 205):
                            for _c in range(1, 9):
                                _cell = _ws.cell(_r, _c)
                                _cell.fill      = _fl(_inp_bg if _c <= 7 else "FFFDF7")
                                _cell.border    = _bd()
                                _cell.alignment = _ca() if _c in [1,2,5] else _la()
                                _cell.font      = _bf()

                        # 카테고리 드롭다운 (직접 입력도 허용 - 유효성 검사 없음)
                        _dv = _DV(type="list", formula1='"조립계획,포장계획,출하계획"',
                                  showDropDown=False, showErrorMessage=False)
                        _dv.sqref = "B5:B204"
                        _ws.add_data_validation(_dv)

                        # 조립수 숫자 유효성
                        _dv2 = _DV(type="whole", operator="greaterThanOrEqual", formula1="0",
                                   showErrorMessage=True, errorTitle="입력 오류", error="0 이상의 숫자만 입력하세요.")
                        _dv2.sqref = "E5:E204"
                        _ws.add_data_validation(_dv2)

                        # 컬럼 너비
                        for _col, _w in zip("ABCDEFGH", [14,13,18,34,10,18,22,18]):
                            _ws.column_dimensions[_col].width = _w
                        _ws.freeze_panes = "A5"

                    # 가이드 시트
                    _wg = _wb.create_sheet(" 작성 가이드")
                    _wg.sheet_properties.tabColor = "8A7F72"
                    _guide = [
                        ["항목","설명"],
                        ["시트명","제조1반 / 제조2반 / 제조3반 → 시트명이 곧 반 (자동 인식)"],
                        ["날짜","YYYY-MM-DD 형식 (예: 2026-03-05)"],
                        ["카테고리","드롭다운: 조립계획 / 포장계획 / 출하계획 / 특이사항 / 기타"],
                        ["P/N","품목코드 (예: TMP6133002) — 선택"],
                        ["모델명","필수 — 조립 라인 모델 목록에 자동 등록됨"],
                        ["처리수","숫자만. 0 또는 빈칸이면 해당 행 스킵"],
                        ["출하계획","자유 텍스트 입력 — 예: 3/15 30 / 3월15일 30대 / 3/15 등 형식 무관, 선택 입력"],
                        ["특이사항","메모 자유 입력 — 선택"],
                        [],
                        [" 주의사항"],
                        ["1. 각 시트에 해당 반 데이터만 입력 (반 혼용 불가)"],
                        ["2. 4행 예시 행은 업로드 시 자동 스킵"],
                        ["3. 처리수 0 또는 빈칸 → 해당 행 스킵"],
                        ["4. 모델명 등록 시 해당 반 조립 라인에 자동 반영"],
                        ["5. 여러 반을 한 번에 → 각 시트 채워서 업로드"],
                    ]
                    for _ri, _row in enumerate(_guide, 1):
                        for _ci, _v in enumerate(_row, 1):
                            _c = _wg.cell(_ri, _ci, _v)
                            _c.font = _bf(); _c.alignment = _la()
                    for _ci in range(1, 3):
                        _c = _wg.cell(1, _ci)
                        _c.font = _hf(); _c.fill = _fl("7EB8E8")
                        _c.alignment = _ca(); _c.border = _bd()
                    _wg.cell(11, 1).font = _Font(name="맑은 고딕", bold=True, size=10, color="C8605A")
                    _wg.column_dimensions["A"].width = 14
                    _wg.column_dimensions["B"].width = 62

                    _buf = _tmpio.BytesIO()
                    _wb.save(_buf)
                    return _buf.getvalue()

                st.download_button(
                    " 업로드 양식 다운로드 (반별 시트)",
                    _make_template(),
                    "PMS_반별_업로드양식.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            except Exception as _e:
                st.warning(f"양식 생성 오류: {_e}")
        with dl2:
            st.markdown("""<p style='color:#5a96c8; font-size:0.88rem; margin:8px 0;'>
             <b>반별 시트 양식</b> (추천): 제조1반·2반·3반 시트 분리 — 시트명이 곧 반, 별도 선택 불필요<br>
            </p>""", unsafe_allow_html=True)

        # 지원 형식 안내
        with st.expander(" 지원 엑셀 형식 안내"):
            st.markdown("""
    <p style='color:#2a2420;'>
    <b>① PMS 반별 시트 양식</b> (위 버튼으로 다운로드) ⭐추천<br>
    &nbsp;&nbsp;• 시트명: <b>제조1반 / 제조2반 / 제조3반</b> — 시트명이 곧 반 정보<br>
    &nbsp;&nbsp;• 컬럼: 날짜 / 카테고리 / P/N / 모델명 / 처리수 / 출하계획 / 특이사항<br>
    &nbsp;&nbsp;• 여러 반을 한 파일에 각 시트별로 입력 후 한 번에 업로드 가능<br><br>
    <b>② PMS 단일 시트 양식</b><br>
    &nbsp;&nbsp;• 시트명: <b>생산계획_업로드</b> / 컬럼에 반 포함<br><br>
    """, unsafe_allow_html=True)

        uploaded_file = st.file_uploader(" 엑셀 파일 선택 (.xlsx)", type=["xlsx"], key="sch_upload")

        if uploaded_file:
            try:
                import openpyxl, io as _io, re as _re

                raw = uploaded_file.read()
                wb  = openpyxl.load_workbook(_io.BytesIO(raw), data_only=True)
                sheet_names = wb.sheetnames

                # ── 양식 자동 감지 ──
                group_sheets = [s for s in sheet_names if s in PRODUCTION_GROUPS]
                if group_sheets:
                    detected_mode = "PMS 반별 시트 양식"
                elif "생산계획_업로드" in sheet_names:
                    detected_mode = "PMS 단일 시트 양식"
                else:
                    detected_mode = "지원하지 않는 양식"

                st.info(f" 감지된 양식: **{detected_mode}**")

                parsed = []

                # ══════════════════════════════════════════
                # ① PMS 반별 시트 양식 파싱
                # ══════════════════════════════════════════
                if detected_mode == "PMS 반별 시트 양식":
                    st.markdown(
                        f"<p style='color:#2a2420;'>감지된 반 시트: "
                        + " ".join([f"<b style='color:#5a96c8;'>[{s}]</b>" for s in group_sheets])
                        + " — 전체 파싱합니다.</p>",
                        unsafe_allow_html=True)

                    for g_sheet in group_sheets:
                        ws = wb[g_sheet]
                        for row in ws.iter_rows(min_row=5, values_only=True):
                            date_val, cat, pn, model, qty, ship, note = (list(row) + [None]*7)[:7]
                            # 완전히 빈 행 스킵
                            if not any([date_val, cat, pn, model, qty, ship, note]):
                                continue
                            # 모델명 없으면 스킵
                            if not model:
                                continue
                            # 날짜 처리 (datetime / 문자열 / 숫자 모두 대응)
                            date_str = None
                            if isinstance(date_val, datetime):
                                date_str = date_val.strftime('%Y-%m-%d')
                            elif hasattr(date_val, 'strftime'):  # date 객체
                                date_str = date_val.strftime('%Y-%m-%d')
                            elif isinstance(date_val, str):
                                dv = date_val.strip()
                                # YYYY-MM-DD 또는 YYYY/MM/DD 형식
                                m = _re.match(r'(\d{4})[-/](\d{1,2})[-/](\d{1,2})', dv)
                                if m:
                                    date_str = f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
                            elif isinstance(date_val, (int, float)):
                                # 엑셀 시리얼 날짜 변환
                                try:
                                    date_str = (date(1899, 12, 30) + timedelta(days=int(date_val))).strftime('%Y-%m-%d')
                                except Exception:
                                    pass
                            if not date_str:
                                continue
                            # 수량 처리 (비어있어도 0으로 등록 허용)
                            qty_int = 0
                            if isinstance(qty, (int, float)):
                                qty_int = max(0, int(qty))
                            elif isinstance(qty, str) and qty.strip():
                                nums = _re.findall(r'\d+', qty)
                                qty_int = int(nums[0]) if nums else 0
                            parsed.append({
                                '날짜':     date_str,
                                '반':       g_sheet,
                                '카테고리': str(cat  or "조립계획").strip(),
                                'pn':       str(pn   or "").strip(),
                                '모델명':   str(model or "").strip(),
                                '조립수':   qty_int,
                                '출하계획': str(ship or "").strip(),
                                '특이사항': str(note or "").strip(),
                                '작성자':   st.session_state.user_id,
                            })

                # ══════════════════════════════════════════
                # ② PMS 단일 시트 양식 파싱
                # ══════════════════════════════════════════
                elif detected_mode == "PMS 단일 시트 양식":
                    ws = wb["생산계획_업로드"]
                    for row in ws.iter_rows(min_row=5, values_only=True):
                        ban, date_val, cat, pn, model, qty, ship, note = (list(row) + [None]*8)[:8]

                        if not ban and not model and not date_val: continue
                        if not model and not note: continue
                        if isinstance(date_val, datetime):
                            date_str = date_val.strftime('%Y-%m-%d')
                        elif isinstance(date_val, str) and len(date_val) == 10:
                            date_str = date_val
                        else:
                            continue
                        qty_int = 0
                        if isinstance(qty, (int, float)) and qty > 0:
                            qty_int = int(qty)
                        elif isinstance(qty, str):
                            nums = _re.findall(r'\d+', qty)
                            qty_int = int(nums[0]) if nums else 0
                        if qty_int <= 0: continue
                        parsed.append({
                            '날짜':     date_str,
                            '반':       str(ban or "").strip(),
                            '카테고리': str(cat or "기타").strip(),
                            'pn':       str(pn  or "").strip(),
                            '모델명':   str(model or "").strip(),
                            '조립수':   qty_int,
                            '출하계획': str(ship or "").strip(),
                            '특이사항': str(note or "").strip(),
                            '작성자':   st.session_state.user_id,
                        })

                # 지원하지 않는 양식
                else:
                    st.error(" 지원하지 않는 파일 형식입니다. 위 [업로드 양식 다운로드] 버튼으로 PMS 반별 시트 양식을 사용해주세요.")

                if parsed:
                    # 미리보기
                    import pandas as _pd
                    preview_df = _pd.DataFrame(parsed)[['날짜','카테고리','pn','모델명','조립수','출하계획']].rename(columns={'조립수': '처리수'})
                    st.markdown(f"<p style='color:#2a2420;'> <b>{len(parsed)}건</b> 파싱 완료 — 미리보기:</p>", unsafe_allow_html=True)
                    st.dataframe(preview_df, use_container_width=True, hide_index=True, height=300)

                    st.divider()

                    # ── 반 선택 (PMS 양식은 이미 반 포함, MNT는 여기서 지정) ──
                    has_ban = all(r.get('반','') in PRODUCTION_GROUPS for r in parsed)
                    if has_ban:
                        st.info(f" 반 정보가 파일에 포함되어 있습니다.")
                        upload_ban = None  # 파일 내 반 사용
                    else:
                        upload_ban = st.selectbox(
                            " 해당 엑셀의 반 선택 필수",
                            PRODUCTION_GROUPS,
                            key="bulk_ban_sel",
                        
                        )

                    # 날짜 범위 필터
                    all_dates = sorted(set(r['날짜'] for r in parsed))
                    fc1, fc2 = st.columns(2)
                    date_from = fc1.selectbox("등록 시작일", all_dates, key="bulk_from")
                    date_to   = fc2.selectbox("등록 종료일", all_dates,
                        index=len(all_dates)-1, key="bulk_to")
                    filtered = [r for r in parsed if date_from <= r['날짜'] <= date_to]

                    # 실제 등록될 반 표시
                    actual_ban = upload_ban if upload_ban else "파일 내 반 정보 사용"
                    st.markdown(
                        f"<p style='color:#5a96c8; font-weight:bold;'>"
                        f"→ 선택 범위 {date_from} ~ {date_to} : <b>{len(filtered)}건</b>"
                        f" | 등록 반: <b>{actual_ban}</b></p>",
                        unsafe_allow_html=True)

                    # 중복 처리 옵션
                    dup_mode = st.radio("기존 일정 처리",
                        ["건너뜀 (중복 날짜+모델 제외)", "모두 등록 (중복 허용)"],
                        horizontal=True, key="bulk_dup_mode")

                    col_reg, _ = st.columns([2,1])
                    if col_reg.button(f" {len(filtered)}건 일정 등록", type="primary", use_container_width=True, key="bulk_register"):
                        # 반 미선택 방어
                        if not has_ban and not upload_ban:
                            st.error("반을 선택해주세요.")
                        else:
                            existing = st.session_state.schedule_db
                            success_cnt = skip_cnt = fail_cnt = 0
                            fail_rows = []
                            
                            #  진행률 표시 추가
                            progress_bar = st.progress(0)
                            status_text = st.empty()
                            total = len(filtered)
                            
                            for idx, row in enumerate(filtered, 1):
                                # 진행률 업데이트
                                progress = idx / total
                                progress_bar.progress(progress)
                                status_text.text(f" 등록 중... {idx}/{total} ({int(progress*100)}%)")
                                
                                # 반 강제 지정
                                if upload_ban:
                                    row['반'] = upload_ban
                                # 반 최종 확인
                                if row.get('반','') not in PRODUCTION_GROUPS:
                                    skip_cnt += 1
                                    continue
                                # 중복 체크
                                if "건너뜀" in dup_mode and not existing.empty:
                                    dup = existing[
                                        (existing['날짜']   == row['날짜']) &
                                        (existing['모델명'] == row['모델명']) &
                                        (existing['카테고리'] == row['카테고리']) &
                                        (existing['반']     == row['반'])
                                    ]
                                    if not dup.empty:
                                        skip_cnt += 1
                                        continue
                                if insert_schedule(row):
                                    success_cnt += 1
                                else:
                                    fail_cnt += 1
                                    fail_rows.append(f"{row.get('날짜','')} / {row.get('모델명','')}")
                            
                            # 완료 표시
                            progress_bar.progress(1.0)
                            status_text.text(f" 등록 완료!")
                            
                            _clear_schedule_cache()
                            st.session_state.schedule_db = load_schedule()
                            if success_cnt > 0:
                                st.toast(f" 등록 완료: {success_cnt}건  |  건너뜀(중복): {skip_cnt}건" + (f"  |  실패: {fail_cnt}건" if fail_cnt else ""))
                            if fail_rows:
                                st.toast("등록 실패 행:\n" + "\n".join(fail_rows))
                            st.rerun()
                else:
                    st.warning("파싱된 일정이 없습니다. 파일 형식을 확인해주세요.")

            except Exception as e:
                st.error(f"파일 파싱 오류: {e}")

    with sch_tab1:
        # 반 선택을 폼 밖에 두어 모델/품목 목록이 즉시 반영되도록
        sch_ban = st.selectbox("반 *", PRODUCTION_GROUPS, key="sch_form_ban")
        _sch_models  = st.session_state.group_master_models.get(sch_ban, [])
        _sch_all_pns = list(dict.fromkeys(
            _pn for _m in _sch_models
            for _pn in st.session_state.group_master_items.get(sch_ban, {}).get(_m, [])
        ))

        with st.form("schedule_form"):
            sc1, sc2 = st.columns(2)
            sch_date = sc1.date_input("날짜")
            sch_cat  = sc2.selectbox("계획 유형 *", PLAN_CATEGORIES)

            sf1, sf2 = st.columns(2)
            # 모델명 — 등록 목록 드롭박스 + 직접 입력 병행
            _sm_opts  = [""] + _sch_models
            sch_model_sel = sf1.selectbox("모델명 (등록 목록)", _sm_opts,
                                          help="목록에서 선택하거나 아래에 직접 입력")
            sch_model_txt = sf1.text_input("모델명 직접 입력", placeholder="목록에 없으면 여기 입력")

            # P/N — 등록 목록 드롭박스 + 직접 입력 병행
            _spn_opts = [""] + _sch_all_pns
            sch_pn_sel = sf2.selectbox("P/N (등록 목록)", _spn_opts,
                                       help="목록에서 선택하거나 아래에 직접 입력")
            sch_pn_txt = sf2.text_input("P/N 직접 입력", placeholder="목록에 없으면 여기 입력")

            sc4, sc5 = st.columns(2)
            sch_qty_str = sc4.text_input("처리수", value="0", placeholder="숫자 입력")
            sch_note    = sc5.text_input("특이사항")

            if st.form_submit_button(" 일정 등록", use_container_width=True, type="primary"):
                try:
                    sch_qty = max(0, int(sch_qty_str.strip() or "0"))
                except ValueError:
                    sch_qty = 0
                # 직접 입력 우선, 없으면 드롭박스 선택값
                sch_model = sch_model_txt.strip() or sch_model_sel
                sch_pn    = sch_pn_txt.strip()    or sch_pn_sel
                if sch_model.strip() or sch_note.strip():
                    if insert_schedule({
                        '날짜': str(sch_date), '반': sch_ban,
                        '카테고리': sch_cat,
                        'pn': sch_pn.strip(), '모델명': sch_model.strip(),
                        '조립수': sch_qty, '출하계획': '',
                        '특이사항': sch_note.strip(), '작성자': st.session_state.user_id
                    }):
                        _clear_schedule_cache()
                        st.session_state.schedule_db = load_schedule()
                        st.session_state["_sch_add_toast"] = f" [{sch_ban}] {sch_date} 일정 등록 완료"
                        st.rerun()
                else:
                    st.warning("모델명 또는 특이사항을 입력해주세요.")

    with sch_tab3:
        sch_list = st.session_state.schedule_db
        if not sch_list.empty:
            # ── 전체 삭제 버튼 ──
            all_del_key = "sch_all_del_confirm"
            if not st.session_state.get(all_del_key, False):
                if st.button(" 전체 일정 삭제", type="secondary", key="sch_all_del"):
                    st.session_state[all_del_key] = True
                    st.rerun()
            else:
                st.error(" 등록된 일정 **전체**를 삭제합니다. 되돌릴 수 없습니다.")
                ac1, ac2, ac3 = st.columns([2, 1, 1])
                ac1.markdown("<p style='color:#c8605a; font-weight:bold; margin-top:8px;'>삭제 후 복구 불가</p>", unsafe_allow_html=True)
                if ac2.button(" 예, 전체 삭제", type="primary", use_container_width=True, key="sch_all_del_yes"):
                    # 성능: iterrows 대신 ID 리스트로 직접 처리
                    for _sid in sch_list['id'].dropna().astype(int).tolist():
                        delete_schedule(_sid)
                    _clear_schedule_cache()                        # ← 캐시 초기화
                    st.session_state.schedule_db = load_schedule()
                    st.session_state[all_del_key] = False
                    st.rerun()
                if ac3.button("취소", use_container_width=True, key="sch_all_del_no"):
                    st.session_state[all_del_key] = False
                    st.rerun()
            st.divider()

            # ── 헤더 행 ──
            hh = st.columns([1.0, 0.8, 1.2, 1.5, 2.2, 0.8, 1.8, 0.5])
            for col, txt in zip(hh, ["유형","반","날짜","P/N","모델명","수량","특이사항",""]):
                col.markdown(f"<p style='font-size:0.72rem;font-weight:700;color:#8a7f72;margin:0;padding-bottom:3px;border-bottom:1px solid #e0d8c8;'>{txt}</p>", unsafe_allow_html=True)

            # 성능: iterrows → to_dict('records') (위젯 키는 row['id'] 사용)
            for row in sch_list.sort_values('날짜').to_dict('records'):
                cat    = row.get('카테고리', '기타')
                color  = SCHEDULE_COLORS.get(cat, "#888")
                row_id = row['id']
                del_ck = f"sch_del_ck_{row_id}"

                # ── 데이터 행 ──
                c1,c2,c3,c4,c5,c6,c7,c8 = st.columns([1.0, 0.8, 1.2, 1.5, 2.2, 0.8, 1.8, 0.5])
                c1.markdown(f"<span style='background:{color}22;border-left:3px solid {color};padding:2px 5px;border-radius:4px;font-size:0.78rem;'>{cat}</span>", unsafe_allow_html=True)
                c2.caption(row.get('반',''))
                c3.caption(row.get('날짜',''))
                c4.caption(row.get('pn',''))
                c5.caption(row.get('모델명',''))
                c6.caption(f"{row.get('조립수',0)}대")
                c7.caption(row.get('특이사항',''))
                if c8.button("삭제", key=f"del_sch_{row_id}", help="삭제"):
                    st.session_state[del_ck] = True
                    st.rerun()

                # ── 확인 팝업: 행 아래에 별도 표시 ──
                if st.session_state.get(del_ck, False):
                    with st.container():
                        cf1, cf2, cf3 = st.columns([3, 1, 1])
                        cf1.warning(f"**[{row.get('날짜','')} / {row.get('모델명','')}]** 일정을 삭제하시겠습니까?")
                        if cf2.button(" 삭제", key=f"del_sch_yes_{row_id}", type="primary", use_container_width=True):
                            delete_schedule(int(row_id))
                            _clear_schedule_cache()                # ← 캐시 초기화
                            st.session_state.schedule_db = load_schedule()
                            st.session_state[del_ck] = False
                            st.rerun()
                        if cf3.button("취소", key=f"del_sch_no_{row_id}", use_container_width=True):
                            st.session_state[del_ck] = False
                            st.rerun()
        else:
            st.info("등록된 일정이 없습니다.")

    st.divider()




